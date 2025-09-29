
# quickstart.py

import os
import sys
import yaml
from datetime import datetime
from urllib.parse import unquote
from openpyxl.utils import get_column_letter

if len(sys.argv) < 3:
    print("Usage: python quickstart.py <yaml_file> <timestamp")
    sys.exit(1)

yaml_file = sys.argv[1]
timestamp = sys.argv[2]

with open(yaml_file, 'r') as f:
    data = yaml.safe_load(f)

fileinfo = data.get('fileinfo', {})
if not fileinfo:
    print("No fileinfo found in the YAML file.")
    sys.exit(1)
basename = fileinfo.get('basename')
tablename = fileinfo.get('table').replace(" ", "_") if fileinfo.get('table') else ""
source = fileinfo.get('source')
scope_file = fileinfo.get('scope file')

if not basename:
    print("No 'basename'found in fileinfo. Expecting 'basename' key.")
    sys.exit(1)

if not tablename:
    print("No 'table' found in fileinfo. Expecting 'table' key.")
    sys.exit(1)


# I don't think this code is doing anything useful. It always sets create_mode to True?! since 
# Determine if we will be INSERTING rows eventually vs just UPDATING existing rows in Excel/SharePoint
if "quickstart.scope" in yaml_file.lower():
    print(f"quickstart detected based on filename: {yaml_file}.")
    # You can set a flag or handle import-specific logic here if needed
    output_file = basename + "." + tablename + timestamp + ".quickstart.jira.csv"
    #mode = "assignee" #"resolved"
else:
    print("Error: YAML filename is not 'quickstart.scope.yaml")
    sys.exit(1)

fields = data.get('fields', [])
# Convert list of dicts into a dictionary
fields_dict = {field["value"]: field.get("index", "<blank>") for field in fields}

field_values = [field.get('value') for field in fields if 'value' in field]
field_indexes = [field.get('index') for field in fields if 'index' in field]
field_values_str = ','.join(field_values)
field_indexes_str = ','.join(map(str, field_indexes))

print("Source file,", source)
print("basename,", basename)
print("Table,", tablename)
print("Field indexes,", field_indexes_str)
print("Field values,", field_values_str)

'''with open(output_file, "w") as outfile:
    outfile.write("Source file," + source + "\n")
    outfile.write("Basename," + basename + "\n")
    outfile.write("Scope file," + scope_file + "\n")    
    outfile.write("Table," + tablename + "\n")
    outfile.write("Field indexes," + field_indexes_str + "\n")
    outfile.write("Field values," + field_values_str + "\n")
'''

# yaml will expected to contain params, eg:
'''
    fileinfo:
    basename: Quickstart-test.xlsx
    scope file: Quickstart-test.xlsx.My_Quick_Start.20250927_224227.quickstart.scope.yaml
    source: Quickstart-test.xlsx
    table: My_Quick_Start
    jira projects:
    - tes
    - fr
    row: 7
    col: 2
    lastrow: 42
'''

# the following were addded by scope.py in runrate yaml file
jira_projects_list = data.get('jira projects',[])
print(f"Jira projects list from yaml: {jira_projects_list}")
quickstart_table_row = data.get('row', None)
print(f"quickstart_table_row from yaml: {quickstart_table_row}")
quickstart_table_col = data.get('col', None)
print(f"quickstart_table_col from yaml: {quickstart_table_col}")
last_excel_row = data.get('lastrow', None)
print(f"last_excel_row from yaml: {last_excel_row}")

# validate
if not jira_projects_list or quickstart_table_row is None or quickstart_table_col is None or last_excel_row is None:
    print("❌ Error: Missing required fields in YAML file")
    sys.exit(1)

'''
The output will be:
    Project Epics<jira> jql project in ({jira_projects_list}) and issuetype = Epic
    <key>	<url>	<children>	<summary>	<description>	<status>	<timestamp>	<children>	<links>	<assignee>	<reporter>	Commentsummary<ai>
'''


# join list into comma-separated string
jira_projects_str = ",".join(jira_projects_list)


# Multiline definition (for readability)
changes_epics = [
    f"Project Epics<jira> jql project in ({jira_projects_str}) and issuetype = Epic",
    "Jira ID <key>",
    "Jink Link <url>",
    "Title <summary>",
    "Description <description>",
    "Status <status>",
    "Updated by AI <timestamp>",
    "Epic Children <children>",
    "Jira Linked <links>",
    "Assignee <assignee>",
    "Reporter <reporter>",
    "Comments Summary <ai>"
]

changes_resolved_velocity = [
    f"Resolved Velocity<rate resolved><weeks>jql project in ({jira_projects_str}) and created >= -30d"
]

changes_assignee_velocity = [
    f"Assignee Velocity<rate assignee><weeks>jql project in ({jira_projects_str}) and created >= -30d"
]

changes_overall_status = [
    f"Project Summary<ai brief> Project Epics"
]

changes_cycletime = [
    f"Cycle Time<cycletime> jql project in ({jira_projects_str}) and updated >= -90d"
]

changes_list = []
coord = get_column_letter(1)
r = last_excel_row

excel_col = quickstart_table_col + 1

changes_list.append(f"{get_column_letter(excel_col)}{r} = {changes_overall_status[0]} || ")
r += 5
changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = {changes_resolved_velocity[0]} || ")
r += 10
changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = {changes_assignee_velocity[0]} || ")
r += 10

changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = {changes_cycletime[0]} || ")
r += 10

# column write always start at column A
for idx, entry in enumerate(changes_epics):
    if idx == 0:
        changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = {entry} || ")
    else:
        coord = f"{get_column_letter(excel_col + 1 + idx)}{r + 1}"
        changes_list.append(f"{coord} = {entry} || ")

    

changes_file = yaml_file.replace("scope.yaml","changes.txt")
print(f"Writing changes to {changes_file}")

if changes_list:
    with open(changes_file, "w") as f:
        for entry in changes_list:
            if "||None" in entry:
                entry = entry.replace("||None", "||")
            f.write(entry + "\n")
            print(entry)
    print(f"Changes written to {changes_file} with ({len(changes_list)} entries).")
else:
    print(f"No changes to write. Not need to create {changes_file}")

