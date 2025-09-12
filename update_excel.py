import sys
from openpyxl.styles import Alignment

import re

#def is_valid_jira_id(jira_id):
#    return re.match(r'^[A-Z][A-Z0-9]+-\d+$', jira_id) is not None or "JQL" not in jira_id

def is_JQL(jira_id):
    return "jql" in jira_id.lower()

def is_valid_jira_id(jira_id: str) -> bool:
    # Case 1: valid Jira issue key (case-insensitive)
    if re.match(r'^[A-Z][A-Z0-9]+-\d+$', jira_id, re.IGNORECASE):
        return True
    
    # Case 2: JQL query (case-insensitive)
    if jira_id.strip().lower().startswith("jql "):
        return True
    
    return False

field_index_map = {}
jira_data = {}
change_list = []  # ← This will hold "A1=newvalue"-style strings
file_info = {}

def load_jira_file(filename):
    print(f"Loading JIRA data from {filename}")
    with open(filename, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if len(lines) < 6:
        raise ValueError("File must contain at least 6 header lines.")

    file_info['source'] = lines[0].strip().replace("Source file,", "").strip()
    file_info['basename'] = lines[1].strip().replace("Basename,", "").strip()
    file_info['scope file'] = lines[2].strip().replace("Scope file,", "").strip()
    file_info['table'] = lines[3].strip().replace("Table,", "").strip()

    print(f"lines[4]: {lines[4].strip()}")
    print(f"lines[5]: {lines[5].strip()}")

    # Parse field indexes and values from the first two lines
    index_line = lines[4].strip().replace("Field indexes,", "").strip()
    value_line = lines[5].strip().replace("Field values,", "").strip()

    #print(f"Index line: {index_line}")
    #print(f"Value line: {value_line}")

    index_values = [int(i.strip()) for i in index_line.split(',')]
    field_names = [v.strip() for v in value_line.split(',')]

    if len(index_values) != len(field_names):
        raise ValueError("Mismatch between number of field indexes and field names.")

    # Create a mapping of field names to their respective indexes.
    field_index_map = dict(zip(field_names, index_values))

    # Now that we have the field names and indexes, we can parse the data rows.
    # Parse the remaining lines (ie. the data rows)
    for line in lines[6:]:
        line = line.strip()
        if not line:
            continue  # skip blank lines
        parts = [p.strip() for p in line.split('|')]
        print(f"Processing line: {line}")

        if len(parts) != len(field_names):
            print(f"Warning: Mismatched field count in {filename}. Make sure number of values ({len(parts)}) provided matches number of fields ({len(field_names)}).")
            continue

        # dictionary that maps field names to their corresponding values that we just read from the line
        # assumes the order of parts matches the order of field_names (in the jira.csv file)
        record = dict(zip(field_names, parts))
        key = record.get("key")
        if not key:
            print("Warning: No <key> field found in csv line. Check that the excel file table has a <key> column")
            continue
        
        print(f"Record for {key}: {record}")
        jira_data[key] = record # record contains all fields values for this jira ID
        print(f"Added record for {key} to jira_data.")

    return field_index_map, jira_data


from openpyxl import load_workbook

from openpyxl.utils import get_column_letter
def convert_row_col_to_excel_coordinate(row, col):
    col_letter = get_column_letter(col)  # 1 → A, 2 → B, 27 → AA, etc.
    print(f"converted row={row} col={col} to sheet coordinate {col_letter}{row}")
    return f"{col_letter}{row}"



def process_jira_table_blocks(filename):
    wb = load_workbook(filename)
    ws = wb.active

    printing = False  # Flag to track when we're in a Jira Table block
    printing_import_mode = False # flag to track when we're in a jira table block in import mode

    last_known_row_num = 0

    for row in ws.iter_rows():
        #print(f"Processing row {row[0].row}: {[cell.value for cell in row]}")

        # skip just 1 more row because it contains the table header. I know it's a hack but it works for now
        #if printing and import_mode: 
        #    print("Skipping header row in import mode")
        #    continue

        last_known_row_num = row[0].row  # save this incase we run out of rows but need to keep updating cells to finish inserting all the jira_id in the sheet

        for cell in row:
            #print(f"Processing cell {cell.coordinate}: {cell.value}")
            cleaned_value = str(cell.value).strip().replace(" ", "_")
            #print(f"Cleaned cell value: {cleaned_value}" + f" | file_info['table']: {file_info['table']}")
            #if file_info["table"] in str(cell).strip().replace(" ", "_"):
            if file_info["table"] in cleaned_value and "<jira>" in cleaned_value:
                print(f"Found table header '{file_info['table']}' in cell {cell.coordinate}")
                printing = True
                break
            
        #print(f"Printing flag is {'ON' if printing else 'OFF'} for row {row[0].row}")
        #print(f"Import mode is {'ON' if import_mode else 'OFF'}")

        if printing and execsummary_mode:
            print("This is ExecSummary table")

        elif printing and not import_mode:
            print("About to process update row.")
            first_cell = row[field_index_map["key"]].value
   
            if first_cell is not None:
                first_cell = str(first_cell).lower()
            else:
                first_cell = ""

            #print(first_cell)
            if (jira_data and first_cell in jira_data):
                print(f"Updating row {row[field_index_map['key']].row} with data for {first_cell}")
                #record = jira_data[first_cell]
                record = jira_data.pop(first_cell)
                for field, index in field_index_map.items():
                    print(f"Checking field {field} at index {index}")
                    if field in record:
                        print(f"Updating field {field} at index {index} with value {record[field]}")
                        cell_value = record[field]
                        print(f"Cell value for {field}: {cell_value}")
                        if cell_value is not None:
                            #print(f"Setting cell value: {cell_value}")
                            #ws.cell(row=row[0].row, column=index + 1, value=cell_value)
                            target_cell = ws.cell(row=row[0].row, column=index + 1)
                            old_value = target_cell.value
                            old_value = str(old_value).replace("\n", ";") if old_value else None  # Replace newlines with semicolons for comparison
                            print(f"Old value for {target_cell.coordinate}: {old_value}")
                            print(f"Setting cell {target_cell.coordinate} = {cell_value}")  # <-- Coordinate logging
                            target_cell.value = cell_value.replace(";", "\n")  # Replace ; with newline
                                                    
                            # Enable text wrapping to show newlines
                            '''target_cell.alignment = Alignment(wrapText=True)

                            if field == "key" and is_valid_jira_id(cell_value):
                                target_cell.hyperlink = "https://fz96tw.atlassian.net/browse/" + cell_value  # The actual link
                                target_cell.style = "Hyperlink"  # Optional: makes it blue and underlined
                            elif field == "key" and is_JQL(cell_value):
                                target_cell.hyperlink = "https://fz96tw.atlassian.net/issues/?jql=" + cell_value.lower().replace("jql", "")  # The actual link
                                target_cell.style = "Hyperlink"
                            '''
                            # Save coordinate + value to list
                            change_list.append(f"{target_cell.coordinate}={cell_value.replace("\n",";")}||{old_value}")
            else:
                print(f"No data found for {first_cell} in jira_data.")
        
        elif printing and not printing_import_mode:
                print("Hunting for starting cell for import mode")
                first_cell = row[field_index_map["key"]].value
                if first_cell is not None:
                    first_cell = str(first_cell).lower()
                else:
                    first_cell = ""

                if "<key>" in str(first_cell).lower():
                    print("Found <key> header row, let's skip to next row")
                    printing_import_mode = True
                    continue
                else:
                    print("need to keep looking...")
                    
            
        elif printing and import_mode and printing_import_mode:
            print(f"About to process import row - checking column {field_index_map['key']}")
            first_cell = row[field_index_map["key"]].value
            
            # do not case normalize in this case because the sheet cell and jira.csv entries (jira_data) are same case already
            #if first_cell is not None:
            #    first_cell = str(first_cell).lower()
            #else:
            #    first_cell = ""
            print(f"First cell in import row: {first_cell}")

            if jira_data and (first_cell is None or first_cell == ""):
                print("First cell is blank")
                print(f"Import mode: Adding new row for {first_cell}")
   
                print(f"Updating row {row[field_index_map["key"]].row} with data for {first_cell}")
                k, record = jira_data.popitem()
                for field, index in field_index_map.items():
                    print(f"Checking field {field} at index {index}")
                    if field in record:
                        print(f"Updating field {field} at index {index} with value {record[field]}")
                        cell_value = record[field]
                        print(f"Cell value for {field}: {cell_value}")
                        if cell_value is not None:
                            target_cell = ws.cell(row=row[0].row, column=index + 1)
                            old_value = target_cell.value
                            old_value = str(old_value).replace("\n", ";") if old_value else None  # Replace newlines with semicolons for comparison
                            print(f"Old value for {target_cell.coordinate}: {old_value}")
                            print(f"Setting cell {target_cell.coordinate} = {cell_value}")  # <-- Coordinate logging
                            target_cell.value = cell_value.replace(";", "\n")  # Replace ; with newline
                                                    
                            # Enable text wrapping to show newlines
                            target_cell.alignment = Alignment(wrapText=True)

                            if field == "key" and is_valid_jira_id(cell_value):
                                target_cell.hyperlink = "https://fz96tw.atlassian.net/browse/" + cell_value  # The actual link
                                target_cell.style = "Hyperlink"  # Optional: makes it blue and underlined
                            elif field == "key" and is_JQL(cell_value):
                                target_cell.hyperlink = "https://fz96tw.atlassian.net/issues/?jql=" + cell_value.lower().replace("jql", "")  # The actual link
                                target_cell.style = "Hyperlink"
                            
                            # Save coordinate + value to list
                            change_list.append(f"{target_cell.coordinate}={cell_value.replace("\n",";")}||{old_value}")
            
            elif not jira_data and first_cell is not None:
                if first_cell is not None: #and is_valid_jira_id(first_cell)
                    print(f"No more items in jira_data but first_cell '{first_cell} has text so let's keep going down until blank cell")
                    # cannot exit loop yet because there may be rows below current row that
                    # were populated previously because they were in the jql result.
                    # And now they're not in current result set and must be STRIKEOUT
                    #break

                    # Mark the cell for strikeout if isn't in jira_ids for this JQL result
                    print(f"About to STRIKEOUT {first_cell} not in jira_data, marking for strikeout.")
                    index = field_index_map["key"]  # get the index of the "key" field to find the blank cell
                    print("target index for 'key': ", index)

                    target_cell = ws.cell(row=row[0].row, column=index + 1)
                    old_value = target_cell.value
                    old_value = str(old_value).replace("\n", ";") if old_value else None  # Replace newlines with semicolons for comparison
                    print(f"Old value for {target_cell.coordinate}: {old_value}")
                    print(f"Setting cell {target_cell.coordinate} = STRIKEOUT {first_cell}")  # <-- Coordinate logging
                    
                    if "STRIKEOUT" in target_cell.value:
                        print(f"target_cell '{target_cell.value}' already has STRIKEOUT prefix. Do nothing")
                    else:
                        target_cell.value = "STRIKEOUT " + first_cell  # add marker to indicate it needs strick out 
                        print(f"STRIKEOUT added to target_cell '{target_cell}'")
                        target_cell.font = target_cell.font.copy(strike=True)  # Apply strikeout
                        target_cell.alignment = Alignment(wrapText=True)                 
                        change_list.append(f"{target_cell.coordinate}={target_cell.value.replace("\n",";")}||{old_value}")
                else:
                    # now we can exit loop since not more jira_data and no more keys in key column              
                    break


            #elif is_valid_jira_id(first_cell):
            elif jira_data and first_cell is not None:
                #print(f"This row already contains valid jira id {first_cell}. Will just update instead of overwriting.")
                print("This rows contains some text already")
                # if the text in cell matches the Jira id we are looking for then
                # pop the first_cell from jira_ids if it exists 
                # and update all the field cells applicable in this row

                # if previously had strikeout then remove it
                if "STRIKEOUT" in first_cell:
                    print(f"Already contains STRIKEOUT removing it from first_cell '{first_cell}'")
                    first_cell = first_cell.replace("STRIKEOUT ","")
                    print(f"STRIKEOUT removed, first_cell is no '{first_cell}'")

                # was this jira previously in the cell? 
                # if yes then update all the fields columns and remove strikeout
                # we want to preserve this row incawe there were user notes in this row as well
                if first_cell is not None:
                    first_cell = str(first_cell).lower()
                else:
                    first_cell = ""

                if first_cell in jira_data:
                    print("This first_cell is still part of the jira_data result set")
                    print(f"Popping {first_cell} from jira_data for update.")
                    record = jira_data.pop(first_cell)
                    print(f"Remaining jira_data items: {len(jira_data)}")
                    
                    # if there are Jira IDs in the table already it means we need to update them instead of insert
                    # Call update process here. 
                    print(f"Updating row {row[field_index_map["key"]].row} with data for {first_cell}")
                    #k, record = jira_data.popitem()
                    for field, index in field_index_map.items():
                        print(f"Checking field {field} at index {index}")
                        if field in record:
                            print(f"Updating field {field} at index {index} with value {record[field]}")
                            cell_value = record[field]
                            print(f"Cell value for {field}: {cell_value}")
                            if cell_value is not None:
                                target_cell = ws.cell(row=row[0].row, column=index + 1)
                                old_value = target_cell.value
                                old_value = str(old_value).replace("\n", ";") if old_value else None  # Replace newlines with semicolons for comparison
                                print(f"Old value for {target_cell.coordinate}: {old_value}")
                                print(f"Setting cell {target_cell.coordinate} = {cell_value}")  # <-- Coordinate logging
                                target_cell.value = cell_value.replace(";", "\n")  # Replace ; with newline
                                                        
                                # Enable text wrapping to show newlines
                                target_cell.alignment = Alignment(wrapText=True)

                                if field == "key" and is_valid_jira_id(cell_value):
                                    target_cell.hyperlink = "https://fz96tw.atlassian.net/browse/" + cell_value  # The actual link
                                    target_cell.style = "Hyperlink"  # Optional: makes it blue and underlined
                                elif field == "key" and is_JQL(cell_value):
                                    target_cell.hyperlink = "https://fz96tw.atlassian.net/issues/?jql=" + cell_value.lower().replace("jql", "")  # The actual link
                                    target_cell.style = "Hyperlink"
                                
                                # Save coordinate + value to list
                                change_list.append(f"{target_cell.coordinate}={cell_value.replace("\n",";")}||{old_value}")

                else:
                    # Mark the cell for strikeout if isn't in jira_ids for this JQL result
                    print(f"About to STRIKEOUT {first_cell} not in jira_data, marking for strikeout.")
                    index = field_index_map["key"]  # get the index of the "key" field to find the blank cell
                    print("target index for 'key': ", index)

                    target_cell = ws.cell(row=row[0].row, column=index + 1)
                    old_value = target_cell.value
                    old_value = str(old_value).replace("\n", ";") if old_value else None  # Replace newlines with semicolons for comparison
                    print(f"Old value for {target_cell.coordinate}: {old_value}")
                    print(f"Setting cell {target_cell.coordinate} = STRIKEOUT {first_cell}")  # <-- Coordinate logging
                    
                    if "STRIKEOUT" in target_cell.value:
                        print(f"target_cell '{target_cell.value}' already has STRIKEOUT prefix. Do nothing")
                    else:
                        target_cell.value = "STRIKEOUT " + first_cell  # add marker to indicate it needs strick out 
                        print(f"STRIKEOUT added to target_cell '{target_cell}'")
                        target_cell.font = target_cell.font.copy(strike=True)  # Apply strikeout
                        target_cell.alignment = Alignment(wrapText=True)                 
                        change_list.append(f"{target_cell.coordinate}={target_cell.value.replace("\n",";")}||{old_value}")
            elif not jira_data and first_cell is None:
                # reached row with first_cell is None. Will Exit Loop"
                print(f"Reached row with None first_cell = {first_cell} so will break out of row scan loop")
                break

    # if there are jira_id still remaining but we ran out of rows then keep going if we were in IMPORT mode
    if printing and import_mode and printing_import_mode:
        print(f"Row count ended row = {last_known_row_num} but will continue with import_mode inserts into sheet")
        #first_cell = row[field_index_map["key"]].value
        #print(f"First cell in import row: {first_cell}")
    
        #while jira_data and (first_cell is None or first_cell == ""):
        while jira_data:
            last_known_row_num += 1
            print(f"Updating row {row[field_index_map["key"]].row}")
            k, record = jira_data.popitem()
            for field, index in field_index_map.items():
                print(f"Checking field {field} at index {index}")
                if field in record:
                    print(f"Updating field {field} at index {index} with value {record[field]}")
                    cell_value = record[field]
                    print(f"Cell value for {field}: {cell_value}")
                    if cell_value is not None:
                        #target_cell = ws.cell(row=row[0].row, column=index + 1)
                        target_cell_coordinate = convert_row_col_to_excel_coordinate(last_known_row_num, index+1)
                        old_value = ""
                        #old_value = target_cell.value
                        #old_value = str(old_value).replace("\n", ";") if old_value else None  # Replace newlines with semicolons for comparison
                        print(f"Old value for {target_cell_coordinate}: {old_value}")
                        print(f"Setting cell {target_cell_coordinate} = {cell_value}")  # <-- Coordinate logging
                        #target_cell.value = cell_value.replace(";", "\n")  # Replace ; with newline
                                                
                        # Enable text wrapping to show newlines
                        #target_cell.alignment = Alignment(wrapText=True)
                        
                        # Save coordinate + value to list
                        change_list.append(f"{target_cell_coordinate}={cell_value.replace("\n",";")}||{old_value}")



    # Save updates to the same file
    #print(f"Saving updates to {filename}")
    #wb.save(filename)                


if __name__ == "__main__":
    #main()
    if len(sys.argv) < 3:
        print("Usage: python update_excel.py <csv file> <xlsx file>")
        sys.exit(1)
    
    jiracsv = sys.argv[1]
    print(f"Received argument: {jiracsv}")
    xlfile = sys.argv[2]
    print(f"Received argument: {xlfile}")

    if "import" in jiracsv.lower():
        print("Import mode detected based on filename containing 'import'.")
        # You can set a flag or handle import-specific logic here if needed
        import_mode = True
    else:
        import_mode = False

    if "aisummary" in jiracsv.lower():
        print("This file is ExecSummary")
        execsummary_mode = True
    else:
        execsummary_mode = False

    field_index_map, jira_data_in = load_jira_file(jiracsv)
    print(f"Loaded {len(jira_data)} records from {jiracsv}")
    
    for key, index in field_index_map.items():
        print(f"{key}: {index}")

    jira_data = {k.lower(): v for k, v in jira_data_in.items()}

    for key, record in jira_data.items():
        print(f"lowercased {key}: {record}")

    if (execsummary_mode):
        print("Call Execsummary processor now")
        #process_execsummary_table_blocks(xlfile)
        #process_jira_table_blocks(xlfile)
    else:
        process_jira_table_blocks(xlfile)
    
    print("Finished processing Jira Table blocks.")

            
    if not change_list:
        print("No changes made.")
        # delete any eisting  changes.txt file from previous runs so sharepoint_udpate.py doesn't get confused
        
    else:
        #changes_file = xlfile.replace(".xlsx", ".changes.txt")

        if import_mode:
            changes_file = jiracsv.replace(".import.jira.csv", ".import.changes.txt")
        else:
            changes_file = jiracsv.replace(".jira.csv", ".changes.txt")
        
        print(f"Writing changes to {changes_file}")
        with open(changes_file, "w") as f:
            for entry in change_list:
                if "||None" in entry:
                    entry = entry.replace("||None", "||")
                f.write(entry + "\n")
                print(entry)
        print(f"Changes written to {changes_file} with ({len(change_list)} entries).")

