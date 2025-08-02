import sys
import os
import yaml
from openpyxl import load_workbook
 
def read_excel_cells(filename):
    wb = load_workbook(filename)
    ws = wb.active
    return ws  # return worksheet object for cell access

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scope.py <filename>")
        sys.exit(1)

    filename = sys.argv[1]
    ws = read_excel_cells(filename)

    jira_ids = []
    jira_fields = []
    fields_found = False
    output_file = ""

    for row in ws.iter_rows():
        for cell in row:
            if "<JIRA>" in str(cell.value):
                value = str(cell.value).replace("<JIRA>", "").strip()
                jira_fields.append({"value": value, "coordinate": cell.coordinate})

        if jira_fields and not fields_found:
            fields_found = True
            base = os.path.splitext(os.path.basename(filename))[0]
            output_file = f"{base}.scope.yaml"
            with open(output_file, 'w') as f:
                yaml.dump({"fields": jira_fields}, f, default_flow_style=False)
            continue

        if fields_found:
            # Lookup the coordinate for the field "key"
            coord_of_key = next((f["coordinate"] for f in jira_fields if f["value"] == "key"), None)
            if coord_of_key:
                col_letter = ''.join(filter(str.isalpha, coord_of_key))
                row_number = row[0].row  # all cells in the row have same row number
                cell_value = ws[f"{col_letter}{row_number}"].value
                jira_ids.append(cell_value)

    if jira_ids:
        print(f"JIRA IDs found: {jira_ids}")
        with open(output_file, 'a') as f:
            yaml.dump({"jira_ids": jira_ids}, f, default_flow_style=False)

    print("JIRA rows have been written to output file:", output_file)
    print("Total rows processed:", ws.max_row)
