from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any, Tuple
import statistics

from bs4 import BeautifulSoup
import json
import glob
import os
import requests

def calculate_average_status_transition_time(jira_issues: List[Any]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """
    Same as before, but now includes an artificial "Created" status to represent
    the time from issue creation to its first status transition.
    """

    transition_durations = defaultdict(list)
    transition_issues = defaultdict(list)

    print(f"Calculating status transition times for {len(jira_issues)} issues...")

    for issue in jira_issues:
        # --- Extract issue info and changelog ---
        if hasattr(issue, 'key'):
            issue_key = issue.key
            changelog = getattr(issue, 'changelog', None)
            creation_time = None
            if hasattr(issue, 'fields') and hasattr(issue.fields, 'created'):
                try:
                    creation_time = parse_jira_timestamp(issue.fields.created)
                except ValueError:
                    pass
        else:
            issue_key = issue.get('key', 'Unknown')
            changelog = issue.get('changelog')
            creation_time = None
            fields = issue.get('fields', {})
            if 'created' in fields:
                try:
                    creation_time = parse_jira_timestamp(fields['created'])
                except ValueError:
                    pass

        if not changelog:
            continue

        # --- Extract histories ---
        histories = getattr(changelog, 'histories', None)
        if isinstance(changelog, dict):
            histories = changelog.get('histories', histories)
        if not histories:
            continue

        # --- Extract status changes ---
        status_changes = []
        for history in sorted(histories, key=lambda h: getattr(h, 'created', '') if hasattr(h, 'created') else h.get('created', '')):
            created_str = getattr(history, 'created', None) or (history.get('created') if isinstance(history, dict) else None)
            if not created_str:
                continue

            try:
                timestamp = parse_jira_timestamp(created_str)
            except ValueError:
                continue

            items = getattr(history, 'items', None) or (history.get('items') if isinstance(history, dict) else [])
            for item in items:
                field = getattr(item, 'field', None) if hasattr(item, 'field') else item.get('field', None)
                if field != 'status':
                    continue
                from_status = getattr(item, 'fromString', None) if hasattr(item, 'fromString') else item.get('fromString', None)
                to_status = getattr(item, 'toString', None) if hasattr(item, 'toString') else item.get('toString', None)
                if from_status and to_status:
                    status_changes.append({'timestamp': timestamp, 'from_status': from_status, 'to_status': to_status})

        # --- Add artificial "Created" → first transition ---
        if creation_time and status_changes:
            first_real_status = status_changes[0]['from_status']
            duration_hours = (status_changes[0]['timestamp'] - creation_time).total_seconds() / 3600.0
            if duration_hours >= 0:
                transition_key = ('Created', first_real_status)
                transition_durations[transition_key].append(duration_hours)
                transition_issues[transition_key].append({
                    'issue_key': issue_key,
                    'duration_hours': round(duration_hours, 6),
                    'start_time': creation_time,
                    'end_time': status_changes[0]['timestamp']
                })

        # --- Standard per-transition durations ---
        for i, change in enumerate(status_changes):
            from_status = change['from_status']
            to_status = change['to_status']
            previous_timestamp = creation_time if i == 0 and creation_time else status_changes[i - 1]['timestamp']
            duration_hours = (change['timestamp'] - previous_timestamp).total_seconds() / 3600.0
            if duration_hours < 0:
                continue

            transition_key = (from_status, to_status)
            transition_durations[transition_key].append(duration_hours)
            transition_issues[transition_key].append({
                'issue_key': issue_key,
                'duration_hours': round(duration_hours, 6),
                'start_time': previous_timestamp,
                'end_time': change['timestamp']
            })

    # --- Compile results ---
    results = {}
    for (from_status, to_status), durations in transition_durations.items():
        if not durations:
            continue
        avg_hours = statistics.mean(durations)
        avg_days = avg_hours / 24
        stddev_hours = statistics.stdev(durations) if len(durations) > 1 else 0.0

        results[(from_status, to_status)] = {
            'average_hours': round(avg_hours, 6),
            'average_minutes': round(avg_hours * 60, 2),
            'average_days': round(avg_days, 2),
            'median_hours': round(statistics.median(durations), 6),
            'min_hours': round(min(durations), 6),
            'max_hours': round(max(durations), 6),
            'stddev_hours': round(stddev_hours, 6),
            'count': len(durations),
            'durations': durations,
            'issues': transition_issues[(from_status, to_status)]
        }

    print(f"Calculated {len(results)} unique status transitions (including 'Created').")
    return results

from collections import defaultdict
from datetime import datetime
from typing import List, Any, Dict
import statistics


from collections import defaultdict
from datetime import datetime, timezone
from typing import List, Any, Dict
import statistics


def calculate_average_chain_cycle_time(jira_issues: List[Any]) -> Dict[str, Dict[str, Any]]:
    """
    Calculate total and average cycle time for all unique end-to-end transition chains,
    starting from an artificial 'Created' status.
    Includes issues with no status transitions, treating them as 'Created' → <current status>.
    """

    chain_durations = defaultdict(list)
    chain_issues = defaultdict(list)

    now_utc = datetime.now(timezone.utc)
    print(f"Calculating end-to-end transition chain cycle times (including 'Created') for {len(jira_issues)} issues...")

    for issue in jira_issues:
        # --- Extract info ---
        if hasattr(issue, 'key'):
            issue_key = issue.key
            changelog = getattr(issue, 'changelog', None)
            fields = getattr(issue, 'fields', None)
            creation_time = getattr(fields, 'created', None) if fields else None
            current_status = getattr(fields.status, 'name', 'Unknown') if fields and hasattr(fields, 'status') else 'Unknown'
            if creation_time:
                creation_time = parse_jira_timestamp(creation_time)
        else:
            issue_key = issue.get('key', 'Unknown')
            changelog = issue.get('changelog')
            fields = issue.get('fields', {})
            creation_time = None
            current_status = fields.get('status', {}).get('name', 'Unknown')
            if 'created' in fields:
                try:
                    creation_time = parse_jira_timestamp(fields['created'])
                except ValueError:
                    pass

        print(f"Parsed creation time for issue {issue_key}: {creation_time}")
        print(f"Current status for issue {issue_key}: {current_status}")

        # --- Handle issues with no changelog or histories ---
        histories = None
        if changelog:
            histories = getattr(changelog, 'histories', None)
            if isinstance(changelog, dict):
                histories = changelog.get('histories', histories)

        status_changes = []
        if histories:
            # --- Extract status changes ---
            for history in sorted(histories, key=lambda h: getattr(h, 'created', '') if hasattr(h, 'created') else h.get('created', '')):
                created_str = getattr(history, 'created', None) or (history.get('created') if isinstance(history, dict) else None)
                if not created_str:
                    continue

                try:
                    timestamp = parse_jira_timestamp(created_str)
                except Exception:
                    continue

                items = getattr(history, 'items', None) or (history.get('items') if isinstance(history, dict) else [])
                for item in items:
                    field = getattr(item, 'field', None) if hasattr(item, 'field') else item.get('field', None)
                    if field != 'status':
                        continue
                    from_status = getattr(item, 'fromString', None) if hasattr(item, 'fromString') else item.get('fromString', None)
                    to_status = getattr(item, 'toString', None) if hasattr(item, 'toString') else item.get('toString', None)
                    if from_status and to_status:
                        status_changes.append((from_status, to_status, timestamp))

        # --- Sort by timestamp ---
        status_changes.sort(key=lambda x: x[2]) if status_changes else None

        # --- Insert artificial "Created" status ---
        if creation_time:
            if status_changes:
                first_status = status_changes[0][0]
                status_changes.insert(0, ("Created", first_status, creation_time))
            else:
                # No status changes — create synthetic chain
                status_changes = [("Created", current_status, creation_time)]
            print(f"Inserted 'Created' status for issue {issue_key} at {creation_time}.")
        else:
            print(f"No creation time for issue {issue_key}, skipping.")
            continue

        # --- Determine end time ---
        if len(status_changes) > 1:
            start_time = status_changes[0][2]
            end_time = status_changes[-1][2]
        else:
            # No transitions — measure up to now
            start_time = creation_time
            end_time = now_utc

        chain = " → ".join([s[0] for s in status_changes] + [status_changes[-1][1]])
        duration_hours = (end_time - start_time).total_seconds() / 3600.0

        if duration_hours < 0:
            print(f"Negative duration for issue {issue_key} in chain {chain}, skipping.")
            continue

        chain_durations[chain].append(duration_hours)
        chain_issues[chain].append({
            'issue_key': issue_key,
            'duration_hours': round(duration_hours, 6),
            'start_time': start_time,
            'end_time': end_time
        })

        print(f"Issue {issue_key} chain: {chain} duration: {duration_hours:.2f} hours")

    # --- Compile results ---
    results = {}
    for chain, durations in chain_durations.items():
        if not durations:
            continue
        avg_hours = statistics.mean(durations)
        stddev_hours = statistics.stdev(durations) if len(durations) > 1 else 0.0
        results[chain] = {
            'average_hours': round(avg_hours, 6),
            'average_days': round(avg_hours / 24, 3),
            'median_hours': round(statistics.median(durations), 6),
            'min_hours': round(min(durations), 6),
            'max_hours': round(max(durations), 6),
            'stddev_hours': round(stddev_hours, 6),
            'count': len(durations),
            'durations': durations,
            'issues': chain_issues[chain]
        }

    print(f"Calculated {len(results)} unique transition chains (including 'Created' and synthetic ones).")
    return results

def calculate_average_chain_cycle_time_old(jira_issues: List[Any]) -> Dict[str, Dict[str, Any]]:
    """
    Calculate total and average cycle time for all unique end-to-end transition chains,
    starting from an artificial 'Created' status.
    """

    chain_durations = defaultdict(list)
    chain_issues = defaultdict(list)

    print(f"Calculating end-to-end transition chain cycle times (including 'Created') for {len(jira_issues)} issues...")

    for issue in jira_issues:
        # --- Extract info ---
        if hasattr(issue, 'key'):
            issue_key = issue.key
            changelog = getattr(issue, 'changelog', None)
            creation_time = getattr(issue.fields, 'created', None) if hasattr(issue, 'fields') else None
            if creation_time:
                creation_time = parse_jira_timestamp(creation_time)
        else:
            issue_key = issue.get('key', 'Unknown')
            changelog = issue.get('changelog')
            creation_time = None
            fields = issue.get('fields', {})
            if 'created' in fields:
                try:
                    creation_time = parse_jira_timestamp(fields['created'])
                except ValueError:
                    pass

        print(f"Parsed creation time for issue {issue_key}: {creation_time}")

        if not changelog:
            print(f"No changelog for issue {issue_key}, skipping.")
            continue

        # --- Get histories ---
        histories = getattr(changelog, 'histories', None)
        if isinstance(changelog, dict):
            histories = changelog.get('histories', histories)
        if not histories:
            print(f"No histories in changelog for issue {issue_key}, skipping.")
            continue

        # --- Extract status changes ---
        status_changes = []
        for history in sorted(histories, key=lambda h: getattr(h, 'created', '') if hasattr(h, 'created') else h.get('created', '')):
            created_str = getattr(history, 'created', None) or (history.get('created') if isinstance(history, dict) else None)
            if not created_str:
                print(f"No created timestamp in history for issue {issue_key}, skipping this history.")
                continue

            try:
                timestamp = parse_jira_timestamp(created_str)
            except Exception:
                print(f"Failed to parse timestamp '{created_str}' for issue {issue_key}, skipping this history.")
                continue

            items = getattr(history, 'items', None) or (history.get('items') if isinstance(history, dict) else [])
            for item in items:
                field = getattr(item, 'field', None) if hasattr(item, 'field') else item.get('field', None)
                if field != 'status':
                    continue
                from_status = getattr(item, 'fromString', None) if hasattr(item, 'fromString') else item.get('fromString', None)
                to_status = getattr(item, 'toString', None) if hasattr(item, 'toString') else item.get('toString', None)
                if from_status and to_status:
                    print(f"Issue {issue_key} status change: {from_status} → {to_status} at {timestamp}")
                    status_changes.append((from_status, to_status, timestamp))

        if not status_changes:
            print(f"No status_changes found for issue {issue_key}, skipping.")
            continue

        # --- Sort by timestamp ---
        status_changes.sort(key=lambda x: x[2])

        # --- Insert artificial "Created" → first status ---
        if creation_time:
            first_status = status_changes[0][0]
            status_changes.insert(0, ("Created", first_status, creation_time))
            print(f"Inserted 'Created' status for issue {issue_key} at {creation_time}.")

        # --- Build full chain and compute total duration ---
        chain = " → ".join([s[0] for s in status_changes] + [status_changes[-1][1]])
        start_time = status_changes[0][2]
        end_time = status_changes[-1][2]
        duration_hours = (end_time - start_time).total_seconds() / 3600.0

        if duration_hours < 0:
            print(f"Negative duration for issue {issue_key} in chain {chain}, skipping (likely timestamp order issue).")
            continue

        chain_durations[chain].append(duration_hours)
        chain_issues[chain].append({
            'issue_key': issue_key,
            'duration_hours': round(duration_hours, 6),
            'start_time': start_time,
            'end_time': end_time
        })

        print(f"Issue {issue_key} chain: {chain} duration: {duration_hours:.2f} hours")

    # --- Compile results ---
    results = {}
    for chain, durations in chain_durations.items():
        if not durations:
            continue
        avg_hours = statistics.mean(durations)
        stddev_hours = statistics.stdev(durations) if len(durations) > 1 else 0.0
        results[chain] = {
            'average_hours': round(avg_hours, 6),
            'average_days': round(avg_hours / 24, 3),
            'median_hours': round(statistics.median(durations), 6),
            'min_hours': round(min(durations), 6),
            'max_hours': round(max(durations), 6),
            'stddev_hours': round(stddev_hours, 6),
            'count': len(durations),
            'durations': durations,
            'issues': chain_issues[chain]
        }

    print(f"Calculated {len(results)} unique transition chains (including 'Created').")
    return results





from datetime import datetime, timezone
from dateutil import parser
from collections import defaultdict
import statistics
from typing import List, Any, Dict

def parse_jira_timestamp(ts_str: str) -> datetime:
    """Parse a Jira timestamp into a timezone-aware UTC datetime."""
    dt = parser.isoparse(ts_str)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt


def parse_jira_timestamp_old(timestamp_str: str) -> datetime:
    import re
    from datetime import datetime, timezone, timedelta

    ts = timestamp_str.strip()

    # Convert "-0400" → "-04:00" to make it ISO-compatible
    tz_pattern = r'([+-])(\d{2})(\d{2})$'
    match = re.search(tz_pattern, ts)
    if match:
        sign, hours, minutes = match.groups()
        ts = re.sub(tz_pattern, f'{sign}{hours}:{minutes}', ts)

    formats = [
        '%Y-%m-%dT%H:%M:%S.%f%z',  # e.g. 2025-09-02T23:31:32.138-04:00
        '%Y-%m-%dT%H:%M:%S%z',     # e.g. 2025-09-02T23:31:32-04:00
        '%Y-%m-%dT%H:%M:%S.%f',    # no timezone, with milliseconds
        '%Y-%m-%dT%H:%M:%S',       # no timezone, no milliseconds
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(ts, fmt)
            # If it has tzinfo, convert to UTC and drop tz awareness
            if dt.tzinfo:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt
        except ValueError:
            continue

    raise ValueError(f"Unable to parse timestamp: {timestamp_str}")


def parse_jira_timestamp_old(timestamp_str: str) -> datetime:
    import re
    from datetime import timezone
    
    ts = timestamp_str.strip()
    
    # Convert -0400 to -04:00 format
    tz_pattern = r'([+-])(\d{2})(\d{2})$'
    match = re.search(tz_pattern, ts)
    if match:
        sign, hours, minutes = match.groups()
        ts = re.sub(tz_pattern, f'{sign}{hours}:{minutes}', ts)
    
    formats = [
        '%Y-%m-%dT%H:%M:%S.%f%z',  # 2025-09-02T23:31:32.138-04:00
        '%Y-%m-%dT%H:%M:%S%z',     # 2025-09-02T23:31:32-04:00
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(ts, fmt)
        except ValueError:
            continue
    
    raise ValueError(f"Unable to parse timestamp: {timestamp_str}")


def print_transition_report(transition_data: Dict[Tuple[str, str], Dict[str, Any]]) -> None:
    """
    Print a formatted report of status transition times.
    """
    if not transition_data:
        print("No status transitions found in the provided issues.")
        return
    
    print("JIRA Status Transition Time Analysis")
    print("=" * 50)
    
    # Sort transitions by average time (descending)
    sorted_transitions = sorted(
        transition_data.items(),
        key=lambda x: x[1]['average_hours'],
        reverse=True
    )
    
    for (from_status, to_status), data in sorted_transitions:
        print(f"\n{from_status} → {to_status}")
        print(f"  Average Time: {data['average_hours']:.1f} hours ({data['average_days']:.1f} days)")
        print(f"  Median Time:  {data['median_hours']:.1f} hours")
        print(f"  Range:        {data['min_hours']:.1f} - {data['max_hours']:.1f} hours")
        print(f"  Sample Size:  {data['count']} transitions")


from collections import defaultdict
from typing import List, Any, Dict

def get_issues_with_no_transitions(jira_issues: List[Any]) -> Dict[str, List[str]]:
    """
    Identify JIRA issues that have no status transitions in their changelog,
    grouped by their current status.

    Returns:
        Dict[status_name, [issue_keys]]
    """
    no_transition_statuses = defaultdict(list)

    for issue in jira_issues:
        # --- Extract issue key safely ---
        if hasattr(issue, "key"):
            issue_key = issue.key
        elif isinstance(issue, dict):
            issue_key = issue.get("key", "Unknown")
        else:
            issue_key = "Unknown"

        # --- Extract changelog safely ---
        if hasattr(issue, "changelog"):
            changelog = issue.changelog
        elif isinstance(issue, dict):
            changelog = issue.get("changelog")
        else:
            changelog = None

        # --- If no changelog at all ---
        if not changelog:
            current_status = None
            if hasattr(issue, "fields") and hasattr(issue.fields, "status"):
                current_status = getattr(issue.fields.status, "name", None)
            elif isinstance(issue, dict):
                status_obj = issue.get("fields", {}).get("status")
                if isinstance(status_obj, dict):
                    current_status = status_obj.get("name")
            no_transition_statuses[current_status or "Unknown"].append(issue_key)
            continue

        # --- Extract histories safely ---
        histories = None
        if hasattr(changelog, "histories"):
            histories = changelog.histories
        elif isinstance(changelog, dict):
            histories = changelog.get("histories", [])
        if not histories:
            current_status = None
            if hasattr(issue, "fields") and hasattr(issue.fields, "status"):
                current_status = getattr(issue.fields.status, "name", None)
            elif isinstance(issue, dict):
                status_obj = issue.get("fields", {}).get("status")
                if isinstance(status_obj, dict):
                    current_status = status_obj.get("name")
            no_transition_statuses[current_status or "Unknown"].append(issue_key)
            continue

        # --- Scan for status changes ---
        has_status_change = False
        for history in histories:
            items = getattr(history, "items", None) or (history.get("items") if isinstance(history, dict) else [])
            for item in items:
                field_name = getattr(item, "field", None) if hasattr(item, "field") else item.get("field", None)
                if field_name == "status":
                    has_status_change = True
                    break
            if has_status_change:
                break

        # --- If no status change found, group issue by current status ---
        if not has_status_change:
            current_status = None
            if hasattr(issue, "fields") and hasattr(issue.fields, "status"):
                current_status = getattr(issue.fields.status, "name", None)
            elif isinstance(issue, dict):
                status_obj = issue.get("fields", {}).get("status")
                if isinstance(status_obj, dict):
                    current_status = status_obj.get("name")
            no_transition_statuses[current_status or "Unknown"].append(issue_key)

    print(f"Found {sum(len(v) for v in no_transition_statuses.values())} issues with no transitions.")
    return dict(no_transition_statuses)


import hashlib, re

def to_filename(path: str, max_len=200):
    # Make filesystem-safe
    clean = re.sub(r'\s*→\s*', '_', path).replace(' ', '-')
    
    # Hash full path for uniqueness
    hash_suffix = hashlib.md5(path.encode()).hexdigest()[:6]
    
    # Truncate safely before appending hash
    trimmed = clean[:max_len]
    
    return f"{trimmed}_{hash_suffix}"



def write_execsummary_yaml(jira_ids, file_info, chain_str, chain_row, timestamp):
    # always create <exec summary> yaml files incase they're needed down the chain
    # step 1 hunt for jira id in all rows and build a list
    # step 2 hunt for jql in all the rows and get list of jira id and add to list from #1
    # step 3 Create ayaml new exec summary scope yaml file with all the Jira found in #1 #2
    # step 4 read_jira will process this yaml downstream

    fname = f"{file_info['basename']}.{file_info['sheet']}.{file_info['table']}.{chain_str}"
    fname_hash = to_filename(fname)
    #execsummary_scope_output_file = f"{file_info['basename']}.{file_info['sheet']}.{file_info['table']}.{table_hash}.{timestamp}.aisummary.scope.yaml"
    execsummary_scope_output_file = f"{fname_hash}.{timestamp}.chain.scope.yaml"
    print(f"write_execsummary_yaml(...) called to write to file={execsummary_scope_output_file}")

    file_info["scope file"] = execsummary_scope_output_file
    #file_info["table"] = cleaned_value

    with open(execsummary_scope_output_file, 'w') as f:
        yaml.dump({ "fileinfo": file_info }, f, default_flow_style=False)
        yaml.dump({ "chain_row": chain_row }, f, default_flow_style=False)

    jira_fields = []

    jira_fields.append({"value": "key", "index": 0})
    jira_fields.append({"value": "summary", "index": 0})
    jira_fields.append({"value": "description", "index": 0})
    jira_fields.append({"value": "status", "index": 0})
    jira_fields.append({"value": "issuetype", "index": 0})
    jira_fields.append({"value": "priority", "index": 0})
    jira_fields.append({"value": "created", "index": 0})
    jira_fields.append({"value": "assignee", "index": 0})
    jira_fields.append({"value": "status", "index": 0})
    jira_fields.append({"value": "comments", "index": 0})


    with open(execsummary_scope_output_file, 'a') as f:
        yaml.dump({ "fields": jira_fields }, f, default_flow_style=False)
   
    if jira_ids:
        print(f"{len(jira_ids)} JIRA IDs found: {jira_ids}")

        # instead just dump the jira_ids to the scope file. read_jira.py will take care of it, ie run jql and get the jira ids.         
        with open(execsummary_scope_output_file, 'a') as f:
            yaml.dump({"jira_ids": jira_ids}, f, default_flow_style=False)
            # commented out since  defautl values feature not supported or needed in this case
            # only interested in generated a yaml file with fields ids and jira ids that will be used
            # by cycletime.py on 2nd pass to fill in aisummary for each chain 
            #print(f"Fieldname args found for: {jira_fields_default_value}")
            #yaml.dump({"field_args": jira_fields_default_value}, f, default_flow_style=False)

    else:
        print(f"ERROR: can't proceed, No JIRA IDs found to write to aisummary yaml file {execsummary_scope_output_file}")
        sys.exit(1)

    f.close()
    print("ExecSummary scope yaml file created successfully:", execsummary_scope_output_file)


# Build LLM context from table rows and corresponding jiracsv file
# so yes, it combines excel table rows (all cells including non-jira ones)
# and combines with the aisummary.jira.csv file contents
def build_llm_context(table_name, timestamp):
    
    context = ""
    
    '''# first put in the sheet contents for this table
    for r in table_rows:
        # r is a list of cell values, join them into a string
        context += " | ".join(str(cell) for cell in r) + "\n"
    '''

     # now append the jiracsv contents
    jiracsv_pattern = f"{basename}.{sheet}.{table_name}.{timestamp}.aisummary.jira.csv"
    dir_path = os.getcwd()  # or the folder where your CSVs live

    # case-insensitive search
    matched_file = None
    for f in os.listdir(dir_path):
        if f.lower() == jiracsv_pattern.lower():
            matched_file = os.path.join(dir_path, f)
            break

    if matched_file:
        with open(matched_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            context += line
    else:
        print(f"Warning: file not found (case-insensitive match) for {jiracsv_pattern}")

    return context

# Default to localhost unless overridden in env variable (set when in Docker)
SUMMARIZER_HOST = os.getenv("SUMMARIZER_HOST", "http://localhost:8000")
#LLMCONFIG_FILE = "../../../config/llmconfig.json"


def get_llm_model(llm_config_file):
    print("Current working directory:", os.getcwd())  # <-- debug
    if os.path.exists(llm_config_file):
        with open(llm_config_file, "r") as f:
            llm_model_set = json.load(f)
            m = llm_model_set.get("model")
            print(f"get_llm_model returning  {m}")
            return m
    else:
        print(f"ERROR: load_llm_config file {llm_config_file} was not found")
    
    return None


# variable names refer to comments but they don't have to be.  This function used for any field that were found to have an LLM prompt 
def get_summarized_comments(comments_list_asc, field_arg=None):
    """
    Summarize comments for LLM processing.
    This function takes a list of comments in ascending order and returns a summarized version.
    Optionally, a specific field name can be passed via `field_arg`.
    """
    try:
        if not comments_list_asc:
            return "No comments available."

        # Only join if it's a list or tuple
        if isinstance(comments_list_asc, (list, tuple)):
            comments_str = "; ".join(comments_list_asc)
        else:
            comments_str = comments_list_asc  # already a string

        comments_str = comments_str.replace("\n", "").replace("\r", "")

        # Prepare the payload for the service
        payload = {
            "comments": [comments_str]  # service expects a list[str]
        }
        if field_arg:
            payload["field"] = field_arg  # include field if provided

        LLMCONFIG_FILE = f"../../../config/llmconfig_{userlogin}.json"
        llm_model = get_llm_model(LLMCONFIG_FILE)
         # Determine endpoint
        if llm_model == "OpenAI":
            ENDPOINT = "/summarize_openai_ex"
        elif llm_model == "Local":
            ENDPOINT = "/summarize_local_ex"
        elif llm_model == "Claude":
            ENDPOINT = "/summarize_claude_ex"
        else:
            ENDPOINT = "/summarize_local_ex"

        # Make the POST request
        resp = requests.post(f"{SUMMARIZER_HOST}{ENDPOINT}", json=payload)

        if resp.status_code == 200:
            full_response = resp.json().get("summary", "")
            print(f"LLM endpoint returned {resp.status_code} OK")
        else:
            print(f"error LLM endpoint returned {resp.status_code}")
            full_response = f"[ERROR] Service call failed: {resp.text}"

        # Clean up response
        full_response = full_response.rstrip("\n").replace("\n", "; ").replace("|", "/")

        print(f"Full response: {full_response}")

        return full_response
    
    except Exception as e:
        # Log the exception and return a safe default
        print(f"[EXCEPTION ERROR] LLM could not be engaged, get_summarized_comments failed: {e}")
        #return "[ERROR] LLM could not be engaged due to exceptions during LLM interaction."
        return f"[EXCEPTION ERROR] LLM could not be engaged, get_summarized_comments failed: {e}"


def get_summarized_comments_old(context, sysprompt):
    """
    Summarize comments for LLM processing.
    This function takes a list of comments in ascending order and returns a summarized version.
    """
    try:
        if not context:
            return "No context provided."

        # Only join if it's a list or tuple
        if isinstance(context, (list, tuple)):
            comments_str = "; ".join(context)
        else:
            comments_str = context  # already a string

        # comments_str was a single string before, but the service expects a list[str].
        # If you only have one string, wrap it in a list.
        #context = [comments_str]

        prompt = sysprompt + ".\n\n" + context
        #prompt = sysprompt + "\n\n" + "\n".join(context)

        prompt_list = [prompt]
        print(f"calling LLM with prompt = {prompt_list[0][:255]}...")

        llm_model = get_llm_model(LLMCONFIG_FILE)

        if llm_model == "OpenAI":
            ENDPOINT = "/summarize_openai"
        else:
            ENDPOINT = "/summarize_local"

        resp = requests.post(f"{SUMMARIZER_HOST}{ENDPOINT}", json=prompt_list)

        if resp.status_code == 200:
            full_response = resp.json()["summary"]
        else:
            full_response = f"[ERROR] Service call failed: {resp.text}"    
        
        
        print(f"Full response: {full_response}")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return  full_response   
    except Exception as e:
        # Log the exception and return a safe default
        print(f"[EXCEPTION THROWN ERROR] get_summarized_comments failed: {e}")
        #return "[ERROR] Summary could not be generated."
        return f"[EXCEPTION THROWN ERROR] get_summarized_comments failed: {e}"



def html_to_text_with_structure(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    def walk(node, depth=0):
        parts = []
        for child in node.children:
            if child.name is None:  # NavigableString
                text = child.strip()
                if text:
                    # add indentation spaces based on depth
                    parts.append(" " * (depth * 2) + text)
            else:
                # Handle block-level tags with newlines
                if child.name in ("p", "div", "section", "article", "header", "footer",
                                  "ul", "ol", "li", "br", "h1", "h2", "h3", "h4", "h5", "h6"):
                    if child.name == "li":
                        parts.append(" " * (depth * 2) + "- " + walk(child, depth + 1).strip())
                    else:
                        inner = walk(child, depth + 1)
                        if inner:
                            parts.append(inner)
                    parts.append("\n")  # newline after block
                else:
                    # Inline tag (span, b, i, etc.)
                    inner = walk(child, depth)
                    if inner:
                        parts.append(inner)
        return "\n".join(p for p in parts if p.strip())

    text = walk(soup).strip()

    # Collapse excessive blank lines
    lines = [line.rstrip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line.strip() != "")


import os
import sys
import yaml
from datetime import datetime
from urllib.parse import unquote
from openpyxl.utils import get_column_letter

if len(sys.argv) < 4:
    print("Usage: python cycletime.py <yaml_file> <timestamp> <userlogin>")
    sys.exit(1)

yaml_file = sys.argv[1]
timestamp = sys.argv[2]
userlogin = sys.argv[3]

print(f"cycletime parameters  yaml_file={yaml_file} timestamp={timestamp} userlogin={userlogin}")

with open(yaml_file, 'r') as f:
    data = yaml.safe_load(f)

fileinfo = data.get('fileinfo', {})
if not fileinfo:
    print("No fileinfo found in the YAML file.")
    sys.exit(1)
basename = fileinfo.get('basename')
tablename = fileinfo.get('table').replace(" ", "_") if fileinfo.get('table') else ""
source = fileinfo.get('source')
scope_file = fileinfo.get('scope file')
sheet = fileinfo.get('sheet')

if not basename:
    print("No 'basename'found in fileinfo. Expecting 'basename' key.")
    sys.exit(1)

if not tablename:
    print("No 'table' found in fileinfo. Expecting 'table' key.")
    sys.exit(1)


# I don't think this code is doing anything useful. It always sets create_mode to True?! since 
# Determine if we will be INSERTING rows eventually vs just UPDATING existing rows in Excel/SharePoint
if "cycletime.scope" in yaml_file.lower():
    print(f"cycletime detected based on filename: {yaml_file}.")
    # You can set a flag or handle import-specific logic here if needed
    output_file = basename + "." + tablename + timestamp + ".cycletime.jira.csv"
    #mode = "assignee" #"resolved"
else:
    print("Error: YAML filename is not 'cycletime.scope.yaml")
    sys.exit(1)

fields = data.get('fields', [])
#scan_ahead_nonblank_rows = data.get('scan_ahead_nonblank_rows', 0)

scan_ahead_nonblank_rows = data.get('last_update_row_count',0) 
print(f"scan_ahead_nonblank_rows initialized = {scan_ahead_nonblank_rows}")


# Convert list of dicts into a dictionary
fields_dict = {field["value"]: field.get("index", "<blank>") for field in fields}

field_values = [field.get('value') for field in fields if 'value' in field]
field_indexes = [field.get('index') for field in fields if 'index' in field]
field_values_str = ','.join(field_values)
field_indexes_str = ','.join(map(str, field_indexes))

print("Source file,", source)
print("basename,", basename)
print("Table,", tablename)
print("Field indexes,", field_indexes_str)
print("Field values,", field_values_str)

'''with open(output_file, "w") as outfile:
    outfile.write("Source file," + source + "\n")
    outfile.write("Basename," + basename + "\n")
    outfile.write("Scope file," + scope_file + "\n")    
    outfile.write("Table," + tablename + "\n")
    outfile.write("Field indexes," + field_indexes_str + "\n")
    outfile.write("Field values," + field_values_str + "\n")
'''

# yaml will expected to contain params, eg:
'''
    fileinfo:
    basename: Quickstart.xlsx
    scope file: Quickstart.xlsx.Cycle_Time.20250928_164350.cycletime.scope.yaml
    source: Quickstart.xlsx
    table: Cycle_Time
    jql: project in (tes,fr) and updated > -90d
    row: 27
    col: 1
    lastrow: 49
'''

# the following were addded by scope.py in runrate yaml file
cycletime_table_row = data.get('row', None)
print(f"quickstart_table_row from yaml: {cycletime_table_row}")
cycletime_table_col = data.get('col', None)
print(f"quickstart_table_col from yaml: {cycletime_table_col}")
last_excel_row = data.get('lastrow', None)
print(f"last_excel_row from yaml: {last_excel_row}")
jql_str = data.get('jql', None)
print(f"jql from yaml: {jql_str}")

llm_user_prompt = data.get('llm', "Read all of it and briefly as possible categorize types of issues and work that was done. Mention any reason you see that could have blocked work on these issues or could have been done more quickly or correctly." )

# other prompt ideas:
# idea 1
# Read all the comments in the jira issues and list out any issues that are blocked and reason.  
# Also mention any improvements we can make that will allow work t be completed more quickly or correctly.
#
# idea 2
# <llm> Read all the comments in the jira issues that provide or suggest  improvements idea to get things done more efficiently or correctly
# add my prefix prompt regardless of user specified prompt or default prompt


sysprompt = "The following text is a delimited data separated by | character. These are rows of jira issues. " + llm_user_prompt

print (f"llm prompt = {sysprompt}")

# validate
if not jql_str or cycletime_table_row is None or cycletime_table_col is None or last_excel_row is None:
    print("❌ Error: Missing required fields in YAML file")
    sys.exit(1)


    
from jira import JIRA
from dotenv import load_dotenv

JIRA_MAX_RESULTS = False

# Load environment variables from a .env file if present
# -------------------------------
#load_dotenv()
# load .env from config folder
ENV_PATH = f"../../../config/env.{userlogin}"
load_dotenv(dotenv_path=ENV_PATH)

JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN")
JIRA_URL = os.environ.get("JIRA_URL")
JIRA_EMAIL = os.environ.get("JIRA_EMAIL")

if not JIRA_API_TOKEN:
    print("Error: JIRA_API_TOKEN environment variable not set.")
    sys.exit(1)



    import re
    from typing import Optional

    def extract_rows_count(text: str) -> Optional[int]:
        """
        Extract the number of rows mentioned in a message.
        Matches numbers like:
        "3 rows updated on 2025-11-03 14:38:54 by Trinket" -> 3
        "1 row updated ..." -> 1
        "1,234 rows ..." -> 1234
        Returns an int if found, otherwise None.
        """
        match = re.search(r'([\d,]+)\s+rows?\b', text, flags=re.IGNORECASE)
        if not match:
            return None
        number_str = match.group(1).replace(',', '')
        try:
            return int(number_str)
        except ValueError:
            return None


#########################


# first check if the chain.jira.csv already exists, which would mean this is 2nd call to cycletime
# only calculate_average_chain_cycle_time on the first call. so we can skip if 2nd call and move 
# to processing chain.jira.csv in this scenario

context = ""
csv = os.path.join(os.getcwd(), f"{basename}.{sheet}.{tablename}.*.{timestamp}.chain.jira.csv")
csv_files = glob.glob(csv)

if csv_files:
    # yes, we have chain.jira.csv files so this must be 2nd pass of cycletime from resync

    for csv_file in csv_files:
        context = ""
        print(f"found csv file {csv_file}")
        with open(csv_file, "r", encoding="utf-8") as f:
            for _ in range(5):  # skip first 5 lines
                next(f, None)
            for line in f:
                context += line 


        match = re.match(rf"{re.escape(basename)}\.{sheet}\.{tablename}\.(.+?)\.chain\.jira\.csv", os.path.basename(csv_file))
        if match:
            substring = match.group(1)
            output_file = f"{basename}.{sheet}.{tablename}.{substring}.chain.llm.txt"
            context_output_file =  f"{basename}.{sheet}.{tablename}.{substring}.chain.context.txt"
            chain_yaml_file =  f"{basename}.{sheet}.{tablename}.{substring}.chain.scope.yaml"
            print(f"context for {csv_file} will be saved in {output_file}")
            print(f"chain row for {csv_file} will be looked in {chain_yaml_file}")
        else:
            print(f"ERROR: could not find chain.jira.csv pattern in {csv_file} to generate context for LLM. Will exit cycletime.py")
            sys.exit(0)


        # read the chain_row that was saved in this chain's scope.yaml file earlier
        with open(chain_yaml_file, 'r') as f:
            data = yaml.safe_load(f)
            rownum = data.get('chain_row', None)
            if rownum:
                print(f"chain row is {rownum}")
            else:
                print(f"No chain_row found in the YAML file {chain_yaml_file}")
                sys.exit(1)
            
        print(f"Calling get_summarized_comments with context={context[:255]}... and sysprompt={sysprompt[:255]}...")
        report = get_summarized_comments(context, sysprompt)

        # Save context to file
        with open(context_output_file, "w", encoding="utf-8") as f:
            f.write(context)
            print(f"LLM context saved to {context_output_file}")

        # save llm's response to file for debugging
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(report)
            print(f"LLM reponse saved to {output_file}")

        # Replace all newlines with semicolons
        cleaned_response = html_to_text_with_structure(report)
        #print(f"html_to_text_with_structure returned cleaned_response={cleaned_response}")
        cleaned_response = cleaned_response.rstrip("\n")
        cleaned_response = cleaned_response.replace("\n", "; ")
        cleaned_response = cleaned_response.replace("|", "^")
        
        # now update the changes.txt with llm response
        # the chain's changes.txt file will be automaically consumed by update_googlesheet (or update_sharepoint) downstream
        # +1 because cols start at 1 = A
        # +8 because it needs to be displayed on right end of cycletime table on sheet 
        entry = f"{get_column_letter(cycletime_table_col + 1 + 8)}{rownum} = {cleaned_response} || "  
    
        changes_file = yaml_file.replace("scope.yaml","import.changes.txt")
        print(f"Will write LLM response into {changes_file}")

        with open(changes_file, "a") as f:
            f.write(entry + "\n")
            print(entry)
       
        print(f"Changes written to {changes_file} entry = {entry}")

    
    # no need to proceed since cycle time has completed 2 passess at this point
    sys.exit(0)    
else:
    print (f"No chain.jira.csv files found matching pattern {csv} so assume it's first pass of cycletime.py")





# Connect to Jira with basic auth
try:
    jira = JIRA(server=JIRA_URL, basic_auth=(JIRA_EMAIL, JIRA_API_TOKEN))
    #print("✅ Successfully connected to Jira.")
except Exception as e:
    print(f"❌ Failed to connect to Jira: {e}")
    sys.exit(1)

print(f"JIRA client connected to {JIRA_URL} login:{JIRA_EMAIL} apitoken:{JIRA_API_TOKEN} ")

issues = []  # global issues list to hold results from both ID and JQL searches



try:
    jql_query = jql_str
    jql_query = jql_query.lower().replace("jql ", "").strip()
    issues = jira.search_issues(jql_query, maxResults = JIRA_MAX_RESULTS, expand='changelog')
    print(f"Found {len(issues)} issues for JQL query '{jql_query}':")
    if len(issues) == 0:
        print(f"No issues found for JQL query '{jql_query}'.")
        sys.exit(1)
    filtered_ids = [issue.key for issue in issues]
    print(f"Filtered IDs from JQL: {filtered_ids}")
    #jql_ids = []  # Clear jql_ids to avoid re-processing later below
except Exception as e:
    print(f"❌ Failed to search issues for JQL query '{jql_query}': {e}")
    sys.exit(1)



# rest of this is done ONLY on first call to cycletime.py

# Calculate chain cycle times
results = calculate_average_chain_cycle_time(issues)
print(f"\nCalculated cycle times for {len(results)} unique transition chains.")

transition_data = results

# Sort chains by average time (descending)
sorted_chains = sorted(
    transition_data.items(),
    key=lambda x: x[1]['average_hours'],
    reverse=True
)

changes_list = []
r = cycletime_table_row 
excel_col = cycletime_table_col + 1


#last_update_count = extract_rows_count(s)

if scan_ahead_nonblank_rows:
    prefix = ""
    #scan_ahead_nonblank_rows -= 1      # don't subtract since we didn't include this row in the scan_ahead count
else:
    prefix = "INSERT"


r += 1  # bump row one more time so INSERT are done at the row right below the cycletime tag

# write out the timestamp in cell adjacent to <> so we can tell when the update occurred
coord = f"{get_column_letter(excel_col)}{r}"
now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
entry = f"{coord} = {prefix} {len(results)} rows updated on {now_str} by Trinket ||"
print(entry)
changes_list.append(entry)

r += 1  # bump row one more time so INSERT are done at the row right below the cycletime tag

if scan_ahead_nonblank_rows:
    prefix = ""
    #scan_ahead_nonblank_rows -= 1  # don't substract since we didn't count the headings row in scan_ahead count 
else:
    prefix = "INSERT"

changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} TRANSITION CHAIN || ")
changes_list.append(f"{get_column_letter(excel_col + 1)}{r} =   AVERAGE || ")
changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =   MEDIAN  || ")
changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =   STD DEV || ")
changes_list.append(f"{get_column_letter(excel_col + 4)}{r} =   MIN || ")
changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =   MAX || ")
changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =   SAMPLE SIZE || ")
changes_list.append(f"{get_column_letter(excel_col + 7)}{r} =   % of Total Jira || ")
changes_list.append(f"{get_column_letter(excel_col + 8)}{r} =   {llm_user_prompt} || ")


total_jira = 0 
for chain_str, data in sorted_chains:
    total_jira += data['count']

print(f"total_jira = {total_jira} for chain % calculation")

for chain_str, data in sorted_chains:

    r += 1

    if scan_ahead_nonblank_rows:
        prefix = ""
        scan_ahead_nonblank_rows -= 1
    else:
        prefix = "INSERT"# get total number of issues

    print(f"\n{chain_str}")
    median_days = data['median_hours'] / 24.0
    stdev_days = data['stddev_hours'] / 24.0
    min_days = data['min_hours'] / 24.0
    max_days = data['max_hours'] / 24.0
    changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} {chain_str} || ")
    changes_list.append(f"{get_column_letter(excel_col + 1)}{r} =  {data['average_hours']:.1f} hrs ({data['average_days']:.1f}d) || ")
    changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =  {data['median_hours']:.1f} hrs ({median_days:.1f}d)|| ")
    changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =  {data['stddev_hours']:.1f} hrs ({stdev_days:.1f} days) || ")
    changes_list.append(f"{get_column_letter(excel_col + 4)}{r} =  {data['min_hours']:.1f} hrs ({min_days:.1f}d) || ")
    changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =  {data['max_hours']:.1f} hrs ({max_days:.1f}d) || ")

    
    # save the row number where this chain is displayed in the sheet
    # will be use by write_execsummary_yaml() function to save it in that chain.scope.yaml file
    # this is required so that we can recall where to write the LLM output for this chain on the sheet
    chain_row = r

    jira_ids = []

    # Build hyperlink JQL for thie issues in this chain
    jql = "key in ("
    issue_keys = [issue['issue_key'] for issue in data['issues']]
    for key in issue_keys:
        jql += key + ","
        jira_ids.append(key)    # also save this key for aisummary yaml later belwo
    jql = jql.rstrip(",") + ") order by key asc"

    from my_utils import *
    hyperlink = make_hyperlink_formula(f"{JIRA_URL}/issues/?jql={jql}", f"{data['count']}") + " || "
    changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =  {hyperlink} || ")

    percent = (data['count']/total_jira) * 100
    print(f"{data['count']}/{total_jira} = {percent:.1f}% ")
    changes_list.append(f"{get_column_letter(excel_col + 7)}{r} =  {percent:.1f}% || ")
   
    '''r += 1

    if scan_ahead_nonblank_rows:
        prefix = ""
        scan_ahead_nonblank_rows -= 1
    else:
        prefix = "INSERT"
    '''
    write_execsummary_yaml(jira_ids, fileinfo, chain_str, chain_row,  timestamp)


    print(f"  Average Time: {data['average_hours']:.1f} hours ({data['average_days']:.1f} days)")
    print(f"  Median Time:  {data['median_hours']:.1f} hours")
    print(f"  Range:        {data['min_hours']:.1f} - {data['max_hours']:.1f} hours")
    print(f"  Sample Size:  {data['count']} issues")

print(f"scan_ahead_nonblank_rows remaining = {scan_ahead_nonblank_rows}")
while (scan_ahead_nonblank_rows):
    print("filling in remaining rows with DELETE")
    r +=1
    scan_ahead_nonblank_rows -= 1
    changes_list.append(f"{get_column_letter(excel_col)}{r} = DELETE || ")

###########################


'''
# Calculate transition times
results = calculate_average_status_transition_time(issues)
print(f"\nCalculated transition times for {len(results)} status transitions.")

# Print results
#print_transition_report(results)

transition_data = results
  
# Sort transitions by average time (descending)
sorted_transitions = sorted(
    transition_data.items(),
    key=lambda x: x[1]['average_hours'],
    reverse=True
)


#changes_list = []
#r = cycletime_table_row + 1
#excel_col = cycletime_table_col + 1

# write out the timestamp in cell adjacent to <> so we can tell when the update occured
coord = f"{get_column_letter(cycletime_table_col + 2)}{cycletime_table_row + 1}"
now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
entry = f"{coord} = {now_str} ||"
print (entry)
changes_list.append(entry)


if scan_ahead_nonblank_rows:
    prefix = ""
    scan_ahead_nonblank_rows -= 1
else:
    prefix = "INSERT"

r += 1  # bump row one more time so INSERT are done at the row right below the cycletime tag 
changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} FROM STATUS || ")  # add INSERT to only first since the rest are on same row
changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = TO STATUS || ")  # add INSERT to only first since the rest are on same row
changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =   Average Time || ")
changes_list.append(f"{get_column_letter(excel_col + 3 )}{r} =   Median Time || ")
changes_list.append(f"{get_column_letter(excel_col + 4 )}{r} =   StdDev Time || ")
#changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =  INSERT Range || ")
changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =   Min Time || ")
changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =   Max Time|| ")
changes_list.append(f"{get_column_letter(excel_col + 7)}{r} =   Sample Size || ")

r +=1

for (from_status, to_status), data in sorted_transitions:

    
    if scan_ahead_nonblank_rows:
        prefix = ""
        scan_ahead_nonblank_rows -= 1
    else:
        prefix = "INSERT"

    print(f"\n{from_status} → {to_status}")    
    changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} {from_status} || ")  # add INSERT only to first row since rest are in sme row
    changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = {to_status} || ")  # add INSERT only to first row since rest are in sme row
    changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =  {data['average_hours']:.1f} hours ({data['average_days']:.1f} days) || ")
    changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =  {data['median_hours']:.1f} hours || ")
    changes_list.append(f"{get_column_letter(excel_col + 4)}{r} =  {data['stddev_hours']:.1f} hours || ")
    #changes_list.append(f"{get_column_letter(excel_col + 3)}{r} = INSERT {data['min_hours']:.1f} - {data['max_hours']:.1f} hours || ")
    changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =  {data['min_hours']:.1f} hours || ")
    changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =  {data['max_hours']:.1f} hours || ")

    jql = "key in ("
    # Print only the issue keys for this transition
    issue_keys = [issue['issue_key'] for issue in data['issues']]
    for key in issue_keys:
        #print(f"{key}")
        jql = jql + key + ","

    jql = jql.rstrip(",") + ")"

    from my_utils import *
    hyperlink = make_hyperlink_formula(f"{JIRA_URL}/issues/?jql={jql}", f"{data['count']}") + " || "
    changes_list.append(f"{get_column_letter(excel_col + 7)}{r} =  {hyperlink} || ")
    #changes_list.append(f"{get_column_letter(excel_col + 4)}{r} = INSERT {data['count']} issues || ")
    
    r += 1

    print(f"  Average Time: {data['average_hours']:.1f} hours ({data['average_days']:.1f} days)")
    print(f"  Median Time:  {data['median_hours']:.1f} hours")
    print(f"  Range:        {data['min_hours']:.1f} - {data['max_hours']:.1f} hours")
    print(f"  Sample Size:  {data['count']} issues")

changes_list.append(f"{get_column_letter(excel_col)}{r} =  INSERT EOL || ")

no_transitions = get_issues_with_no_transitions(issues)
print("\nIssues with no transitions:")
for status, issue_keys in no_transitions.items():
    print(f"{status}: {', '.join(issue_keys)}")
'''








changes_file = yaml_file.replace("scope.yaml","import.changes.txt")
print(f"Writing changes to {changes_file}")

if changes_list:
    with open(changes_file, "w") as f:
        for entry in changes_list:
            if "||None" in entry:
                entry = entry.replace("||None", "||")
            f.write(entry + "\n")
            print(entry)
    print(f"Changes written to {changes_file} with ({len(changes_list)} entries).")
else:
    print(f"No changes to write. Not need to create {changes_file}")
