import sys
import yaml
from jira import JIRA
import os
from dotenv import load_dotenv
import re
import ollama
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



# this is how this script is called from resync, note the paramter list
# run_and_log(["python", "-u", proc_aisummary_script, yaml_file, timestamp], log, f"proc_aisummary.py {yaml_file}, {timestamp}")


import json
LLMCONFIG_FILE = "../../../config/llmconfig.json"

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import markdown


from bs4 import BeautifulSoup

def html_to_text_with_structure(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    def walk(node, depth=0):
        parts = []
        for child in node.children:
            if child.name is None:  # NavigableString
                text = child.strip()
                if text:
                    # add indentation spaces based on depth
                    parts.append(" " * (depth * 2) + text)
            else:
                # Handle block-level tags with newlines
                if child.name in ("p", "div", "section", "article", "header", "footer",
                                  "ul", "ol", "li", "br", "h1", "h2", "h3", "h4", "h5", "h6"):
                    if child.name == "li":
                        parts.append(" " * (depth * 2) + "- " + walk(child, depth + 1).strip())
                    else:
                        inner = walk(child, depth + 1)
                        if inner:
                            parts.append(inner)
                    parts.append("\n")  # newline after block
                else:
                    # Inline tag (span, b, i, etc.)
                    inner = walk(child, depth)
                    if inner:
                        parts.append(inner)
        return "\n".join(p for p in parts if p.strip())

    text = walk(soup).strip()

    # Collapse excessive blank lines
    lines = [line.rstrip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line.strip() != "")

def send_markdown_email(subject: str, from_address: str, to_address: str, data: str):
    print(f"send_markdown_email called params= {subject}, {from_address}, {to_address}, {data[:20]}{'...' if len(data) > 20 else ''}")
    
    # Gmail SMTP settings
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_user = "fz96tw@gmail.com"                # your Gmail address
    smtp_password = "tgpmcbauhlligxvi"        # your 16-char app password

    #html_content = markdown.markdown(data)
    #html_content = markdown.markdown(data, extensions=[])  # no "sane_lists"
    html_content = data
    
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_address
    msg["To"] = to_address

    # Attach HTML only (simpler, avoids client showing raw markdown)
    msg.attach(MIMEText(html_content, "html"))

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login("fz96tw@gmail.com", "tgpmcbauhlligxvi")
        server.send_message(msg)

    print(f"Email sent to {to_address}")


def get_llm_model(llm_config_file):
    print("Current working directory:", os.getcwd())  # <-- debug
    if os.path.exists(llm_config_file):
        with open(llm_config_file, "r") as f:
            llm_model_set = json.load(f)
            m = llm_model_set.get("model")
            print(f"get_llm_model returning  {m}")
            return m
    else:
        print(f"ERROR: load_llm_config file {llm_config_file} was not found")
    
    return None


if len(sys.argv) < 5:
    print("Usage: python aibrief.py <url> <yaml_file> <timestamp> <userlogin> optional[delegated_auth]")
    sys.exit(1)

url = sys.argv[1]
yaml_file = sys.argv[2]
timestamp = sys.argv[3]
userlogin = sys.argv[4] if len(sys.argv) > 4 else None
delegated_auth = "--user_auth" in sys.argv




llm_model = get_llm_model(LLMCONFIG_FILE)
if not llm_model:
    llm_model = "Local"


with open(yaml_file, 'r') as f:
    data = yaml.safe_load(f)

fileinfo = data.get('fileinfo', {})
if not fileinfo:
    print("No fileinfo found in the YAML file.")
    sys.exit(1)
basename = fileinfo.get('basename')
tablename = fileinfo.get('table').replace(" ", "_") if fileinfo.get('table') else ""
source = fileinfo.get('source')
scope_file = fileinfo.get('scope file') # in this case this will match the yaml_file param passed to this script
sheet = fileinfo.get('sheet')

if not basename:
    print("No 'basename'found in fileinfo. Expecting 'basename' key.")
    sys.exit(1)

if not tablename:
    print("No 'table' found in fileinfo. Expecting 'table' key.")
    sys.exit(1)

# Determine if we will be INSERTING rows eventually vs just UPDATING existing rows in Excel/SharePoint
if "aibrief.scope.yaml" not in yaml_file.lower():
    print(f"ERROR: {yaml_file} is not a aibrief.scope.yaml file, exiting without action.")
    sys.exit(1)


print("Scope file:", scope_file )
print("Source file:", source)
print("basename:", basename)
print("Table,", tablename)
#print("Field indexes,", field_indexes_str)
#print("Field values,", field_values_str)


refer_tables = data.get('tables',[])
print(f"{yaml_file} contains refer_tables: {refer_tables}")

for ref_table in refer_tables:
    print(f"Checking refer_tables: {ref_table} for unwanted worksheet prefix...")
    # there may be a worksheet prefix with dotted notation 
    # need to remove worksheet prefix if both sheet and aibrief are on the same sheet
    if "." in ref_table:
        possible_sheet, possible_table = ref_table.split(".", 1)
        print(f"Found . in refer_tables, possible_sheet={possible_sheet}, possible_table={possible_table}")
        if possible_sheet == sheet:
            # remove the prefix
            refer_tables.remove(ref_table)
            refer_tables.append(possible_table)
            print(f"removed worksheet prefix '{possible_table}' leaving tablename={possible_table} in refer_tables")


email_list = data.get('email',[])
print(f"{yaml_file} contain email: {email_list}")

import requests
from requests.auth import HTTPBasicAuth


# Default to localhost unless overridden in env variable (set when in Docker)
SUMMARIZER_HOST = os.getenv("SUMMARIZER_HOST", "http://localhost:8000")

def get_summarized_comments(context, sysprompt):
    """
    Summarize comments for LLM processing.
    This function takes a list of comments in ascending order and returns a summarized version.
    """
    try:
        if not context:
            return "No context provided."

        # Only join if it's a list or tuple
        if isinstance(context, (list, tuple)):
            comments_str = "; ".join(context)
        else:
            comments_str = context  # already a string

        # comments_str was a single string before, but the service expects a list[str].
        # If you only have one string, wrap it in a list.
        #context = [comments_str]

        prompt = sysprompt + ".\n\n" + context
        #prompt = sysprompt + "\n\n" + "\n".join(context)

        prompt_list = [prompt]
        print(f"calling LLM with prompt = {prompt_list[0][:255]}...")

        if llm_model == "OpenAI":
            ENDPOINT = "/summarize_openai"
        else:
            ENDPOINT = "/summarize_local"

        #resp = requests.post("http://localhost:8000/summarize", json=prompt_list)
        resp = requests.post(f"{SUMMARIZER_HOST}{ENDPOINT}", json=prompt_list)

        if resp.status_code == 200:
            full_response = resp.json()["summary"]
        else:
            full_response = f"[ERROR] Service call failed: {resp.text}"    
        
        
        print(f"Full response: {full_response}")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return  full_response   
    except Exception as e:
        # Log the exception and return a safe default
        print(f"[EXCEPTION THROWN ERROR] get_summarized_comments failed: {e}")
        #return "[ERROR] Summary could not be generated."
        return f"[EXCEPTION THROWN ERROR] get_summarized_comments failed: {e}"
    
# Build LLM context from table rows and corresponding jiracsv file
# so yes, it combines excel table rows (all cells including non-jira ones)
# and combines with the aisummary.jira.csv file contents
def build_llm_context(table_rows, table_name, timestamp):
    
    context = ""
    # first put in the sheet contents for this table
    for r in table_rows:
        # r is a list of cell values, join them into a string
        context += " | ".join(str(cell) for cell in r) + "\n"

     # now append the jiracsv contents
    jiracsv_pattern = f"{basename}.{sheet}.{table_name}.{timestamp}.aisummary.jira.csv"
    dir_path = os.getcwd()  # or the folder where your CSVs live

    # case-insensitive search
    matched_file = None
    for f in os.listdir(dir_path):
        if f.lower() == jiracsv_pattern.lower():
            matched_file = os.path.join(dir_path, f)
            break

    if matched_file:
        with open(matched_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            context += line
    else:
        print(f"Warning: file not found (case-insensitive match) for {jiracsv_pattern}")

    return context



def check_missing_table_files(refer_tables, folder_path='.'):
    """
    Check which table names from refer_tables don't have corresponding .jira.csv files.
    
    Args:
        refer_tables: List of table names to check
        folder_path: Path to the folder to search (default is current directory)
    
    Returns:
        List of table names that don't have matching .jira.csv files
    """
    import os
    
    # Get all filenames in the folder that end with .jira.csv
    files = [f for f in os.listdir(folder_path) if f.endswith('.jira.csv')]
    
    # Track tables without matching files
    missing_tables = []
    
    # Check each table name
    for table in refer_tables:
        # Check if any .jira.csv file contains the table name
        found = any(table in filename for filename in files)
        
        if not found:
            print(f"Table '{table}' is missing corresponding .jira.csv file. Added to missing_tables={missing_tables}")
            missing_tables.append(table)
        else:
            print(f"Table '{table}' has corresponding .jira.csv file so no need to add to missing_tables={missing_tables}")
    
    return missing_tables


def get_unique_sheet_names(missing_tables):
    """
    Extract unique sheet names from missing table list.
    
    Args:
        missing_tables: List of table names in format 'sheet.tablename'
    
    Returns:
        List of unique sheet names
    """
    sheet_names = set()
    
    for table_name in missing_tables:
        if '.' in table_name:
            sheet = table_name.split('.', 1)[0]
            sheet_names.add(sheet)
        else:
            # If no dot, treat entire string as sheet name
            print(f"Warning: table name '{table_name}' does not contain sheet name")

    return list(sheet_names)


from openpyxl import load_workbook



from google_oauth import *

# don't be confused, googlelogin param is just userlogin param passed to identify which google token to use
# this var name is misleading and i need to fix it.
def read_google_rows(googlelogin, spreadsheet_url_or_id, sheet_name=None):
    """
    Reads all rows from a Google Sheet, returns as list of lists.
    sheet_name: name of the sheet; defaults to first sheet if None.
    """
    # Extract spreadsheet ID if URL is given
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", spreadsheet_url_or_id)
    spreadsheet_id = match.group(1) if match else spreadsheet_url_or_id.strip()

    # Load user's saved credentials
    creds = load_google_token(googlelogin)
    if not creds or not creds.valid:
        raise Exception(f"❌ User {googlelogin} not logged in to Google Drive")

    service = build("sheets", "v4", credentials=creds)

    # If no sheet_name provided, get the first sheet
    if sheet_name is None:
        metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_name = metadata["sheets"][0]["properties"]["title"]

    print(f"Reading Google Sheet ID={spreadsheet_id}, sheet='{sheet_name}' for user={googlelogin}")
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=sheet_name
    ).execute()

    return result.get("values", [])


def read_google_sheet_as_openpyxl(spreadsheet_url_or_id, sheet_name, userlogin="default"):
    """
    Read a Google Sheet and return an openpyxl-compatible worksheet object.
    Uses existing read_google_rows() function for authentication and data retrieval.
    
    Args:
        spreadsheet_url_or_id: Google Sheet ID or full URL
        sheet_name: Name of the worksheet/tab
        userlogin: Google credentials identifier (passed to read_google_rows)
    
    Returns:
        openpyxl Worksheet object
    """
    from openpyxl import Workbook
    
    # Use your existing function to get the data
    all_values = read_google_rows(userlogin, spreadsheet_url_or_id, sheet_name)
    
    # Create openpyxl workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Populate the worksheet with data from Google Sheets
    for row_idx, row_data in enumerate(all_values, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws



def extract_table_rows(source, basename, worksheet_name=None, tablename=None):
    """
    Extract table rows from a specific worksheet in an Excel workbook.
    
    Parameters:
    -----------
    basename : str
        Path to the Excel workbook file
    worksheet_name : str
        Name of the worksheet to process
    tablename : str, optional
        Specific table name to search for in <ai brief> tags
        
    Returns:
    --------
    dict
        Dictionary mapping table_name -> list of rows
    """

 
     # Detect if source is Google Sheet (simple heuristic: URL or just ID)
    is_google_sheet = isinstance(source, str) and ("docs.google.com/spreadsheets" in source or re.match(r"^[a-zA-Z0-9-_]{20,}$", source))

    if is_google_sheet:
        # For Google Sheets, assume userlogin is 'default' for now
        print(f"Reading Google Sheet: {basename}, sheet: {worksheet_name}")
        ws = read_google_sheet_as_openpyxl(source, worksheet_name, userlogin)

    else:
        wb = load_workbook(basename)
        ws = wb[worksheet_name] if worksheet_name else wb.active

    table_rows = {}     # maps table_name -> list of rows
    curr_table = None

    SPECIAL_TAGS = ("<jira>", "<ai brief>", "<rate assignee>", "<rate resolved>", "<cycletime>", "<statustime>")
    aibrief_cells = []   # list to store coordinates of <aibrief> cells

    # Iterate with cell objects, not just values
    for row in ws.iter_rows(values_only=False):
        row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]

        print(f"reading row = {row_values}")

        # skip the row if all cells are blank
        if all(v == "" for v in row_values):
            continue

        # check each cell for special tags (only if tablename is provided)
        if tablename:
            for cell in row:
                print(f"checking cell_value: {cell.value}")
                if cell.value and any(tag in str(cell.value) for tag in SPECIAL_TAGS):
                    if "<ai brief>" in str(cell.value):
                        print(f"found <ai brief> {cell.value} at coordinate={cell.coordinate} in row={row_values}")
                        print(f"checking if tablename: {tablename} present")
                        if tablename.replace("_", " ") in str(cell.value):
                            print(f"found target tablename")
                            # Get the cell one row below in the same column
                            below_cell = ws.cell(row=cell.row + 1, column=cell.column)

                            print(f"Cell below is at {below_cell.coordinate} with value: {below_cell.value}")

                            aibrief_cells.append({
                                "coordinate": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                                "below_coordinate": below_cell.coordinate,
                                "below_value": below_cell.value
                            })

                            break

        # find if this row contains any special tag
        tag_cells = [v for v in row_values if any(tag in v for tag in SPECIAL_TAGS)]

        if tag_cells:
            # close current table if we were collecting
            if curr_table:
                print(f"Finished collecting rows for table '{curr_table}'")

            # check if this is a <jira> start
            jira_cells = [v for v in tag_cells if "<jira>" in v]
            if jira_cells:
                #curr_table = jira_cells[0].rsplit("<jira>", 1)[0].rstrip().replace(" ", "_").lower()
                curr_table = jira_cells[0].rsplit("<jira>", 1)[0].rstrip().replace(" ", " ")  # keep case as-is
                print(f"curr_table set to {curr_table}")                
                print(f"checking worksheet_name={worksheet_name} against sheet={sheet}")
                if worksheet_name and worksheet_name != sheet:
                    # we are extracting from specific worksheet, prefix table name with sheet name
                    # because this will be lookup with with refer_tables which have sheet.tablename format
                    curr_table = f"{worksheet_name}.{curr_table}"
                    print(f"updating currtable to {curr_table} because sheet {worksheet_name} is different from active sheet {sheet}")

                table_rows[curr_table] = []   # start fresh list
                print(f"Found <jira> table '{curr_table}'")
            else:
                # other tag encountered: stop collecting
                curr_table = None

            continue  # skip storing this header row itself

        # if inside a <jira> table, add rows
        if curr_table:
            table_rows[curr_table].append(row_values)

    if curr_table:
        print(f"Finished collecting rows for table '{curr_table}'")

    return table_rows, aibrief_cells




# we pass tablename in this case to identify which <aibrief> tag to process
table_rows, aibrief_cells = extract_table_rows(source, basename, sheet, tablename)

# Debug: print collected <aibrief> positions
# not sure why this is needed but leaving it in for now.
# we will only have single aibrief tag per yaml_file processed i don't below is needed or makes sense
print("Found <aibrief> tags at:")
for cell in aibrief_cells:
    print(f" - {cell['coordinate']} (row {cell['row']}, col {cell['column']})")


#------------------------------------------------------------------
# Dump all tables and their rows
# -----------------------------------------------------------------

context = ""
aibrief_context = ""


# Refresh.py calls aibrief.py after processing all other tables in the seet.
# So make sure all the refer_tables have data files in this folder
# if not then assume the table is from a different sheet and
# we need to run resync on that sheet before proceeding.
# Once we have generated data files for other sheets we will
# also have to update table_rows by scanning those sheets too to
# collect the rows for those tables.



from refresh import *
from my_utils import *  

missing_tables = check_missing_table_files(refer_tables, os.getcwd())

sheets_to_resync = get_unique_sheet_names(missing_tables)


if sheets_to_resync:
    print(f"Warning: The following tables do not have matching .jira.csv files in the folder: {missing_tables}")
    #print("Please run resync on the corresponding sheets to generate the required data files before proceeding.")
    #sys.exit(1)
    for wsheet in sheets_to_resync:
        url = url + "#" + wsheet # append sheet fragment to URL because that's the format resync expects
        val= clean_sharepoint_url(url)  
        cwd = os.getcwd()

        config_dir_backup = CONFIG_DIR
        print(f"saving CONFIG_DIR value {CONFIG_DIR} before calling resync")
        print(f"Running resync({val},{userlogin},{cwd},{timestamp}) on sheet '{url}' to generate data for sheet '{wsheet}'")
        resync(val,userlogin, delegated_auth, cwd, timestamp)  # call your function with the string value file URL and userlogin (used for working folder for script)
        #CONFIG_DIR = config_dir_backup
        CONFIG_DIR = "../../../config"
        print(f"restoring CONFIG_DIR value to {CONFIG_DIR}")

        from google_oauth import *

        # After resync, extract table rows from this sheet and update table_rows
        print(f"Extracting table rows from sheet '{wsheet}' after resync")
        # in this case we do not care about the aibrief_cells returned because we already have them from the original sheet
        # _ variable is used to ignore the second return value
        new_table_rows, _ = extract_table_rows(source, basename, wsheet)
        if (new_table_rows):
            print(f"Extracted tables from sheet '{wsheet}': {list(new_table_rows.keys())}")
            table_rows.update(new_table_rows)
        else:
            print(f"Warning: No tables extracted from sheet '{wsheet}' after resync.")

# At this point we should have all the data from all the sheets needed to proceed.
# if some tables are still missing then there must be some other issue in resync and we can skip those files
# not fatal but are missing context data for LLM


print("\n=== Aggregating data from all the collected tables that are referred by this <aibrief> ===")
for table_name, rows in table_rows.items():
    print(f"\nTable: {table_name}")
    if (table_name in refer_tables):
        print(f"{table_name} is in refer_tables={refer_tables}, generating context")
        for r in rows:
            print("  ", r)
        context = build_llm_context(table_rows[table_name], table_name, timestamp)
        print(f"context generated = {context}")

        # Compose filename
        #filename = f"{source}.{sheet}.{tablename}.{table_name}.aisummary.context.txt"
        filename = f"{basename}.{table_name}.aisummary.context.txt"
        # Save context to file
        with open(filename, "w", encoding="utf-8") as f:
            f.write(context)
        print(f"Context saved to {filename}")

        aibrief_context += "\n\n" + context
    else:
        print(f"{table_name} is not referred in {yaml_file} -> {refer_tables}")

# Compose filename
filename = f"{basename}.{sheet}.{tablename}.aibrief.context.txt"
# Save context to file
with open(filename, "w", encoding="utf-8") as f:
    f.write(aibrief_context)

print(f"<aibrief> Context saved to {filename}")

sysprompt = f"The following text is a csv data separated by | character. Refer to this project as Project {tablename}.  Read all of it and briefly as possible in the form of project status report for executive summary. highlight all milestones,  risks or blocking issues. Also add a paragraph of titled 'Executive Summary' at the top of your response with very brief business executive summary. Provide your summary in HTML format. Convert an items that have corresponding Jira id into URL link for that jira id listed below"

if isinstance(aibrief_context, list):
    aibrief_context = "\n".join(aibrief_context)

print(f"Calling get_summarized_comments with aibrief_context={aibrief_context[:255]}... and sysprompt={sysprompt[:255]}...")
report = get_summarized_comments(aibrief_context, sysprompt)

# Compose filename
#filename = f"{os.path.splitext(source)[0]}.{tablename}.llm.txt"
filename = f"{basename}.{sheet}.{tablename}.aibrief.llm.txt"
# Save context to file
with open(filename, "w", encoding="utf-8") as f:
    f.write(report)
    print(f"LLM reponse saved to {filename}")

# Replace all newlines with semicolons
cleaned_response = html_to_text_with_structure(report)
print(f"html_to_text_with_structure returned cleaned_response={cleaned_response}")
cleaned_response = cleaned_response.rstrip("\n")
cleaned_response = cleaned_response.replace("\n", "; ")
cleaned_response = cleaned_response.replace("|", "^")

if aibrief_cells:
    #changes = f"{aibrief_cells[0]["coordinate"]} = {report} || "
    first_abrief =  aibrief_cells[0]
    below_cell = first_abrief["below_coordinate"]

    changes = f"{below_cell} = {timestamp}:{cleaned_response} || "
    #changes = f"{below_coordinate} = {timestamp}:{cleaned_response} || "
    #filename = f"{os.path.splitext(source)[0]}.{tablename}.changes.txt"
    filename = f"{basename}.{sheet}.{tablename}.aibrief.changes.txt"
    # Save context to file
    with open(filename, "w", encoding="utf-8") as f:
        f.write(changes)
    print(f"changes written to {filename}")

    for email_id in email_list:
        print(f"Sending email to {email_id}")
        send_markdown_email(f"AI Connector update {basename}.{sheet}.{tablename}","fz96tw@gmail.com", email_id, report )

else:
    print(f"aibrief_cells is empty so no changes.txt to write")


