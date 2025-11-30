import sys
import yaml
from jira import JIRA
import os
from dotenv import load_dotenv
import re
import ollama
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Cache dictionary to avoid repeated calls
user_cache = {}

import requests
from requests.auth import HTTPBasicAuth


# Cache dictionary to avoid repeated calls
user_cache = {}



def get_user_display_name(account_id):
    if account_id in user_cache:
        return user_cache[account_id]

    url = f"{JIRA_URL}/rest/api/3/user?accountId={account_id}"
    response = requests.get(url, auth=HTTPBasicAuth(JIRA_EMAIL, JIRA_API_TOKEN))

    if response.status_code == 200:
        user_data = response.json()
        display_name = user_data.get("displayName", "unknown")
    else:
        display_name = "unknown"

    user_cache[account_id] = display_name
    return display_name

def get_user_display_name2(account_id, jira_client):
    if account_id in user_cache:
        return user_cache[account_id]

    # Search users with account_id as query
    users = jira_client.search_users(query=account_id, maxResults=1)

    if users:
        # Check if the first user matches the accountId exactly
        user = users[0]
        if user.accountId == account_id:
            display_name = user.displayName
        else:
            display_name = "unknown"
    else:
        display_name = "unknown"

    user_cache[account_id] = display_name
    return display_name


def replace_account_ids_with_names(text):
    def replacer(match):
        account_id = match.group(1)
        display_name = get_user_display_name(account_id)
        return f"@{display_name}"
    return re.sub(r"\[~accountid:([a-zA-Z0-9:\-]+)\]", replacer, text)


def replace_account_ids_with_names2(text, jira_client):
    import re
    def replacer(match):
        account_id = match.group(1)
        display_name = get_user_display_name(account_id, jira_client)
        return f"@{display_name}"
    return re.sub(r"\[~accountid:([a-zA-Z0-9:\-]+)\]", replacer, text)


import re

def move_brackets_to_front(lines):
    """
    Moves [JIRA-KEY] from end of string to the front.

    Args:
        lines (list[str]): List of strings like "Task name [KEY-1]"

    Returns:
        list[str]: Reformatted strings like "[KEY-1] Task name"
    """
    result = []
    
    pattern = re.compile(r'^(.*?)\s+(?:▫️\s*)?(\[[^\]]+\])$')



    for line in lines:
        match = pattern.match(line.strip())
        if match:
            text, ticket = match.groups()
            result.append(f"{ticket} {text}")
        else:
            result.append(line)  # leave unchanged if no match
    return result


user_cache = {}

def get_account_id(jira_client, username_or_email):
    """
    Look up the Jira accountId for a given username/email.
    Caches results to avoid repeated lookups.
    """
    if not username_or_email:
        return None

    if username_or_email in user_cache:
        return user_cache[username_or_email]

    try:
        users = jira_client.search_users(query=username_or_email, maxResults=1)
        if users:
            account_id = users[0].accountId
            user_cache[username_or_email] = account_id
            return account_id
    except Exception as e:
        print(f"⚠️ Failed to find accountId for {username_or_email}: {e}")

    return None

if len(sys.argv) < 4:
    print("Usage: python create_jira.py <yaml_file> <xlsx file>")
    sys.exit(1)

yaml_file = sys.argv[1]
filename = sys.argv[2]
timestamp = sys.argv[3]

with open(yaml_file, 'r') as f:
    data = yaml.safe_load(f)

fileinfo = data.get('fileinfo', {})
if not fileinfo:
    print("No fileinfo found in the YAML file.")
    sys.exit(1)
basename = fileinfo.get('basename')
tablename = fileinfo.get('table').replace(" ", "_") if fileinfo.get('table') else ""
source = fileinfo.get('source')
scope_file = fileinfo.get('scope file')

if not basename:
    print("No 'basename'found in fileinfo. Expecting 'basename' key.")
    sys.exit(1)

if not tablename:
    print("No 'table' found in fileinfo. Expecting 'table' key.")
    sys.exit(1)

# I don't think this code is doing anything useful. It always sets create_mode to True?! since 
# Determine if we will be INSERTING rows eventually vs just UPDATING existing rows in Excel/SharePoint
if "create" in yaml_file.lower():
    print(f"CREATE mode detected based on filename: {yaml_file}.")
    # You can set a flag or handle import-specific logic here if needed
    create_mode = True
    output_file = basename + "." + tablename + timestamp + ".create.jira.csv"
else:
    create_mode = False
    output_file = basename + "." + tablename + timestamp + ".create.jira.csv"

fields = data.get('fields', [])
# Convert list of dicts into a dictionary
fields_dict = {field["value"]: field.get("index", "<blank>") for field in fields}

field_values = [field.get('value') for field in fields if 'value' in field]
field_indexes = [field.get('index') for field in fields if 'index' in field]
field_values_str = ','.join(field_values)
field_indexes_str = ','.join(map(str, field_indexes))

print("Source file,", source)
print("basename,", basename)
print("Table,", tablename)
print("Field indexes,", field_indexes_str)
print("Field values,", field_values_str)

with open(output_file, "w") as outfile:
    outfile.write("Source file," + source + "\n")
    outfile.write("Basename," + basename + "\n")
    outfile.write("Scope file," + scope_file + "\n")    
    outfile.write("Table," + tablename + "\n")
    outfile.write("Field indexes," + field_indexes_str + "\n")
    outfile.write("Field values," + field_values_str + "\n")

#jira_ids = data.get('jira_ids', [])
jira_create_row = data.get('jira_create_rows',[])


# Replace with your Jira Cloud credentials and URL

# Load environment variables from a .env file if present
#load_dotenv()
#ENV_PATH = "../../../config/.env"
#load_dotenv(dotenv_path=ENV_PATH)

# load .env from config folder
ENV_PATH = f"../../../config/env.{userlogin}"
load_dotenv(dotenv_path=ENV_PATH)


JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN")
JIRA_URL = os.environ.get("JIRA_URL")
JIRA_EMAIL = os.environ.get("JIRA_EMAIL")

if not JIRA_API_TOKEN:
    print("Error: JIRA_API_TOKEN environment variable not set.")
    sys.exit(1)


# Connect to Jira with basic auth
try:
    jira = JIRA(server=JIRA_URL, basic_auth=(JIRA_EMAIL, JIRA_API_TOKEN))
    #print("✅ Successfully connected to Jira.")
except Exception as e:
    print(f"❌ Failed to connect to Jira: {e}")
    sys.exit(1)

#issues = []  # global issues list to hold results from both ID and JQL searches


project_key = ""
summary = ""
description = ""
issuetype = ""
labels = []
assignee = ""
priority = ""
epic = ""   # if the jira needs to be child of existing Epic (must exist apriori!)
epic_name_field = ""    # contains string only if Epic is being created otherwise empty

changes_list = []

for line in jira_create_row:
    if not line:
        print("skipping blank line")
        continue  # skip blank lines
    parts = [p.strip() for p in line.split(',')]
    print(f"Processing line: {line}")

    if len(parts) != len(fields_dict):
        print(f"Warning: Mismatched field count in {filename}. Make sure number of values provided matches number of fields.")
        continue

    # dictionary that maps field names to their corresponding values that we just read from the line
    # assumes the order of parts matches the order of field_names (in the jira.csv file)
    record = dict(zip(field_values, parts))
    print(f"record = {record}")
    
    key = record.get("key")
    if not key:
        print("Warning: No <key> field found in csv line. Check that the excel file table has a <key> column")
        continue

    for field in record:
        record[field] = record[field].replace("<blank>","")
        if "project" in field:
            project_key = (record.get("project") or "").strip().upper()
        elif "summary" in field:
            summary = record["summary"]
        elif "description" in field:
            description = "Created by AI Connector behalf of requestor:" + record["requestor"] + ". " + record["description"]
        elif "epic" in field:
            epic = record["epic"]
            epic = (record.get("epic") or "").strip().upper()
        elif "issuetype" in field:
            #issuetype = record["issuetype"]
            issuetype = record.get("issuetype", "").strip().capitalize() if record.get("issuetype") else ""
            '''
            if "epic" in issuetype.lower()
                print("Epic issuetype to be created")
                # Find out the field ID for Epic Name (commonly customfield_10011)
                # You can inspect jira.fields() to see all fields
                fields_all = jira.fields()
                epic_name_field = None
                for f in fields_all:
                    if f['name'] == 'Epic Name':
                        epic_name_field = f['id']
                        continue
            '''

        elif "priority" in field:
            priority = record.get("priority", "").strip().capitalize() if record.get("priority") else ""
        elif "assignee" in field:
            assignee = record["assignee"]
        elif "labels" in field:
            labels = record["labels"]
            labels_list = [x.strip() for x in labels.split("|") if x.strip()]  # need to make it into list
        else:
            print(f"unable to process this field '{field}' no matching if condition")

    print(f"project_key = {project_key}")
    print(f"summary = {summary}")
    print(f"description = {description}")
    print(f"issuetype = {issuetype}")
    print(f"priority = {priority}")

    # make sure we have minimum fields need to create jira
    if all([project_key, summary, description, issuetype, priority]):
        issue_fields = {
            "project": {"key": project_key},
            "summary": summary,
            "description": description,
            "issuetype": {"name": issuetype},
            "priority": {"name": priority},
        }

        if assignee:
            account_id = get_account_id(jira, assignee)
            if account_id:
                issue_fields["assignee"] = {"id": account_id}
            else:
                print(f"⚠️ Warning: Could not resolve accountId for assignee '{assignee}'")

        if labels:
            issue_fields["labels"] = labels_list  # must be a list of strings

        print(f"constructed issue_field: {issue_fields}")

        # check if this issue already exists
        jql = f'project = "{project_key}" AND summary ~ "{summary}" AND issuetype = {issuetype}'
        print(f"jql = {jql}")
        issues = jira.search_issues(jql, maxResults=10)
        print(f"jira search for duplicate returned {len(issues)} records.")
        exact_matches = [i for i in issues if i.fields.summary == summary and len(i.fields.summary) == len(summary)]
        print(f"Potential exact_matches {len(exact_matches)}.")
        for i in exact_matches:
            print(f"       possibly match: {i}")


    

        print(f"checking if duplicate exists for '{project_key}:{issuetype}:{summary}")
        if exact_matches:
            issue = exact_matches[0]
            #issue = exact_matches[0]
            #issue.update(fields=issue_fields)
            print(f"duplicate JIRA found '{issue.key}:{issue.fields.issuetype.name}:{issue.fields.summary}' already exists, will not create.")
        else:
            try:

                if "epic" in issuetype.lower() and len(epic_name_field):
                    print("Epic issuetype to be created")
                    #issue_fields[epic_name_field] = summary  # or another string for Epic Name

                if len(epic):
                    print(f"Jira is to have a parent Epic {epic}")
                    issue_fields["parent"] = {"key": epic}  # Replace with your Epic key

                print(f"About to call jira.create_issue ")
                issue = jira.create_issue(fields=issue_fields)
                print(f"✅ Created issue: {issue.key}")
                    
            except Exception as e:
                print(f"❌ Failed to create issue: {e}")
                # pass on the error in the key field
                # create a dummy object to hold the error
                class DummyIssue:
                    def __init__(self, error):
                        self.key = f"ERROR: {error}"
                issue = DummyIssue(e)


                    # update changes_list here!
        if "key" in fields_dict:
            print("Found <key> field so will write to changes.txt file")
            row_num = record["row"]
            col_num = fields_dict["key"]
            # If col_num is 0-based (Python index), add +1
            coord = f"{get_column_letter(col_num + 1)}{row_num}"
            print (f"row_num = {row_num}   col_num = {col_num}")
            print (f"Adding to Changes.txt    {coord} = {issue.key} || ")
            changes_list.append(f"{coord} = {issue.key} || ")

        if "url" in fields_dict:
            print("Found <key> field so will write to changes.txt file")
            row_num = record["row"]
            col_num = fields_dict["url"]
            # If col_num is 0-based (Python index), add +1
            coord = f"{get_column_letter(col_num + 1)}{row_num}"
            print (f"row_num = {row_num}   col_num = {col_num}")
            print (f"Adding to Changes.txt    {coord} = {issue.key} || ")
            changes_list.append(f"{coord} = URL {issue.key} || ")


        if "timestamp" in fields_dict:
            row_num = record["row"]
            col_num = fields_dict["timestamp"]
            coord = f"{get_column_letter(col_num + 1)}{row_num}"
            currtime =  datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print (f"row_num = {row_num}   col_num = {col_num}  excel = {coord}")
            print (f"Adding to Changes.txt    {coord} = {currtime} || ")
            changes_list.append(f"{coord} = {currtime} || ")
    else:
        print("Skipping Jira creation, required field(s) is missing")



if not changes_list:
    print("No changes made.")
else:
    changes_file = yaml_file.replace("scope.yaml","changes.txt")
    
    print(f"Writing changes to {changes_file}")
    with open(changes_file, "w") as f:
        for entry in changes_list:
            if "||None" in entry:
                entry = entry.replace("||None", "||")
            f.write(entry + "\n")
            print(entry)
    print(f"Changes written to {changes_file} with ({len(changes_list)} entries).")

print(f"Data written to {output_file}")
print(f"CSV_CREATED:{output_file}")