from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Any, Tuple
import statistics

def calculate_average_status_transition_time(jira_issues: List[Any]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """
    Same as before, but now includes an artificial "Created" status to represent
    the time from issue creation to its first status transition.
    """

    transition_durations = defaultdict(list)
    transition_issues = defaultdict(list)

    print(f"Calculating status transition times for {len(jira_issues)} issues...")

    for issue in jira_issues:
        # --- Extract issue info and changelog ---
        if hasattr(issue, 'key'):
            issue_key = issue.key
            changelog = getattr(issue, 'changelog', None)
            creation_time = None
            if hasattr(issue, 'fields') and hasattr(issue.fields, 'created'):
                try:
                    creation_time = parse_jira_timestamp(issue.fields.created)
                except ValueError:
                    pass
        else:
            issue_key = issue.get('key', 'Unknown')
            changelog = issue.get('changelog')
            creation_time = None
            fields = issue.get('fields', {})
            if 'created' in fields:
                try:
                    creation_time = parse_jira_timestamp(fields['created'])
                except ValueError:
                    pass

        if not changelog:
            continue

        # --- Extract histories ---
        histories = getattr(changelog, 'histories', None)
        if isinstance(changelog, dict):
            histories = changelog.get('histories', histories)
        if not histories:
            continue

        # --- Extract status changes ---
        status_changes = []
        for history in sorted(histories, key=lambda h: getattr(h, 'created', '') if hasattr(h, 'created') else h.get('created', '')):
            created_str = getattr(history, 'created', None) or (history.get('created') if isinstance(history, dict) else None)
            if not created_str:
                continue

            try:
                timestamp = parse_jira_timestamp(created_str)
            except ValueError:
                continue

            items = getattr(history, 'items', None) or (history.get('items') if isinstance(history, dict) else [])
            for item in items:
                field = getattr(item, 'field', None) if hasattr(item, 'field') else item.get('field', None)
                if field != 'status':
                    continue
                from_status = getattr(item, 'fromString', None) if hasattr(item, 'fromString') else item.get('fromString', None)
                to_status = getattr(item, 'toString', None) if hasattr(item, 'toString') else item.get('toString', None)
                if from_status and to_status:
                    status_changes.append({'timestamp': timestamp, 'from_status': from_status, 'to_status': to_status})

        # --- Add artificial "Created" → first transition ---
        if creation_time and status_changes:
            first_real_status = status_changes[0]['from_status']
            duration_hours = (status_changes[0]['timestamp'] - creation_time).total_seconds() / 3600.0
            if duration_hours >= 0:
                transition_key = ('Created', first_real_status)
                transition_durations[transition_key].append(duration_hours)
                transition_issues[transition_key].append({
                    'issue_key': issue_key,
                    'duration_hours': round(duration_hours, 6),
                    'start_time': creation_time,
                    'end_time': status_changes[0]['timestamp']
                })

        # --- Standard per-transition durations ---
        for i, change in enumerate(status_changes):
            from_status = change['from_status']
            to_status = change['to_status']
            previous_timestamp = creation_time if i == 0 and creation_time else status_changes[i - 1]['timestamp']
            duration_hours = (change['timestamp'] - previous_timestamp).total_seconds() / 3600.0
            if duration_hours < 0:
                continue

            transition_key = (from_status, to_status)
            transition_durations[transition_key].append(duration_hours)
            transition_issues[transition_key].append({
                'issue_key': issue_key,
                'duration_hours': round(duration_hours, 6),
                'start_time': previous_timestamp,
                'end_time': change['timestamp']
            })

    # --- Compile results ---
    results = {}
    for (from_status, to_status), durations in transition_durations.items():
        if not durations:
            continue
        avg_hours = statistics.mean(durations)
        avg_days = avg_hours / 24
        stddev_hours = statistics.stdev(durations) if len(durations) > 1 else 0.0

        results[(from_status, to_status)] = {
            'average_hours': round(avg_hours, 6),
            'average_minutes': round(avg_hours * 60, 2),
            'average_days': round(avg_days, 2),
            'median_hours': round(statistics.median(durations), 6),
            'min_hours': round(min(durations), 6),
            'max_hours': round(max(durations), 6),
            'stddev_hours': round(stddev_hours, 6),
            'count': len(durations),
            'durations': durations,
            'issues': transition_issues[(from_status, to_status)]
        }

    print(f"Calculated {len(results)} unique status transitions (including 'Created').")
    return results

from collections import defaultdict
from datetime import datetime
from typing import List, Any, Dict
import statistics


from collections import defaultdict
from datetime import datetime, timezone
from typing import List, Any, Dict
import statistics


def calculate_average_chain_cycle_time(jira_issues: List[Any]) -> Dict[str, Dict[str, Any]]:
    """
    Calculate total and average cycle time for all unique end-to-end transition chains,
    starting from an artificial 'Created' status.
    Includes issues with no status transitions, treating them as 'Created' → <current status>.
    """

    chain_durations = defaultdict(list)
    chain_issues = defaultdict(list)

    now_utc = datetime.now(timezone.utc)
    print(f"Calculating end-to-end transition chain cycle times (including 'Created') for {len(jira_issues)} issues...")

    for issue in jira_issues:
        # --- Extract info ---
        if hasattr(issue, 'key'):
            issue_key = issue.key
            changelog = getattr(issue, 'changelog', None)
            fields = getattr(issue, 'fields', None)
            creation_time = getattr(fields, 'created', None) if fields else None
            current_status = getattr(fields.status, 'name', 'Unknown') if fields and hasattr(fields, 'status') else 'Unknown'
            if creation_time:
                creation_time = parse_jira_timestamp(creation_time)
        else:
            issue_key = issue.get('key', 'Unknown')
            changelog = issue.get('changelog')
            fields = issue.get('fields', {})
            creation_time = None
            current_status = fields.get('status', {}).get('name', 'Unknown')
            if 'created' in fields:
                try:
                    creation_time = parse_jira_timestamp(fields['created'])
                except ValueError:
                    pass

        print(f"Parsed creation time for issue {issue_key}: {creation_time}")
        print(f"Current status for issue {issue_key}: {current_status}")

        # --- Handle issues with no changelog or histories ---
        histories = None
        if changelog:
            histories = getattr(changelog, 'histories', None)
            if isinstance(changelog, dict):
                histories = changelog.get('histories', histories)

        status_changes = []
        if histories:
            # --- Extract status changes ---
            for history in sorted(histories, key=lambda h: getattr(h, 'created', '') if hasattr(h, 'created') else h.get('created', '')):
                created_str = getattr(history, 'created', None) or (history.get('created') if isinstance(history, dict) else None)
                if not created_str:
                    continue

                try:
                    timestamp = parse_jira_timestamp(created_str)
                except Exception:
                    continue

                items = getattr(history, 'items', None) or (history.get('items') if isinstance(history, dict) else [])
                for item in items:
                    field = getattr(item, 'field', None) if hasattr(item, 'field') else item.get('field', None)
                    if field != 'status':
                        continue
                    from_status = getattr(item, 'fromString', None) if hasattr(item, 'fromString') else item.get('fromString', None)
                    to_status = getattr(item, 'toString', None) if hasattr(item, 'toString') else item.get('toString', None)
                    if from_status and to_status:
                        status_changes.append((from_status, to_status, timestamp))

        # --- Sort by timestamp ---
        status_changes.sort(key=lambda x: x[2]) if status_changes else None

        # --- Insert artificial "Created" status ---
        if creation_time:
            if status_changes:
                first_status = status_changes[0][0]
                status_changes.insert(0, ("Created", first_status, creation_time))
            else:
                # No status changes — create synthetic chain
                status_changes = [("Created", current_status, creation_time)]
            print(f"Inserted 'Created' status for issue {issue_key} at {creation_time}.")
        else:
            print(f"No creation time for issue {issue_key}, skipping.")
            continue

        # --- Determine end time ---
        if len(status_changes) > 1:
            start_time = status_changes[0][2]
            end_time = status_changes[-1][2]
        else:
            # No transitions — measure up to now
            start_time = creation_time
            end_time = now_utc

        chain = " → ".join([s[0] for s in status_changes] + [status_changes[-1][1]])
        duration_hours = (end_time - start_time).total_seconds() / 3600.0

        if duration_hours < 0:
            print(f"Negative duration for issue {issue_key} in chain {chain}, skipping.")
            continue

        chain_durations[chain].append(duration_hours)
        chain_issues[chain].append({
            'issue_key': issue_key,
            'duration_hours': round(duration_hours, 6),
            'start_time': start_time,
            'end_time': end_time
        })

        print(f"Issue {issue_key} chain: {chain} duration: {duration_hours:.2f} hours")

    # --- Compile results ---
    results = {}
    for chain, durations in chain_durations.items():
        if not durations:
            continue
        avg_hours = statistics.mean(durations)
        stddev_hours = statistics.stdev(durations) if len(durations) > 1 else 0.0
        results[chain] = {
            'average_hours': round(avg_hours, 6),
            'average_days': round(avg_hours / 24, 3),
            'median_hours': round(statistics.median(durations), 6),
            'min_hours': round(min(durations), 6),
            'max_hours': round(max(durations), 6),
            'stddev_hours': round(stddev_hours, 6),
            'count': len(durations),
            'durations': durations,
            'issues': chain_issues[chain]
        }

    print(f"Calculated {len(results)} unique transition chains (including 'Created' and synthetic ones).")
    return results

def calculate_average_chain_cycle_time_old(jira_issues: List[Any]) -> Dict[str, Dict[str, Any]]:
    """
    Calculate total and average cycle time for all unique end-to-end transition chains,
    starting from an artificial 'Created' status.
    """

    chain_durations = defaultdict(list)
    chain_issues = defaultdict(list)

    print(f"Calculating end-to-end transition chain cycle times (including 'Created') for {len(jira_issues)} issues...")

    for issue in jira_issues:
        # --- Extract info ---
        if hasattr(issue, 'key'):
            issue_key = issue.key
            changelog = getattr(issue, 'changelog', None)
            creation_time = getattr(issue.fields, 'created', None) if hasattr(issue, 'fields') else None
            if creation_time:
                creation_time = parse_jira_timestamp(creation_time)
        else:
            issue_key = issue.get('key', 'Unknown')
            changelog = issue.get('changelog')
            creation_time = None
            fields = issue.get('fields', {})
            if 'created' in fields:
                try:
                    creation_time = parse_jira_timestamp(fields['created'])
                except ValueError:
                    pass

        print(f"Parsed creation time for issue {issue_key}: {creation_time}")

        if not changelog:
            print(f"No changelog for issue {issue_key}, skipping.")
            continue

        # --- Get histories ---
        histories = getattr(changelog, 'histories', None)
        if isinstance(changelog, dict):
            histories = changelog.get('histories', histories)
        if not histories:
            print(f"No histories in changelog for issue {issue_key}, skipping.")
            continue

        # --- Extract status changes ---
        status_changes = []
        for history in sorted(histories, key=lambda h: getattr(h, 'created', '') if hasattr(h, 'created') else h.get('created', '')):
            created_str = getattr(history, 'created', None) or (history.get('created') if isinstance(history, dict) else None)
            if not created_str:
                print(f"No created timestamp in history for issue {issue_key}, skipping this history.")
                continue

            try:
                timestamp = parse_jira_timestamp(created_str)
            except Exception:
                print(f"Failed to parse timestamp '{created_str}' for issue {issue_key}, skipping this history.")
                continue

            items = getattr(history, 'items', None) or (history.get('items') if isinstance(history, dict) else [])
            for item in items:
                field = getattr(item, 'field', None) if hasattr(item, 'field') else item.get('field', None)
                if field != 'status':
                    continue
                from_status = getattr(item, 'fromString', None) if hasattr(item, 'fromString') else item.get('fromString', None)
                to_status = getattr(item, 'toString', None) if hasattr(item, 'toString') else item.get('toString', None)
                if from_status and to_status:
                    print(f"Issue {issue_key} status change: {from_status} → {to_status} at {timestamp}")
                    status_changes.append((from_status, to_status, timestamp))

        if not status_changes:
            print(f"No status_changes found for issue {issue_key}, skipping.")
            continue

        # --- Sort by timestamp ---
        status_changes.sort(key=lambda x: x[2])

        # --- Insert artificial "Created" → first status ---
        if creation_time:
            first_status = status_changes[0][0]
            status_changes.insert(0, ("Created", first_status, creation_time))
            print(f"Inserted 'Created' status for issue {issue_key} at {creation_time}.")

        # --- Build full chain and compute total duration ---
        chain = " → ".join([s[0] for s in status_changes] + [status_changes[-1][1]])
        start_time = status_changes[0][2]
        end_time = status_changes[-1][2]
        duration_hours = (end_time - start_time).total_seconds() / 3600.0

        if duration_hours < 0:
            print(f"Negative duration for issue {issue_key} in chain {chain}, skipping (likely timestamp order issue).")
            continue

        chain_durations[chain].append(duration_hours)
        chain_issues[chain].append({
            'issue_key': issue_key,
            'duration_hours': round(duration_hours, 6),
            'start_time': start_time,
            'end_time': end_time
        })

        print(f"Issue {issue_key} chain: {chain} duration: {duration_hours:.2f} hours")

    # --- Compile results ---
    results = {}
    for chain, durations in chain_durations.items():
        if not durations:
            continue
        avg_hours = statistics.mean(durations)
        stddev_hours = statistics.stdev(durations) if len(durations) > 1 else 0.0
        results[chain] = {
            'average_hours': round(avg_hours, 6),
            'average_days': round(avg_hours / 24, 3),
            'median_hours': round(statistics.median(durations), 6),
            'min_hours': round(min(durations), 6),
            'max_hours': round(max(durations), 6),
            'stddev_hours': round(stddev_hours, 6),
            'count': len(durations),
            'durations': durations,
            'issues': chain_issues[chain]
        }

    print(f"Calculated {len(results)} unique transition chains (including 'Created').")
    return results





from datetime import datetime, timezone
from dateutil import parser
from collections import defaultdict
import statistics
from typing import List, Any, Dict

def parse_jira_timestamp(ts_str: str) -> datetime:
    """Parse a Jira timestamp into a timezone-aware UTC datetime."""
    dt = parser.isoparse(ts_str)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt


def parse_jira_timestamp_old(timestamp_str: str) -> datetime:
    import re
    from datetime import datetime, timezone, timedelta

    ts = timestamp_str.strip()

    # Convert "-0400" → "-04:00" to make it ISO-compatible
    tz_pattern = r'([+-])(\d{2})(\d{2})$'
    match = re.search(tz_pattern, ts)
    if match:
        sign, hours, minutes = match.groups()
        ts = re.sub(tz_pattern, f'{sign}{hours}:{minutes}', ts)

    formats = [
        '%Y-%m-%dT%H:%M:%S.%f%z',  # e.g. 2025-09-02T23:31:32.138-04:00
        '%Y-%m-%dT%H:%M:%S%z',     # e.g. 2025-09-02T23:31:32-04:00
        '%Y-%m-%dT%H:%M:%S.%f',    # no timezone, with milliseconds
        '%Y-%m-%dT%H:%M:%S',       # no timezone, no milliseconds
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(ts, fmt)
            # If it has tzinfo, convert to UTC and drop tz awareness
            if dt.tzinfo:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt
        except ValueError:
            continue

    raise ValueError(f"Unable to parse timestamp: {timestamp_str}")


def parse_jira_timestamp_old(timestamp_str: str) -> datetime:
    import re
    from datetime import timezone
    
    ts = timestamp_str.strip()
    
    # Convert -0400 to -04:00 format
    tz_pattern = r'([+-])(\d{2})(\d{2})$'
    match = re.search(tz_pattern, ts)
    if match:
        sign, hours, minutes = match.groups()
        ts = re.sub(tz_pattern, f'{sign}{hours}:{minutes}', ts)
    
    formats = [
        '%Y-%m-%dT%H:%M:%S.%f%z',  # 2025-09-02T23:31:32.138-04:00
        '%Y-%m-%dT%H:%M:%S%z',     # 2025-09-02T23:31:32-04:00
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(ts, fmt)
        except ValueError:
            continue
    
    raise ValueError(f"Unable to parse timestamp: {timestamp_str}")


def print_transition_report(transition_data: Dict[Tuple[str, str], Dict[str, Any]]) -> None:
    """
    Print a formatted report of status transition times.
    """
    if not transition_data:
        print("No status transitions found in the provided issues.")
        return
    
    print("JIRA Status Transition Time Analysis")
    print("=" * 50)
    
    # Sort transitions by average time (descending)
    sorted_transitions = sorted(
        transition_data.items(),
        key=lambda x: x[1]['average_hours'],
        reverse=True
    )
    
    for (from_status, to_status), data in sorted_transitions:
        print(f"\n{from_status} → {to_status}")
        print(f"  Average Time: {data['average_hours']:.1f} hours ({data['average_days']:.1f} days)")
        print(f"  Median Time:  {data['median_hours']:.1f} hours")
        print(f"  Range:        {data['min_hours']:.1f} - {data['max_hours']:.1f} hours")
        print(f"  Sample Size:  {data['count']} transitions")


from collections import defaultdict
from typing import List, Any, Dict

def get_issues_with_no_transitions(jira_issues: List[Any]) -> Dict[str, List[str]]:
    """
    Identify JIRA issues that have no status transitions in their changelog,
    grouped by their current status.

    Returns:
        Dict[status_name, [issue_keys]]
    """
    no_transition_statuses = defaultdict(list)

    for issue in jira_issues:
        # --- Extract issue key safely ---
        if hasattr(issue, "key"):
            issue_key = issue.key
        elif isinstance(issue, dict):
            issue_key = issue.get("key", "Unknown")
        else:
            issue_key = "Unknown"

        # --- Extract changelog safely ---
        if hasattr(issue, "changelog"):
            changelog = issue.changelog
        elif isinstance(issue, dict):
            changelog = issue.get("changelog")
        else:
            changelog = None

        # --- If no changelog at all ---
        if not changelog:
            current_status = None
            if hasattr(issue, "fields") and hasattr(issue.fields, "status"):
                current_status = getattr(issue.fields.status, "name", None)
            elif isinstance(issue, dict):
                status_obj = issue.get("fields", {}).get("status")
                if isinstance(status_obj, dict):
                    current_status = status_obj.get("name")
            no_transition_statuses[current_status or "Unknown"].append(issue_key)
            continue

        # --- Extract histories safely ---
        histories = None
        if hasattr(changelog, "histories"):
            histories = changelog.histories
        elif isinstance(changelog, dict):
            histories = changelog.get("histories", [])
        if not histories:
            current_status = None
            if hasattr(issue, "fields") and hasattr(issue.fields, "status"):
                current_status = getattr(issue.fields.status, "name", None)
            elif isinstance(issue, dict):
                status_obj = issue.get("fields", {}).get("status")
                if isinstance(status_obj, dict):
                    current_status = status_obj.get("name")
            no_transition_statuses[current_status or "Unknown"].append(issue_key)
            continue

        # --- Scan for status changes ---
        has_status_change = False
        for history in histories:
            items = getattr(history, "items", None) or (history.get("items") if isinstance(history, dict) else [])
            for item in items:
                field_name = getattr(item, "field", None) if hasattr(item, "field") else item.get("field", None)
                if field_name == "status":
                    has_status_change = True
                    break
            if has_status_change:
                break

        # --- If no status change found, group issue by current status ---
        if not has_status_change:
            current_status = None
            if hasattr(issue, "fields") and hasattr(issue.fields, "status"):
                current_status = getattr(issue.fields.status, "name", None)
            elif isinstance(issue, dict):
                status_obj = issue.get("fields", {}).get("status")
                if isinstance(status_obj, dict):
                    current_status = status_obj.get("name")
            no_transition_statuses[current_status or "Unknown"].append(issue_key)

    print(f"Found {sum(len(v) for v in no_transition_statuses.values())} issues with no transitions.")
    return dict(no_transition_statuses)


import os
import sys
import yaml
from datetime import datetime
from urllib.parse import unquote
from openpyxl.utils import get_column_letter

if len(sys.argv) < 4:
    print("Usage: python statustime.py <yaml_file> <timestamp> <userlogin>")
    sys.exit(1)

yaml_file = sys.argv[1]
timestamp = sys.argv[2]
userlogin = sys.argv[3]

print(f"statustime parameters  yaml_file={yaml_file} timestamp={timestamp} userlogin={userlogin}")

with open(yaml_file, 'r') as f:
    data = yaml.safe_load(f)

fileinfo = data.get('fileinfo', {})
if not fileinfo:
    print("No fileinfo found in the YAML file.")
    sys.exit(1)
basename = fileinfo.get('basename')
tablename = fileinfo.get('table').replace(" ", "_") if fileinfo.get('table') else ""
source = fileinfo.get('source')
scope_file = fileinfo.get('scope file')

if not basename:
    print("No 'basename'found in fileinfo. Expecting 'basename' key.")
    sys.exit(1)

if not tablename:
    print("No 'table' found in fileinfo. Expecting 'table' key.")
    sys.exit(1)


if "statustime.scope" in yaml_file.lower():
    print(f"statustime detected based on filename: {yaml_file}.")
    # You can set a flag or handle import-specific logic here if needed
    output_file = basename + "." + tablename + timestamp + ".statustime.jira.csv"
    #mode = "assignee" #"resolved"
else:
    print("Error: YAML filename is not 'statustime.scope.yaml")
    sys.exit(1)

fields = data.get('fields', [])
scan_ahead_nonblank_rows = data.get('scan_ahead_nonblank_rows', 0)

# Convert list of dicts into a dictionary
fields_dict = {field["value"]: field.get("index", "<blank>") for field in fields}

field_values = [field.get('value') for field in fields if 'value' in field]
field_indexes = [field.get('index') for field in fields if 'index' in field]
field_values_str = ','.join(field_values)
field_indexes_str = ','.join(map(str, field_indexes))

print("Source file,", source)
print("basename,", basename)
print("Table,", tablename)
print("Field indexes,", field_indexes_str)
print("Field values,", field_values_str)

'''with open(output_file, "w") as outfile:
    outfile.write("Source file," + source + "\n")
    outfile.write("Basename," + basename + "\n")
    outfile.write("Scope file," + scope_file + "\n")    
    outfile.write("Table," + tablename + "\n")
    outfile.write("Field indexes," + field_indexes_str + "\n")
    outfile.write("Field values," + field_values_str + "\n")
'''

# yaml will expected to contain params, eg:
'''
    fileinfo:
    basename: Quickstart.xlsx
    scope file: Quickstart.xlsx.Cycle_Time.20250928_164350.statustime.scope.yaml
    source: Quickstart.xlsx
    table: Cycle_Time
    jql: project in (tes,fr) and updated > -90d
    row: 27
    col: 1
    lastrow: 49
'''

# the following were addded by scope.py in runrate yaml file
statustime_table_row = data.get('row', None)
print(f"quickstart_table_row from yaml: {statustime_table_row}")
statustime_table_col = data.get('col', None)
print(f"quickstart_table_col from yaml: {statustime_table_col}")
last_excel_row = data.get('lastrow', None)
print(f"last_excel_row from yaml: {last_excel_row}")
jql_str = data.get('jql', None)
print(f"jql from yaml: {jql_str}")

# validate
if not jql_str or statustime_table_row is None or statustime_table_col is None or last_excel_row is None:
    print("❌ Error: Missing required fields in YAML file")
    sys.exit(1)


    
from jira import JIRA
from dotenv import load_dotenv

JIRA_MAX_RESULTS = False

# Load environment variables from a .env file if present
# -------------------------------
#load_dotenv()
# load .env from config folder
ENV_PATH = f"../../../config/env.{userlogin}"
load_dotenv(dotenv_path=ENV_PATH)

JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN")
JIRA_URL = os.environ.get("JIRA_URL")
JIRA_EMAIL = os.environ.get("JIRA_EMAIL")

if not JIRA_API_TOKEN:
    print("Error: JIRA_API_TOKEN environment variable not set.")
    sys.exit(1)


# Connect to Jira with basic auth
try:
    jira = JIRA(server=JIRA_URL, basic_auth=(JIRA_EMAIL, JIRA_API_TOKEN))
    #print("✅ Successfully connected to Jira.")
except Exception as e:
    print(f"❌ Failed to connect to Jira: {e}")
    sys.exit(1)

print(f"JIRA client connected to {JIRA_URL} login:{JIRA_EMAIL} apitoken:{JIRA_API_TOKEN} ")

issues = []  # global issues list to hold results from both ID and JQL searches



try:
    jql_query = jql_str
    jql_query = jql_query.lower().replace("jql ", "").strip()
    issues = jira.search_issues(jql_query, maxResults = JIRA_MAX_RESULTS, expand='changelog')
    print(f"Found {len(issues)} issues for JQL query '{jql_query}':")
    if len(issues) == 0:
        print(f"No issues found for JQL query '{jql_query}'.")
        sys.exit(1)
    filtered_ids = [issue.key for issue in issues]
    print(f"Filtered IDs from JQL: {filtered_ids}")
    #jql_ids = []  # Clear jql_ids to avoid re-processing later below
except Exception as e:
    print(f"❌ Failed to search issues for JQL query '{jql_query}': {e}")
    sys.exit(1)


changes_list = []

r = statustime_table_row + 1
excel_col = statustime_table_col + 1

#########################
'''
# Calculate chain cycle times
results = calculate_average_chain_cycle_time(issues)
print(f"\nCalculated cycle times for {len(results)} unique transition chains.")

transition_data = results

# Sort chains by average time (descending)
sorted_chains = sorted(
    transition_data.items(),
    key=lambda x: x[1]['average_hours'],
    reverse=True
)

r = statustime_table_row + 1
excel_col = statustime_table_col + 1

# write out the timestamp in cell adjacent to <> so we can tell when the update occurred
coord = f"{get_column_letter(statustime_table_col + 2)}{statustime_table_row + 1}"
now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
entry = f"{coord} = {now_str} ||"
print(entry)
changes_list.append(entry)

if scan_ahead_nonblank_rows:
    prefix = ""
    scan_ahead_nonblank_rows -= 1
else:
    prefix = "INSERT"

r += 1  # bump row one more time so INSERT are done at the row right below the statustime tag
changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} TRANSITION CHAIN || ")
changes_list.append(f"{get_column_letter(excel_col + 1)}{r} =   AVERAGE || ")
changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =   MEDIAN  || ")
changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =   STD DEV || ")
changes_list.append(f"{get_column_letter(excel_col + 4)}{r} =   MIN || ")
changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =   MAX || ")
changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =   SAMPLE SIZE || ")

r += 1

for chain_str, data in sorted_chains:

    if scan_ahead_nonblank_rows:
        prefix = ""
        scan_ahead_nonblank_rows -= 1
    else:
        prefix = "INSERT"

    print(f"\n{chain_str}")
    median_days = data['median_hours'] / 24.0
    stdev_days = data['stddev_hours'] / 24.0
    min_days = data['min_hours'] / 24.0
    max_days = data['max_hours'] / 24.0
    changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} {chain_str} || ")
    changes_list.append(f"{get_column_letter(excel_col + 1)}{r} =  {data['average_hours']:.1f} hours ({data['average_days']:.1f} days) || ")
    changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =  {data['median_hours']:.1f} hours ({median_days:.1f}) days|| ")
    changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =  {data['stddev_hours']:.1f} hours ({stdev_days:.1f} days) || ")
    changes_list.append(f"{get_column_letter(excel_col + 4)}{r} =  {data['min_hours']:.1f} hours ({min_days:.1f} days) || ")
    changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =  {data['max_hours']:.1f} hours ({max_days:.1f} days) || ")

    # Build hyperlink JQL for the issues in this chain
    jql = "key in ("
    issue_keys = [issue['issue_key'] for issue in data['issues']]
    for key in issue_keys:
        jql += key + ","
    jql = jql.rstrip(",") + ") order by key asc"

    from my_utils import *
    hyperlink = make_hyperlink_formula(f"{JIRA_URL}/issues/?jql={jql}", f"{data['count']}") + " || "
    changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =  {hyperlink} || ")

    r += 1

    print(f"  Average Time: {data['average_hours']:.1f} hours ({data['average_days']:.1f} days)")
    print(f"  Median Time:  {data['median_hours']:.1f} hours")
    print(f"  Range:        {data['min_hours']:.1f} - {data['max_hours']:.1f} hours")
    print(f"  Sample Size:  {data['count']} issues")

changes_list.append(f"{get_column_letter(excel_col)}{r} =  INSERT EOL || ")
'''
###########################



# Calculate transition times
results = calculate_average_status_transition_time(issues)
print(f"\nCalculated transition times for {len(results)} status transitions.")

# Print results
#print_transition_report(results)

transition_data = results
  
# Sort transitions by average time (descending)
sorted_transitions = sorted(
    transition_data.items(),
    key=lambda x: x[1]['average_hours'],
    reverse=True
)


#changes_list = []
#r = statustime_table_row + 1
#excel_col = statustime_table_col + 1



if scan_ahead_nonblank_rows:
    prefix = ""
    scan_ahead_nonblank_rows -= 1
else:
    prefix = "INSERT"

r += 1  # bump row one more time so INSERT are done at the row right below the cycletime tag

# write out the timestamp in cell adjacent to <> so we can tell when the update occurred
coord = f"{get_column_letter(excel_col)}{r}"
now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
entry = f"{coord} = {prefix} {now_str} ||"
print(entry)
changes_list.append(entry)

r += 1  # bump row one more time so INSERT are done at the row right below the statustime tag 
changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} FROM STATUS || ")  # add INSERT to only first since the rest are on same row
changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = TO STATUS || ")  # add INSERT to only first since the rest are on same row
changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =   Average Time || ")
changes_list.append(f"{get_column_letter(excel_col + 3 )}{r} =   Median Time || ")
changes_list.append(f"{get_column_letter(excel_col + 4 )}{r} =   StdDev Time || ")
#changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =  INSERT Range || ")
changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =   Min Time || ")
changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =   Max Time|| ")
changes_list.append(f"{get_column_letter(excel_col + 7)}{r} =   Sample Size || ")

r +=1

for (from_status, to_status), data in sorted_transitions:

    
    if scan_ahead_nonblank_rows:
        prefix = ""
        scan_ahead_nonblank_rows -= 1
    else:
        prefix = "INSERT"

    print(f"\n{from_status} → {to_status}")    
    changes_list.append(f"{get_column_letter(excel_col)}{r} = {prefix} {from_status} || ")  # add INSERT only to first row since rest are in sme row
    changes_list.append(f"{get_column_letter(excel_col + 1)}{r} = {to_status} || ")  # add INSERT only to first row since rest are in sme row
    changes_list.append(f"{get_column_letter(excel_col + 2)}{r} =  {data['average_hours']:.1f} hours ({data['average_days']:.1f} days) || ")
    changes_list.append(f"{get_column_letter(excel_col + 3)}{r} =  {data['median_hours']:.1f} hours || ")
    changes_list.append(f"{get_column_letter(excel_col + 4)}{r} =  {data['stddev_hours']:.1f} hours || ")
    #changes_list.append(f"{get_column_letter(excel_col + 3)}{r} = INSERT {data['min_hours']:.1f} - {data['max_hours']:.1f} hours || ")
    changes_list.append(f"{get_column_letter(excel_col + 5)}{r} =  {data['min_hours']:.1f} hours || ")
    changes_list.append(f"{get_column_letter(excel_col + 6)}{r} =  {data['max_hours']:.1f} hours || ")

    jql = "key in ("
    # Print only the issue keys for this transition
    issue_keys = [issue['issue_key'] for issue in data['issues']]
    for key in issue_keys:
        #print(f"{key}")
        jql = jql + key + ","

    jql = jql.rstrip(",") + ")"

    from my_utils import *
    hyperlink = make_hyperlink_formula(f"{JIRA_URL}/issues/?jql={jql}", f"{data['count']}") + " || "
    changes_list.append(f"{get_column_letter(excel_col + 7)}{r} =  {hyperlink} || ")
    #changes_list.append(f"{get_column_letter(excel_col + 4)}{r} = INSERT {data['count']} issues || ")
    
    r += 1

    print(f"  Average Time: {data['average_hours']:.1f} hours ({data['average_days']:.1f} days)")
    print(f"  Median Time:  {data['median_hours']:.1f} hours")
    print(f"  Range:        {data['min_hours']:.1f} - {data['max_hours']:.1f} hours")
    print(f"  Sample Size:  {data['count']} issues")

changes_list.append(f"{get_column_letter(excel_col)}{r} =  INSERT EOL || ")

no_transitions = get_issues_with_no_transitions(issues)
print("\nIssues with no transitions:")
for status, issue_keys in no_transitions.items():
    print(f"{status}: {', '.join(issue_keys)}")









changes_file = yaml_file.replace("scope.yaml","import.changes.txt")
print(f"Writing changes to {changes_file}")

if changes_list:
    with open(changes_file, "w") as f:
        for entry in changes_list:
            if "||None" in entry:
                entry = entry.replace("||None", "||")
            f.write(entry + "\n")
            print(entry)
    print(f"Changes written to {changes_file} with ({len(changes_list)} entries).")
else:
    print(f"No changes to write. Not need to create {changes_file}")
