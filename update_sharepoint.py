import requests
from collections import defaultdict
import argparse
from urllib.parse import urlparse, quote, unquote
import msal
import os
import json
import string
import shutil
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re
from dotenv import load_dotenv
import time

# -------------------------------
# Config from environment variables
# -------------------------------
ENV_PATH = "../../../config/env.system"
load_dotenv(dotenv_path=ENV_PATH)

SCOPES = ["https://graph.microsoft.com/.default"]
CLIENT_ID = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]
TENANT_ID = os.environ["TENANT_ID"]
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
TOKEN_CACHE_FILE = "../../../config/token_cache.json"

# -------------------------------
# Token cache functions
# -------------------------------
def load_cache(userlogin=None):
    global TOKEN_CACHE_FILE
    if userlogin:
        TOKEN_CACHE_FILE = f"../../../config/token_cache_{userlogin}.json"
    
    cache = msal.SerializableTokenCache()
    if os.path.exists(TOKEN_CACHE_FILE):
        cache.deserialize(open(TOKEN_CACHE_FILE, "r").read())
        print(f"Loaded cache from '{TOKEN_CACHE_FILE}'")
    return cache

def save_cache(cache, userlogin=None):
    if cache.has_state_changed:
        global TOKEN_CACHE_FILE
        if userlogin:
            TOKEN_CACHE_FILE = f"../../../config/token_cache_{userlogin}.json"
        open(TOKEN_CACHE_FILE, "w").write(cache.serialize())

def get_app_token_delegated(userlogin):
    print(f"🔑 Acquiring delegated user token for user={userlogin}")
    cache = load_cache(userlogin)
    
    cca = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET,
        token_cache=cache
    )
    
    accounts = cca.get_accounts()
    if accounts:
        result = cca.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            print("✅ Using cached user token")
            return result["access_token"]
    
    raise Exception("❌ No cached user token found. Please log in first.")

def get_app_token():
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    cca = msal.ConfidentialClientApplication(
        CLIENT_ID, authority=authority, client_credential=CLIENT_SECRET
    )
    result = cca.acquire_token_for_client(scopes=SCOPES)
    if "access_token" not in result:
        raise Exception(f"Failed to get token: {result}")
    return result["access_token"]

# -------------------------------
# Command line arguments
# -------------------------------
parser = argparse.ArgumentParser(description="Update Excel file on SharePoint with changes from a file")
parser.add_argument("file_url", help="Full URL to the Excel file in SharePoint")
parser.add_argument("changes_file", help="Path to the local changes file")
parser.add_argument("timestamp", help="String tag (timestamp) for local output temp files")
parser.add_argument("userlogin", help="String userlogin")
parser.add_argument("worksheet", help="String worksheet name")
parser.add_argument("--user_auth", action="store_true", help="Enable delegated user auth flow")

args = parser.parse_args()

auth_user = args.user_auth
userlogin = args.userlogin
worksheet_name = args.worksheet
file_url = args.file_url
changes_file = args.changes_file
timestamp = args.timestamp

print(f"Worksheet name = {worksheet_name}")
print(f"User login = {userlogin}")

if userlogin is None:
    print("ERROR! Required 'userlogin' argument is missing.")
    exit(1)

delegated_auth = False
if auth_user and userlogin:
    print(f"Using delegated authorization for user: {userlogin}")
    delegated_auth = True
    CLIENT_ID = os.environ["CLIENT_ID2"]
    CLIENT_SECRET = os.environ["CLIENT_SECRET2"]
    SCOPES = ["User.Read", "Files.ReadWrite.All", "Sites.ReadWrite.All"]
    AUTHORITY = "https://login.microsoftonline.com/common"
else:
    print("Using application authorization")

# Load user settings
ENV_PATH_USER = os.path.join(os.path.dirname(__file__), "config", f"env.{userlogin}")
load_dotenv(dotenv_path=ENV_PATH_USER)

# Get access token
if delegated_auth:
    access_token = get_app_token_delegated(userlogin)
else:
    access_token = get_app_token()

# Check for import/runrate modes
import_mode = "import.changes.txt" in changes_file
runrate_mode = "rate.import.changes.txt" in changes_file

if import_mode:
    print(f"Import mode enabled")
if runrate_mode:
    print(f"Runrate mode enabled")

jira_base_url = os.environ.get("JIRA_URL", "")
print(f"Using JIRA base URL: {jira_base_url}")

# -------------------------------
# Helper functions
# -------------------------------
def is_valid_jira_id(value):
    if not isinstance(value, str):
        return False
    return bool(re.fullmatch(r"[A-Z][A-Z0-9]+-\d+", value))

def is_jql(value):
    return value.strip().upper().startswith("JQL")

def create_hyperlink(value, jira_base_url):
    if value.startswith("URL "):
        value = value.replace("URL", "").strip()
        if is_valid_jira_id(value):
            return f"{jira_base_url}/browse/{value}"
        elif is_jql(value):
            jql_query = str(value).lower().replace("jql", "").strip()
            return f"{jira_base_url}/issues/?jql={jql_query}"
    return None

def _excel_escape_quotes(s: str) -> str:
    return s.replace('"', '""')

def _make_hyperlink_formula(url: str, text: str) -> str:
    text = text.replace("\n", " ")
    return f'=HYPERLINK("{_excel_escape_quotes(url)}","{_excel_escape_quotes(text)}")'

# -------------------------------
# Optimized batch update functions with retry logic
# -------------------------------
def build_batch_requests(row_updates, worksheet_name, site_id, item_id, jira_base_url, runrate_mode=False):
    """
    Build all batch requests for updating multiple rows at once.
    Returns list of batch request objects.
    """
    requests_list = []
    req_id = 1
    
    for row_num in sorted(row_updates.keys()):
        cols = row_updates[row_num]
        
        # Group columns into contiguous range for this row
        col_letters = sorted(cols.keys(), key=lambda c: ord(c))
        if not col_letters:
            continue
            
        start_col = col_letters[0]
        end_col = col_letters[-1]
        row_range = f"{start_col}{row_num}:{end_col}{row_num}"
        
        # Build values array for the entire range
        all_cols = [chr(c) for c in range(ord(start_col), ord(end_col) + 1)]
        values = []
        
        for col_letter in all_cols:
            if col_letter in cols:
                new_val = cols[col_letter]["new"]
                
                # Handle hyperlinks
                if new_val.startswith("URL "):
                    hyperlink = create_hyperlink(new_val, jira_base_url)
                    if hyperlink:
                        new_val = new_val.replace("URL", "").strip()
                        #if "jql" in str(new_val).lower():
                        #    new_val = "Link"
                        new_val = _make_hyperlink_formula(hyperlink, new_val)
                else:
                    # Convert semicolons to newlines
                    new_val = new_val.replace(";", "\n") if ";" in new_val else new_val
                
                values.append(new_val)
            else:
                # For runrate mode, use empty space for non-updated cells
                values.append(" " if runrate_mode else "")
        
        # Add PATCH request for this row
        requests_list.append({
            "id": str(req_id),
            "method": "PATCH",
            "url": f"/sites/{site_id}/drive/items/{item_id}/workbook/worksheets('{worksheet_name}')/range(address='{row_range}')",
            "headers": {"Content-Type": "application/json"},
            "body": {"values": [values]}
        })
        req_id += 1
    
    return requests_list

def build_insert_requests(insert_rows, worksheet_name, site_id, item_id):
    """
    Build insert dimension requests for multiple rows.
    Returns list of Graph API requests (not batch format).
    """
    requests = []
    
    # Sort in descending order to avoid index shifting
    for row_num in sorted(insert_rows.keys(), reverse=True):
        count = 1  # Insert one row at a time for each entry
        end_row = row_num + count - 1
        insert_range = f"{row_num}:{end_row}"
        
        requests.append({
            "row_num": row_num,
            "url": f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{item_id}/workbook/worksheets('{worksheet_name}')/range(address='{insert_range}')/insert",
            "body": {"shift": "Down"}
        })
    
    return requests

def execute_batch_requests(headers, batch_requests, batch_size=5, max_retries=5):
    """
    Execute batch requests with retry logic for rate limiting.
    Reduced batch_size to 5 to avoid rate limits.
    """
    batch_url = "https://graph.microsoft.com/v1.0/$batch"
    total_success = 0
    total_failed = 0
    
    # Split into smaller batches
    for i in range(0, len(batch_requests), batch_size):
        batch = batch_requests[i:i + batch_size]
        batch_num = (i // batch_size) + 1
        
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            batch_body = {"requests": batch}
            resp = requests.post(batch_url, headers=headers, json=batch_body)
            
            if resp.status_code == 429:
                # Rate limit hit - wait and retry
                retry_after = int(resp.headers.get('Retry-After', 10))
                print(f"⏳ Rate limit hit on batch {batch_num}. Waiting {retry_after} seconds...")
                time.sleep(retry_after)
                retry_count += 1
                continue
            
            if resp.status_code != 200:
                print(f"❌ Batch {batch_num} failed: {resp.status_code} {resp.text}")
                total_failed += len(batch)
                break
            
            # Check individual responses
            results = resp.json().get("responses", [])
            batch_failed = 0
            batch_success = 0
            
            for r in results:
                status = r.get("status")
                if status == 429:
                    batch_failed += 1
                elif status >= 400:
                    print(f"⚠️ Batch {batch_num}, request {r.get('id')} failed: {status}")
                    batch_failed += 1
                else:
                    batch_success += 1
            
            if batch_failed > 0 and retry_count < max_retries:
                # Some requests failed due to rate limiting - retry the batch
                wait_time = (2 ** retry_count) * 2  # Exponential backoff
                print(f"⏳ {batch_failed} requests rate limited in batch {batch_num}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
                retry_count += 1
            else:
                total_success += batch_success
                total_failed += batch_failed
                success = True
                print(f"✅ Batch {batch_num}/{(len(batch_requests) + batch_size - 1) // batch_size} completed ({batch_success} success, {batch_failed} failed)")
        
        # Add delay between batches to avoid rate limiting
        if i + batch_size < len(batch_requests):
            time.sleep(2)
    
    return total_success, total_failed

def execute_insert_requests(headers, insert_requests, max_retries=5):
    """
    Execute row insert requests with retry logic.
    """
    success_count = 0
    
    for req in insert_requests:
        row_num = req["row_num"]
        url = req["url"]
        body = req["body"]
        
        retry_count = 0
        while retry_count < max_retries:
            resp = requests.post(url, headers=headers, json=body)
            
            if resp.status_code == 429:
                retry_after = int(resp.headers.get('Retry-After', 5))
                print(f"⏳ Rate limit hit inserting row {row_num}. Waiting {retry_after}s...")
                time.sleep(retry_after)
                retry_count += 1
                continue
            
            if resp.status_code in (200, 201):
                print(f"✅ Inserted blank row at {row_num}")
                success_count += 1
                break
            else:
                print(f"❌ Failed to insert row {row_num}: {resp.status_code} {resp.text}")
                break
        
        # Small delay between inserts
        time.sleep(1)
    
    return success_count

# -------------------------------
# Parse changes file
# -------------------------------
insert_row_values = defaultdict(dict)
row_values = defaultdict(dict)

with open(changes_file, "r") as f:
    print(f"📖 Parsing changes file: {changes_file}")
    for line in f:
        line = line.strip()
        if not line or line.startswith("Changes"):
            continue
        if "=" not in line:
            continue
        
        cell, value_pair = line.split("=", 1)
        if "||" in value_pair:
            new_value, old_value = value_pair.split("||", 1)
        else:
            new_value, old_value = value_pair, ""
        
        col_letter = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell)))
        
        if import_mode and "INSERT" in new_value.upper():
            print(f"Import mode: removing INSERT prefix")
            new_value = new_value.replace("INSERT", "").replace("insert", "").strip()
            insert_row_values[row_num][col_letter.upper()] = {
                "new": new_value.strip(),
                "old": old_value.strip()
            }
        else:
            row_values[row_num][col_letter.upper()] = {
                "new": new_value.strip(),
                "old": old_value.strip()
            }
        
        print(f"Processed {cell}: new='{new_value.strip()}'")

# -------------------------------
# Main execution
# -------------------------------
file_url = unquote(file_url)

if "http" in file_url:
    print(f"📊 SharePoint path: {file_url}")
    
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }
    
    # Extract site and file path
    parsed = urlparse(file_url)
    hostname = parsed.netloc
    path_parts = parsed.path.strip("/").split("/", 2)
    if len(path_parts) < 3:
        raise ValueError("Invalid SharePoint URL format")
    
    site_path = f"/{path_parts[0]}/{path_parts[1]}"
    file_path = "/" + path_parts[2]
    
    # Load metadata
    meta_filename = file_path.strip("/").replace("/", "_") + "." + timestamp + ".meta.json"
    if not os.path.exists(meta_filename):
        raise FileNotFoundError(f"Metadata file {meta_filename} not found")
    
    with open(meta_filename, "r") as f:
        saved_meta = json.load(f)
    saved_etag = saved_meta.get("etag")
    print(f"📂 Loaded saved eTag: {saved_etag}")
    
    # Get site ID (1 API call)
    site_api_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:{site_path}"
    resp_site = requests.get(site_api_url, headers=headers)
    resp_site.raise_for_status()
    site_id = resp_site.json()["id"]
    
    # Get file ID and check eTag (1 API call)
    file_api_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{quote(file_path)}"
    resp_file = requests.get(file_api_url, headers=headers)
    resp_file.raise_for_status()
    file_meta = resp_file.json()
    item_id = file_meta["id"]
    current_etag = file_meta["eTag"]
    
    print(f"🔍 Current eTag: {current_etag}")
    
    # Verify eTag
    if current_etag != saved_etag:
        print("❌ eTag mismatch! File has been modified since last download.")
        print("👉 Aborting update to prevent overwriting newer changes.")
        exit(1)
    
    print("✅ eTag matches. Safe to apply updates.")
    
    print(f"\n🚀 Starting optimized batch updates with rate limit handling...")
    
    # Step 1: Execute all row inserts (sequential - cannot batch)
    if insert_row_values:
        print(f"📝 Processing {len(insert_row_values)} row insertions...")
        insert_reqs = build_insert_requests(insert_row_values, worksheet_name, site_id, item_id)
        insert_count = execute_insert_requests(headers, insert_reqs)
        print(f"✅ Inserted {insert_count} rows")
    
    # Step 2: Execute runrate blank rows if needed
    if runrate_mode and row_values:
        last_row = max(row_values.keys())
        print(f"🧹 Runrate mode: inserting 2 blank rows after row {last_row}")
        runrate_reqs = build_insert_requests({last_row + 1: {}}, worksheet_name, site_id, item_id)
        # Insert 2 rows
        for _ in range(2):
            execute_insert_requests(headers, runrate_reqs)
    
    # Step 3: Build all update requests (inserted rows + regular updates)
    all_updates = {**insert_row_values, **row_values}
    
    if all_updates:
        print(f"📦 Building batch requests for {len(all_updates)} rows...")
        batch_reqs = build_batch_requests(
            all_updates, worksheet_name, site_id, item_id, jira_base_url, runrate_mode
        )
        
        print(f"📤 Executing {len(batch_reqs)} update requests in batches of 5 with retry logic...")
        success, failed = execute_batch_requests(headers, batch_reqs, batch_size=5)
        
        print(f"\n✅ All updates completed!")
        print(f"📊 Summary: {success} successful, {failed} failed")
    else:
        print("⚠️ No updates to apply")

else:
    # Local file handling
    print(f"📁 Processing local file: {file_url}")
    
    if not os.path.isfile(file_url):
        raise FileNotFoundError(f"Excel file not found: {file_url}")
    
    # Create backup
    base, ext = os.path.splitext(file_url)
    backup_path = f"{base}.{timestamp}{ext}"
    
    if not os.path.exists(backup_path):
        shutil.copy(file_url, backup_path)
        print(f"📋 Created backup: {backup_path}")
    
    # Load workbook
    wb = load_workbook(file_url)
    ws = wb[worksheet_name] if worksheet_name else wb.active
    
    # Apply updates
    for row_num, cols in row_values.items():
        for col_key, val_dict in cols.items():
            new_val = val_dict.get("new")
            if new_val is not None:
                row_idx = int(row_num)
                if isinstance(col_key, str) and col_key.isalpha():
                    col_idx = column_index_from_string(col_key)
                else:
                    col_idx = int(col_key)
                
                ws.cell(row=row_idx, column=col_idx).value = new_val
    
    print(f"💾 Saving updated file: {file_url}")
    wb.save(file_url)
    print("✅ Local file updated successfully!")