import sys
import yaml
from jira import JIRA
import os
from dotenv import load_dotenv
import re
import ollama
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
from collections import defaultdict
import calendar
from datetime import datetime, timedelta
from collections import defaultdict
import calendar
import glob
import hashlib
from bs4 import BeautifulSoup


# Cache dictionary to avoid repeated calls
user_cache = {}


def get_week_number(date):
    """Get the ISO week number for a given date"""
    return date.isocalendar()[1]

def get_year_week(date):
    """Get year and week as a tuple for proper sorting across years"""
    iso_cal = date.isocalendar()
    return (iso_cal[0], iso_cal[1])  # (year, week)

def get_year_month(date):
    """Get year and month as a tuple"""
    return (date.year, date.month)

def get_year_day(date):
    """Get year and day of year as a tuple"""
    return (date.year, date.timetuple().tm_yday)

def get_period_key(date, interval):
    """Get the appropriate period key based on interval type"""
    if interval == "days":
        return (date.year, date.month, date.day)
    elif interval == "weeks":
        return get_year_week(date)
    elif interval == "months":
        return get_year_month(date)
    elif interval == "years":
        return (date.year,)
    else:
        raise ValueError(f"Invalid interval: {interval}")

def get_period_bounds(date, interval):
    """Get start and end dates for a period based on interval type"""
    if interval == "days":
        start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1) - timedelta(seconds=1)
        return start, end
    
    elif interval == "weeks":
        # Start on Monday
        start = date - timedelta(days=date.weekday())
        end = start + timedelta(days=6)
        return start, end
    
    elif interval == "months":
        # First day of month
        start = date.replace(day=1)
        # Last day of month
        if date.month == 12:
            end = date.replace(day=31)
        else:
            next_month = date.replace(month=date.month + 1, day=1)
            end = next_month - timedelta(days=1)
        return start, end
    
    elif interval == "years":
        # First day of year
        start = date.replace(month=1, day=1)
        # Last day of year
        end = date.replace(month=12, day=31)
        return start, end
    
    else:
        raise ValueError(f"Invalid interval: {interval}")

def advance_period(date, interval):
    """Advance date to the next period based on interval type"""
    if interval == "days":
        return date + timedelta(days=1)
    elif interval == "weeks":
        return date + timedelta(weeks=1)
    elif interval == "months":
        # Handle month rollover
        if date.month == 12:
            return date.replace(year=date.year + 1, month=1)
        else:
            return date.replace(month=date.month + 1)
    elif interval == "years":
        return date.replace(year=date.year + 1)
    else:
        raise ValueError(f"Invalid interval: {interval}")

def get_period_label(period_key, interval):
    """Generate a human-readable label for the period"""
    if interval == "days":
        year, month, day = period_key
        return f"{year}-{month:02d}-{day:02d}"
    elif interval == "weeks":
        year, week_num = period_key
        return f"{year}-W{week_num:02d}"
    elif interval == "months":
        year, month = period_key
        month_name = calendar.month_abbr[month]
        return f"{year}-{month_name}"
    elif interval == "years":
        year = period_key[0]
        return f"{year}"
    else:
        return str(period_key)

def bucketize_issues_by_interval(issues, mode, interval="weeks", assignee_filter=None):
    """
    Bucketize Jira issues by time interval based on resolution/creation date.
    
    Args:
        issues: List of Jira issues from jira.search_issues()
        mode: String indicating which date field to use ("resolved", "open", "closed")
        interval: String indicating time interval ("days", "weeks", "months", "years")
        
    Returns:
        tuple: (period_buckets, period_info)
        - period_buckets: List of lists, each containing issues for that period
        - period_info: List of tuples with period metadata (varies by interval)
    """
    
    # Validate interval
    valid_intervals = ["days", "weeks", "months", "years"]
    if interval not in valid_intervals:
        print(f"❌ Invalid interval '{interval}'. Must be one of: {', '.join(valid_intervals)}")
        return [], []
    
    # Dictionary to group issues by period
    issues_by_period = defaultdict(list)
    resolution_dates = []
    
    print(f"\n🔍 Processing issues for {interval} bucketization for assignee={assignee_filter}...")
    print(f"Mode: {mode}")
    print(f"Interval: {interval}")
    
    # Group issues by period and collect resolution dates
    for issue in issues:
        #a = issue.fields.assignee.displayName if issue.fields.assignee else "Unassigned"  
        a = get_resolved_by_user(issue) 

        if assignee_filter is not None and assignee_filter not in a:
            continue  # skip issues not matching assignee filter

        #print(f"Processing issue {issue.key} resolved by {a}")

        try:
            # Get resolution date - try different field names based on mode
            resolved_date = None
            
            if "resolved" in mode:
                if hasattr(issue.fields, 'resolutiondate') and issue.fields.resolutiondate:
                    resolved_date = datetime.strptime(issue.fields.resolutiondate[:19], '%Y-%m-%dT%H:%M:%S')
                elif hasattr(issue.fields, 'resolved') and issue.fields.resolved:
                    resolved_date = datetime.strptime(issue.fields.resolved[:19], '%Y-%m-%dT%H:%M:%S')
            elif "open" in mode: 
                if hasattr(issue.fields, 'created') and issue.fields.created:
                    resolved_date = datetime.strptime(issue.fields.created[:19], '%Y-%m-%dT%H:%M:%S')
                elif hasattr(issue.fields, 'createddate') and issue.fields.createddate:
                    resolved_date = datetime.strptime(issue.fields.createddate[:19], '%Y-%m-%dT%H:%M:%S')
            elif "closed" in mode: 
                if hasattr(issue.fields, 'closed') and issue.fields.closed:
                    resolved_date = datetime.strptime(issue.fields.closed[:19], '%Y-%m-%dT%H:%M:%S')
            else:
                print(f"❌ Unknown mode '{mode}' specified for bucketization.")
                return [], []
                                
            if resolved_date:
                period_key = get_period_key(resolved_date, interval)
                issues_by_period[period_key].append(issue)
                resolution_dates.append(resolved_date)
                print(f"created period_key={period_key}")
                print(f"created resolved_date={resolved_date}")
                
                period_label = get_period_label(period_key, interval)
                print(f"  📅 {issue.key}: {mode} {resolved_date.strftime('%Y-%m-%d')} (period_label={period_label})")
            else:
                print(f"  ⚠️  {issue.key}: No {mode} date found - skipping")
                
        except Exception as e:
            print(f"  ❌ Error processing {issue.key}: {e}")
            continue
    
    if not resolution_dates:
        print(f"❌ No issues with valid {mode} dates found!")
        return [], []
    
    # Find the earliest and latest resolution dates
    earliest_date = min(resolution_dates)
    latest_date = max(resolution_dates)
    
    print(f"\n📊 Date range: {earliest_date.strftime('%Y-%m-%d')} to {latest_date.strftime('%Y-%m-%d')}")
    
    # Generate all periods in the range
    current_date = earliest_date
    
    # Align to the start of the first period
    period_start, _ = get_period_bounds(current_date, interval)
    current_date = period_start
    
    period_buckets = []
    period_info = []
    
    while current_date <= latest_date:
        period_key = get_period_key(current_date, interval)
        period_start, period_end = get_period_bounds(current_date, interval)
        
        # Get issues for this period (if any)
        period_issues = issues_by_period.get(period_key, [])
        period_buckets.append(period_issues)
        
        # Store period information with appropriate metadata
        if interval == "weeks":
            year_week = get_year_week(current_date)
            period_info.append((year_week[0], year_week[1], period_start, period_end))
        elif interval == "months":
            year_month = get_year_month(current_date)
            period_info.append((year_month[0], year_month[1], period_start, period_end))
        elif interval == "years":
            period_info.append((current_date.year, -1, period_start, period_end))  #return a fake -1 param to keep return count consistent
        #elif interval == "days":
        #    period_info.append((current_date.year, current_date.month, current_date.day, period_start, period_end))
        
        # Move to next period
        current_date = advance_period(current_date, interval)
        
        # Safety check to prevent infinite loops
        if len(period_buckets) > 100:
            print(f"⚠️  Warning: Generated {period_buckets} which more than max 100 periods. Stopping.")
            break
    
    print(f"\n✅ Created {len(period_buckets)} {interval} buckets")
    
    return period_buckets, period_info


# Backwards compatibility - keep the old function name
def bucketize_issues_by_weeks(issues, mode):
    """Legacy function - calls bucketize_issues_by_interval with weeks"""
    return bucketize_issues_by_interval(issues, mode, interval="weeks")


def get_week_number(date):
    """Get the ISO week number for a given date"""
    return date.isocalendar()[1]

def get_year_week(date):
    """Get year and week as a tuple for proper sorting across years"""
    iso_cal = date.isocalendar()
    return (iso_cal[0], iso_cal[1])  # (year, week)


def print_weekly_summary(weekly_buckets, week_info):
    """Print a summary of issues bucketized by interval"""
    
    print(f"\n📈 INTERVAL BUCKETIZATION SUMMARY")
    print(f"{'='*60}")
    
    total_issues = sum(len(bucket) for bucket in weekly_buckets)
    print(f"Total issues processed: {total_issues}")
    print(f"Total intervals in range: {len(weekly_buckets)}")
    print(f"{'='*60}")
    
    for i, (bucket, (year, week_num, start_date, end_date)) in enumerate(zip(weekly_buckets, week_info)):
        week_str = f"Interval {i+1} ({year}-{week_num:02d})"
        date_range = f"{start_date.strftime('%m/%d')} - {end_date.strftime('%m/%d/%Y')}"
        issue_count = len(bucket)
        
        print(f"{week_str:<20} {date_range:<20} Issues: {issue_count}")
        
        # Print issue keys for this week
        if bucket:
            issue_keys = [issue.key for issue in bucket]
            print(f"{'':>20} └─ {', '.join(issue_keys)}")
        
        print()


def bucketize_issues_by_weeks_foo(issues, mode):
    """
    Bucketize Jira issues by calendar weeks based on resolution date.
    
    Args:
        issues: List of Jira issues from jira.search_issues()
        
    Returns:
        tuple: (weekly_buckets, week_info)
        - weekly_buckets: List of lists, each containing issues for that week
        - week_info: List of tuples with (year, week_num, start_date, end_date) for each bucket
    """
    
    # Dictionary to group issues by year-week
    issues_by_week = defaultdict(list)
    resolution_dates = []
    
    print("\n🔍 Processing issues for weekly bucketization...")
    print(f"Mode: {mode}")
    
    # Group issues by week and collect resolution dates
    for issue in issues:
        try:
            # Get resolution date - try different field names
            resolved_date = None
            
            if "resolved" in mode:
                if hasattr(issue.fields, 'resolutiondate') and issue.fields.resolutiondate:
                    resolved_date = datetime.strptime(issue.fields.resolutiondate[:19], '%Y-%m-%dT%H:%M:%S')
                elif hasattr(issue.fields, 'resolved') and issue.fields.resolved:
                    resolved_date = datetime.strptime(issue.fields.resolved[:19], '%Y-%m-%dT%H:%M:%S')
            elif "open" in mode: 
                if hasattr(issue.fields, 'created') and issue.fields.created:
                    resolved_date = datetime.strptime(issue.fields.created[:19], '%Y-%m-%dT%H:%M:%S')
                elif hasattr(issue.fields, 'createddate') and issue.fields.createddate:
                    resolved_date = datetime.strptime(issue.fields.createddate[:19], '%Y-%m-%dT%H:%M:%S')
            elif "closed" in mode: 
                if hasattr(issue.fields, 'closed') and issue.fields.created:
                    resolved_date = datetime.strptime(issue.fields.closed[:19], '%Y-%m-%dT%H:%M:%S')
            else:
                print(f"❌ Unknown mode '{mode}' specified for bucketization.")
                return [], []
                                
            if resolved_date:
                year_week = get_year_week(resolved_date)
                issues_by_week[year_week].append(issue)
                resolution_dates.append(resolved_date)
                print(f"  📅 {issue.key}: Resolved {resolved_date.strftime('%Y-%m-%d')} (Week {year_week[1]}, {year_week[0]})")
            else:
                print(f"  ⚠️  {issue.key}: No resolution date found - skipping")
                
        except Exception as e:
            print(f"  ❌ Error processing {issue.key}: {e}")
            continue
    
    if not resolution_dates:
        print("❌ No issues with valid resolution dates found!")
        return [], []
    
    # Find the earliest and latest resolution dates
    earliest_date = min(resolution_dates)
    latest_date = max(resolution_dates)
    
    print(f"\n📊 Date range: {earliest_date.strftime('%Y-%m-%d')} to {latest_date.strftime('%Y-%m-%d')}")
    
    # Generate all weeks in the range
    current_date = earliest_date
    # Go to the start of the week (Monday)
    current_date = current_date - timedelta(days=current_date.weekday())
    
    weekly_buckets = []
    week_info = []
    
    while current_date <= latest_date:
        year_week = get_year_week(current_date)
        week_start = current_date
        week_end = current_date + timedelta(days=6)
        
        # Get issues for this week (if any)
        week_issues = issues_by_week.get(year_week, [])
        weekly_buckets.append(week_issues)
        
        # Store week information
        week_info.append((year_week[0], year_week[1], week_start, week_end))
        
        # Move to next week
        current_date += timedelta(weeks=1)
    
    return weekly_buckets, week_info


def parse_runrate_params(runrate_params_list):
    print(f"entered parse_runrate_params(...) with arg={runrate_params_list} ")
    params = {}
    params["mode"] = "weeks"  # default mode

    for entry in runrate_params_list:
        entry = str(entry).strip()

        
        print(f"checking entry={entry.lower()}")
        if entry.lower().startswith("weeks"):
            print("entry startswith weeks")
            # e.g. "weeks 6"
            params["mode"] = "weeks"

            # dead code since i decided not to support number of weeks. Durations is completed determiend by jql filter time frame
            '''parts = entry.split(maxsplit=1)
            if len(parts) == 2 and parts[1].isdigit():
                params["weeks"] = int(parts[1])
            else:
                params["weeks"] = parts[1] if len(parts) == 2 else None
        '''
        elif entry.lower().startswith("days"):
            print("entry startswith days")
            # e.g. "days 30"
            params["mode"] = "days"
        elif entry.lower().startswith("months"):
            print("entry startswith months")
             # e.g. "months 3"
            params["mode"] = "months"
        elif entry.lower().startswith("years"):
            print("entry startswith years")
             # e.g. "years 1"
            params["mode"] = "years"
        elif entry.lower().startswith("jql"):
            print("entry startswith jql")
             # e.g. "JQL project = tes and assignee = nadeem"
            jql_query = entry[3:].strip()  # remove "JQL"
            params["jql"] = jql_query

    print(f"exiting parse_runrate_params(...) return mode:{params['mode']}, jql:{params['jql']}")
    return params



# Without escaping, Excel would see the unescaped quote as the end of the string, breaking the formula:
# So _excel_escape_quotes() prevents that by doubling the quotes inside the formula string
# the correct way to represent quotes inside Excel string literals.
def _excel_escape_quotes(s: str) -> str:
    # Excel doubles double-quotes inside string literals
    return s.replace('"', '""')

# use TinyURL when hyperlink length exceeds 255 which break excel hyperlinks
import requests
def shorten_url(url: str) -> str:
    """Shorten a URL using TinyURL."""
    try:
        api_url = f"http://tinyurl.com/api-create.php?url={url}"
        response = requests.get(api_url, timeout=5)
        if response.status_code == 200:
            return response.text.strip()
    except Exception as e:
        print(f"⚠️ URL shortening failed for {url}: {e}")
    return url  # fallback to original if shortening fails


def _make_hyperlink_formula(url: str, text: str) -> str:
    """Create an Excel HYPERLINK formula; shorten URL only if it's too long."""
    text = text.replace("\n", " ")

    # Excel's HYPERLINK() formula limit for URLs is ~255 characters
    short_url = url
    if len(url) > 255:
        print(f"URL too long ({len(url)} chars), shortening with TinyURL...")
        short_url = shorten_url(url)

    return f'=HYPERLINK("{_excel_escape_quotes(short_url)}","{_excel_escape_quotes(text)}")'


# takes jira issue object (that include changelog) previously returned by jira client search. this new function will 
# search the changelog history in the issue and return the username that changed the issue status to resolved status
from datetime import datetime
def get_resolved_by_user(issue):
    """
    Extract the username or display name of the user who changed the issue status to 'Resolved'.

    Args:
        issue: A JIRA Issue object (must include changelog, e.g., retrieved with expand='changelog').

    Returns:
        str: The display name or username of the user who resolved the issue,
             or 'Unknown' if not found.
    """
    if not hasattr(issue, "changelog") or not issue.changelog:
        print(f"⚠️ Issue {getattr(issue, 'key', 'Unknown')} has no changelog data.")
        return "Unknown"

    histories = getattr(issue.changelog, "histories", [])
    if not histories:
        print(f"⚠️ Issue {issue.key} changelog has no histories.")
        return "Unknown"

    for history in sorted(histories, key=lambda h: getattr(h, "created", "")):
        # Extract author (safe for JIRA objects)
        author = None
        if hasattr(history, "author"):
            author = getattr(history.author, "displayName", None) or getattr(history.author, "name", None)
        elif isinstance(history, dict):
            author_data = history.get("author", {})
            author = author_data.get("displayName") or author_data.get("name")

        # Loop through all items in this history
        items = getattr(history, "items", [])
        for item in items:
            # Handle both object and dict
            if hasattr(item, "field"):
                field = getattr(item, "field", None)
                from_status = getattr(item, "fromString", None)
                to_status = getattr(item, "toString", None)
            elif isinstance(item, dict):
                field = item.get("field")
                from_status = item.get("fromString")
                to_status = item.get("toString")
            else:
                continue  # unexpected type

            if field == "status" and to_status and to_status.lower() in {"resolved", "done", "completed"}:
                #print(f"✅ Issue {issue.key} resolved by {author or 'Unknown'} "
                #      f"at {getattr(history, 'created', 'unknown time')}")
                return author or "Unknown"
                                
    # it's possible issue went straight to close without a resolved statue.
    # we will make resolved by UNKNOWN in this case.  Alternatively you can use CLOSED author but leave that for future if needed.
    print(f"ℹ️ Issue {issue.key} was never transitioned to 'Resolved'. Current status={issue.fields.status.name}")
    return "Unknown"



def to_filename(path: str, max_len=200):
    # Make filesystem-safe
    clean = re.sub(r'\s*→\s*', '_', path).replace(' ', '-')
    
    # Hash full path for uniqueness
    hash_suffix = hashlib.md5(path.encode()).hexdigest()[:6]
    
    # Truncate safely before appending hash
    trimmed = clean[:max_len]
    
    return f"{trimmed}_{hash_suffix}"


def write_execsummary_yaml(jira_ids, file_info, chain_str, chain_row, timestamp):
    # always create <exec summary> yaml files incase they're needed down the chain
    # step 1 hunt for jira id in all rows and build a list
    # step 2 hunt for jql in all the rows and get list of jira id and add to list from #1
    # step 3 Create ayaml new exec summary scope yaml file with all the Jira found in #1 #2
    # step 4 read_jira will process this yaml downstream

    fname = f"{file_info['basename']}.{file_info['sheet']}.{file_info['table']}.{chain_str}"
    fname_hash = to_filename(fname)
    #execsummary_scope_output_file = f"{file_info['basename']}.{file_info['sheet']}.{file_info['table']}.{table_hash}.{timestamp}.aisummary.scope.yaml"
    execsummary_scope_output_file = f"{fname_hash}.{timestamp}.assignee.scope.yaml"
    print(f"write_execsummary_yaml(...) called to write to file={execsummary_scope_output_file}")

    file_info["scope file"] = execsummary_scope_output_file
    #file_info["table"] = cleaned_value

    with open(execsummary_scope_output_file, 'w') as f:
        yaml.dump({ "fileinfo": file_info }, f, default_flow_style=False)
        yaml.dump({ "chain_row": chain_row }, f, default_flow_style=False)

    jira_fields = []

    jira_fields.append({"value": "key", "index": 0})
    jira_fields.append({"value": "summary", "index": 0})
    jira_fields.append({"value": "description", "index": 0})
    jira_fields.append({"value": "status", "index": 0})
    jira_fields.append({"value": "issuetype", "index": 0})
    jira_fields.append({"value": "priority", "index": 0})
    jira_fields.append({"value": "created", "index": 0})
    jira_fields.append({"value": "assignee", "index": 0})
    jira_fields.append({"value": "status", "index": 0})
    jira_fields.append({"value": "comments", "index": 0})


    with open(execsummary_scope_output_file, 'a') as f:
        yaml.dump({ "fields": jira_fields }, f, default_flow_style=False)
   
    if jira_ids:
        print(f"{len(jira_ids)} JIRA IDs found: {jira_ids}")

        # instead just dump the jira_ids to the scope file. read_jira.py will take care of it, ie run jql and get the jira ids.         
        with open(execsummary_scope_output_file, 'a') as f:
            yaml.dump({"jira_ids": jira_ids}, f, default_flow_style=False)
            # commented out since  defautl values feature not supported or needed in this case
            # only interested in generated a yaml file with fields ids and jira ids that will be used
            # by cycletime.py on 2nd pass to fill in aisummary for each chain 
            #print(f"Fieldname args found for: {jira_fields_default_value}")
            #yaml.dump({"field_args": jira_fields_default_value}, f, default_flow_style=False)

    else:
        print(f"ERROR: can't proceed, No JIRA IDs found to write to aisummary yaml file {execsummary_scope_output_file}")
        sys.exit(1)

    f.close()
    print("ExecSummary scope yaml file created successfully:", execsummary_scope_output_file)


# Build LLM context from table rows and corresponding jiracsv file
# so yes, it combines excel table rows (all cells including non-jira ones)
# and combines with the aisummary.jira.csv file contents
def build_llm_context(table_name, timestamp):
    
    context = ""
    
    '''# first put in the sheet contents for this table
    for r in table_rows:
        # r is a list of cell values, join them into a string
        context += " | ".join(str(cell) for cell in r) + "\n"
    '''

     # now append the jiracsv contents
    jiracsv_pattern = f"{basename}.{sheet}.{table_name}.{timestamp}.aisummary.jira.csv"
    dir_path = os.getcwd()  # or the folder where your CSVs live

    # case-insensitive search
    matched_file = None
    for f in os.listdir(dir_path):
        if f.lower() == jiracsv_pattern.lower():
            matched_file = os.path.join(dir_path, f)
            break

    if matched_file:
        with open(matched_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            context += line
    else:
        print(f"Warning: file not found (case-insensitive match) for {jiracsv_pattern}")

    return context

# Default to localhost unless overridden in env variable (set when in Docker)
SUMMARIZER_HOST = os.getenv("SUMMARIZER_HOST", "http://localhost:8000")
#LLMCONFIG_FILE = "../../../config/llmconfig.json"


def get_llm_model(llm_config_file):
    print("Current working directory:", os.getcwd())  # <-- debug
    if os.path.exists(llm_config_file):
        with open(llm_config_file, "r") as f:
            llm_model_set = json.load(f)
            m = llm_model_set.get("model")
            print(f"get_llm_model returning  {m}")
            return m
    else:
        print(f"ERROR: load_llm_config file {llm_config_file} was not found")
    
    return None


# variable names refer to comments but they don't have to be.  This function used for any field that were found to have an LLM prompt 
def get_summarized_comments(comments_list_asc, field_arg=None):
    """
    Summarize comments for LLM processing.
    This function takes a list of comments in ascending order and returns a summarized version.
    Optionally, a specific field name can be passed via `field_arg`.
    """
    try:
        if not comments_list_asc:
            return "No comments available."

        # Only join if it's a list or tuple
        if isinstance(comments_list_asc, (list, tuple)):
            comments_str = "; ".join(comments_list_asc)
        else:
            comments_str = comments_list_asc  # already a string

        comments_str = comments_str.replace("\n", "").replace("\r", "")

        # Prepare the payload for the service
        payload = {
            "comments": [comments_str]  # service expects a list[str]
        }
        if field_arg:
            payload["field"] = field_arg  # include field if provided

        LLMCONFIG_FILE = f"../../../config/llmconfig_{userlogin}.json"
        llm_model = get_llm_model(LLMCONFIG_FILE)
         # Determine endpoint
        if llm_model == "OpenAI":
            ENDPOINT = "/summarize_openai_ex"
        elif llm_model == "Local":
            ENDPOINT = "/summarize_local_ex"
        elif llm_model == "Claude":
            ENDPOINT = "/summarize_claude_ex"
        else:
            ENDPOINT = "/summarize_local_ex"

        # Make the POST request
        resp = requests.post(f"{SUMMARIZER_HOST}{ENDPOINT}", json=payload)

        if resp.status_code == 200:
            full_response = resp.json().get("summary", "")
            print(f"LLM endpoint returned {resp.status_code} OK")
        else:
            print(f"error LLM endpoint returned {resp.status_code}")
            full_response = f"[ERROR] Service call failed: {resp.text}"

        # Clean up response
        full_response = full_response.rstrip("\n").replace("\n", "; ").replace("|", "/")

        print(f"Full response: {full_response}")

        return full_response
    
    except Exception as e:
        # Log the exception and return a safe default
        print(f"[EXCEPTION ERROR] LLM could not be engaged, get_summarized_comments failed: {e}")
        #return "[ERROR] LLM could not be engaged due to exceptions during LLM interaction."
        return f"[EXCEPTION ERROR] LLM could not be engaged, get_summarized_comments failed: {e}"




def html_to_text_with_structure(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    def walk(node, depth=0):
        parts = []
        for child in node.children:
            if child.name is None:  # NavigableString
                text = child.strip()
                if text:
                    # add indentation spaces based on depth
                    parts.append(" " * (depth * 2) + text)
            else:
                # Handle block-level tags with newlines
                if child.name in ("p", "div", "section", "article", "header", "footer",
                                  "ul", "ol", "li", "br", "h1", "h2", "h3", "h4", "h5", "h6"):
                    if child.name == "li":
                        parts.append(" " * (depth * 2) + "- " + walk(child, depth + 1).strip())
                    else:
                        inner = walk(child, depth + 1)
                        if inner:
                            parts.append(inner)
                    parts.append("\n")  # newline after block
                else:
                    # Inline tag (span, b, i, etc.)
                    inner = walk(child, depth)
                    if inner:
                        parts.append(inner)
        return "\n".join(p for p in parts if p.strip())

    text = walk(soup).strip()

    # Collapse excessive blank lines
    lines = [line.rstrip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line.strip() != "")





if len(sys.argv) < 4:
    print("Usage: python runrate_resolved.py <yaml_file> <timestamp> <userlogin>")
    sys.exit(1)

yaml_file = sys.argv[1]
#filename = sys.argv[2]
timestamp = sys.argv[2]
userlogin = sys.argv[3]

with open(yaml_file, 'r') as f:
    data = yaml.safe_load(f)

fileinfo = data.get('fileinfo', {})
if not fileinfo:
    print("No fileinfo found in the YAML file.")
    sys.exit(1)
basename = fileinfo.get('basename')
tablename = fileinfo.get('table').replace(" ", "_") if fileinfo.get('table') else ""
source = fileinfo.get('source')
scope_file = fileinfo.get('scope file')
sheet = fileinfo.get('sheet')

docs_list = data.get('docs', [])
print(f"docs_list read from yaml file: {docs_list}")

if not basename:
    print("No 'basename'found in fileinfo. Expecting 'basename' key.")
    sys.exit(1)

if not tablename:
    print("No 'table' found in fileinfo. Expecting 'table' key.")
    sys.exit(1)

mode = ""


import json

# remove below because fileinfo is already ready above from yaml file.  ALSO WORSE fileinfo.json doesn't have table which is needed later.
# future - add table to fileinfo.json if needed.
'''with open("fileinfo.json", "r") as f:
    fileinfo = json.load(f)

print("read fileinfo.json file:", fileinfo)
'''

# I don't think this code is doing anything useful. It always sets create_mode to True?! since 
# Determine if we will be INSERTING rows eventually vs just UPDATING existing rows in Excel/SharePoint
if "assignee.rate" in yaml_file.lower():
    print(f"assignee rate mode detected based on filename: {yaml_file}.")
    # You can set a flag or handle import-specific logic here if needed
    output_file = basename + "." + tablename + timestamp + ".assignee.rate.jira.csv"
    #mode = "assignee" #"resolved"
else:
    print("Error: YAML filename does not indicate 'assignee.rate' mode.")
    sys.exit(1)

fields = data.get('fields', [])
# Convert list of dicts into a dictionary
fields_dict = {field["value"]: field.get("index", "<blank>") for field in fields}

field_values = [field.get('value') for field in fields if 'value' in field]
field_indexes = [field.get('index') for field in fields if 'index' in field]
field_values_str = ','.join(field_values)
field_indexes_str = ','.join(map(str, field_indexes))

print("Source file,", source)
print("basename,", basename)
print("Table,", tablename)
print("Field indexes,", field_indexes_str)
print("Field values,", field_values_str)



'''with open(output_file, "w") as outfile:
    outfile.write("Source file," + source + "\n")
    outfile.write("Basename," + basename + "\n")
    outfile.write("Scope file," + scope_file + "\n")    
    outfile.write("Table," + tablename + "\n")
    outfile.write("Field indexes," + field_indexes_str + "\n")
    outfile.write("Field values," + field_values_str + "\n")
'''

# yaml will expected to contain params, eg:
'''
    fileinfo:
    basename: children test.xlsx
    scope file: children test.xlsx.Resolved_Run_rate.20250921_183156.rate.yaml
    source: children test.xlsx
    table: Resolved_Run_rate
    params:
    - weeks 6
    - JQL project = tes and assignee = nadeem
'''

# the following were addded by scope.py in runrate yaml file
runrate_params_list = data.get('params',[])
runrate_table_row = data.get('row', None)
runrate_table_col = data.get('col', None)
#scan_ahead_nonblank_rows = data.get('scan_ahead_nonblank_rows', 0)
scan_ahead_nonblank_rows = data.get('last_update_row_count',-2) 
print(f"scan_ahead_nonblank_rows initialized = {scan_ahead_nonblank_rows}")

llm_user_prompt = data.get('llm', "Read all of it and briefly as possible categorize types of issues and work that was done. Mention any reason you see that could have blocked work on these issues or could have been done more quickly or correctly." )
sysprompt = "The following text is a delimited data separated by | character. These are rows of jira issues. " + llm_user_prompt

print (f"llm prompt = {sysprompt}")



runrate_params = parse_runrate_params(runrate_params_list)
print("Runrate params:", runrate_params)

jql_query = runrate_params.get("jql", "")
if not jql_query:
    print("Error: No JQL query specified in params.")
    sys.exit(1)

jira_filter_str = jql_query.replace("jql", "").strip()
JIRA_MAX_RESULTS = 1000  # adjust as needed

weeks = runrate_params.get("weeks", None)  # default to 6 weeks if not specified
days = runrate_params.get("days", None)  # optional days param
months =  runrate_params.get("months", None)  # optional months param
years = runrate_params.get("years", None)  # optional years param

#if (not weeks) and (not days) and (not months) and (not years):
if runrate_params.get("mode", "") == "":
    print("ERROR: No valid time frame (weeks/days/months/years) specified in params. Defaulting to 6 weeks.")
    sys.exit(1)

# Replace with your Jira Cloud credentials and URL

# Load environment variables from a .env file if present
#load_dotenv()
ENV_PATH = f"../../../config/env.{userlogin}"
load_dotenv(dotenv_path=ENV_PATH)
JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN")
JIRA_URL = os.environ.get("JIRA_URL")
JIRA_EMAIL = os.environ.get("JIRA_EMAIL")

if not JIRA_API_TOKEN:
    print("Error: JIRA_API_TOKEN environment variable not set.")
    sys.exit(1)



#########################


# first check if the assignee.jira.csv already exists, which would mean this is 2nd call to runrate_assignee.py

context = ""
csv = os.path.join(os.getcwd(), f"{basename}.{sheet}.{tablename}.*.{timestamp}.assignee.jira.csv")
csv_files = glob.glob(csv)

if csv_files:
    # yes, we have assignee.jira.csv files so this must be 2nd pass of cycletime from resync

    for csv_file in csv_files:
        context = ""
        print(f"found assignee.jira.csv file {csv_file}")
        with open(csv_file, "r", encoding="utf-8") as f:
            for _ in range(5):  # skip first 5 lines
                next(f, None)
            for line in f:
                context += line 


        match = re.match(rf"{re.escape(basename)}\.{sheet}\.{tablename}\.(.+?)\.assignee\.jira\.csv", os.path.basename(csv_file))
        if match:
            substring = match.group(1)
            output_file = f"{basename}.{sheet}.{tablename}.{substring}.assignee.llm.txt"
            context_output_file =  f"{basename}.{sheet}.{tablename}.{substring}.assignee.context.txt"
            chain_yaml_file =  f"{basename}.{sheet}.{tablename}.{substring}.assignee.scope.yaml"
            print(f"context for {csv_file} will be saved in {output_file}")
            print(f"assignee row for {csv_file} will be looked in {chain_yaml_file}")
        else:
            print(f"ERROR: could not find assignee.jira.csv pattern in {csv_file} to generate context for LLM. Will exit runrate_assignee.py")
            sys.exit(0)


        # read the chain_row that was saved in this chain's scope.yaml file earlier
        with open(chain_yaml_file, 'r') as f:
            data = yaml.safe_load(f)
            rownum = data.get('chain_row', None)
            if rownum:
                print(f"chain row is {rownum}")
            else:
                print(f"No chain_row found in the YAML file {chain_yaml_file}")
                sys.exit(1)
        
        print(f"Preparing to call LLM with llm_user_prompt={llm_user_prompt} for userlogin={userlogin} at timestamp={timestamp}")
        from vector_rag_retriever import *
        rag_result = search_and_prepare_for_llm(userlogin, llm_user_prompt, docs_list)
        if rag_result and rag_result.get('has_context'):
            rag_context = rag_result.get('context', '')
            print(f"RAG context for llm_user_prompt {llm_user_prompt}: {rag_context[:500]}{'...' if len(rag_context) > 500 else ''}")
            context = f"Context: {rag_context}\n\n{context}"
        else:
            print(f"No RAG context for llm_user_prompt {llm_user_prompt}. Proceeding without context.")
        
        print(f"Calling get_summarized_comments with context={context[:255]}... and sysprompt={sysprompt[:255]}...")
        report = get_summarized_comments(context, sysprompt)

        # Save context to file
        with open(context_output_file, "w", encoding="utf-8") as f:
            f.write(context)
            print(f"LLM context saved to {context_output_file}")

        # save llm's response to file for debugging
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(report)
            print(f"LLM reponse saved to {output_file}")

        # Replace all newlines with semicolons
        cleaned_response = html_to_text_with_structure(report)
        #print(f"html_to_text_with_structure returned cleaned_response={cleaned_response}")
        cleaned_response = cleaned_response.rstrip("\n")
        cleaned_response = cleaned_response.replace("\n", "; ")
        cleaned_response = cleaned_response.replace("|", "^")
        
        # now update the changes.txt with llm response
        # the chain's changes.txt file will be automaically consumed by update_googlesheet (or update_sharepoint) downstream
        # +1 because cols start at 1 = A
        # +8 because it needs to be displayed on right end of cycletime table on sheet 
        entry = f"{get_column_letter(runrate_table_col + 1 + 6)}{rownum} = {cleaned_response} || "  
    
        changes_file = yaml_file.replace("scope.yaml","import.changes.txt")
        print(f"Will write LLM response into {changes_file}")

        with open(changes_file, "a") as f:
            f.write(entry + "\n")
            print(entry)
       
        print(f"Changes written to {changes_file} entry = {entry}")

    
    # no need to proceed since cycle time has completed 2 passess at this point
    sys.exit(0)    
else:
    print (f"No chain.jira.csv files found matching pattern {csv} so assume it's first pass of cycletime.py")





# Connect to Jira with basic auth
try:
    jira = JIRA(server=JIRA_URL, basic_auth=(JIRA_EMAIL, JIRA_API_TOKEN))
    #print("✅ Successfully connected to Jira.")
except Exception as e:
    print(f"❌ Failed to connect to Jira: {e}")
    sys.exit(1)


try:
    issues = jira.search_issues(jira_filter_str, maxResults=False, expand='changelog')
    print(f"✅ Found {len(issues)} issue(s) matching the filter={jira_filter_str}")
    for i, issue in enumerate(issues, start=1):
        print(f"{i}. {issue.key} — {issue.fields.summary}")
except Exception as e:
    print(f"❌ Failed to search issues: {e}")




# Now we need to write out the changes to a text file that can be picked up by update_sharepoint.py
changes_list = []

row = runrate_table_row + 1  # skip to next row after <> tag cell
col = runrate_table_col # integer! and will need to be convered to letter for openpyxl

from openpyxl.utils import get_column_letter

print(f"\n🔄 Bucketizing {len(issues)} issues by calendar weeks...")

assignee_list = []

# need list of all unique assignee since they will be in the excel cells table
for issue in issues:
    #assignee  = issue.fields.assignee.displayName if issue.fields.assignee else "Unassigned"   
    assignee = get_resolved_by_user(issue)
    # we don't want to show Unknown assignee in the sheet
    if (assignee != "Unknown"):
        assignee_list.append(assignee)

unique_assignee_list = list(set(assignee_list))  # unique values

unique_assignee_list = sorted(set(assignee_list))  # unique values, sorted

print(f"Found {len(unique_assignee_list)} unique assignees: {unique_assignee_list}")



# FUTURE: delete rows to remove data from previous resync
# if scan_ahead_nonblank_rows > 0:
print(f"scan_ahead_nonblank_rows remaining = {scan_ahead_nonblank_rows}")
temp_row = row
while (scan_ahead_nonblank_rows + 1 >= 0):
    print("filling in remaining rows with DELETE")
    changes_list.append(f"{get_column_letter(runrate_table_col + 1)}{temp_row + scan_ahead_nonblank_rows + 1 } = DELETE || ")
    scan_ahead_nonblank_rows -= 1
    #temp_row +=1


# write out the timestamp in cell adjacent to <> so we can tell when the update occured
coord = f"{get_column_letter(col + 1)}{row}"
now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
entry = f"{coord} = {len(unique_assignee_list)} rows updated on {now_str} by Trinket ||"
print (entry)
changes_list.append(entry)
row += 1

# write out the headers first
coord = f"{get_column_letter(col + 1)}{row}"

if scan_ahead_nonblank_rows:
    prefix = ""
    scan_ahead_nonblank_rows -= 1
else:
    prefix = "INSERT"

entry = f"{coord} = {prefix} Resolved By ||"
print (entry)
changes_list.append(entry)


# write out all the assignees in the header column
for index, assignee in enumerate(unique_assignee_list):
    print(f"{index}: Processing assignee {assignee}")    
    coord = f"{get_column_letter(col + 1)}{row + 1 + index}"

    if scan_ahead_nonblank_rows >= 0:
        prefix = ""
        scan_ahead_nonblank_rows -= 1
    else:
        prefix = "INSERT"

    entry = f"{coord} = {prefix} {assignee} ||"
    print (entry)
    changes_list.append(entry)



if issues:
    # Bucketize issues by weeks
    #weekly_buckets_resolved, week_info_resolved = bucketize_issues_by_interval(issues,"open", runrate_params.get("mode", "weeks"),assignee_filter=assignee)
    weekly_buckets_resolved, week_info_resolved = bucketize_issues_by_interval(issues,"resolved",runrate_params.get("mode", "weeks"))
    #weekly_buckets_closed, week_info_closed = bucketize_issues_by_weeks(issues,"closed")
    
    print("Here's the weekly breakdown:")
    if weekly_buckets_resolved:
        # Print summary
        print_weekly_summary(weekly_buckets_resolved, week_info_resolved)

    '''print("Here's the resolved weekly breakdown:")
    if weekly_buckets_resolved:
        # Print summary
        print_weekly_summary(weekly_buckets_resolved, week_info_resolved)
    '''

# need to keep track of which week maps to which column so that both OPEN and CLOSE rates are under same colums per week
week_to_col = {}

from collections import OrderedDict

interval = runrate_params.get("mode", "weeks")  

week_to_col = {}

# Populate week_to_col
if interval == "years":
    for year, dud, start_date, end_date in week_info_resolved:
        week_to_col[year] = f"{year}"

elif interval == "months":
    for year, month_num, start_date, end_date in week_info_resolved:
        # ensure two-digit month formatting
        month_str = f"{month_num:02d}"
        week_to_col[f"{year}-{month_str}"] = f"{year}-{month_str}"

elif interval == "weeks":
    for year, week_num, start_date, end_date in week_info_resolved:
        # BUG! BUG!  week_num needs to be crafted like "months" - it's multipart not single number
        week_to_col[f"{str(start_date.date())}"] = f"{str(start_date.date())}"

else:
    print(f"Error: unrecognized period interval found in runrate_params mode={interval}")
    sys.exit(1)

# Fill in any gaps in the sequence (only makes sense for numeric keys)
# works for years only whenere key is numeric.   weeks did not work because keys are strings
if interval in ("years") and week_to_col:
    min_week = min(week_to_col.keys())
    max_week = max(week_to_col.keys())
    print(f"Week range: {min_week} to {max_week}")
    
    for week_num in range(min_week, max_week + 1):
        if week_num not in week_to_col:
            week_to_col[week_num] = week_num
            print(f"  Added missing week {week_num}")

# Sort week_to_col properly
if interval == "months":
    # Sort "YYYY-MM" strings as (year, month) tuples
    def parse_year_month(key: str):
        y, m = key.split("-")
        return int(y), int(m)

    sorted_keys = sorted(week_to_col.keys(), key=parse_year_month)
else:
    # numeric sort for weeks or years
    sorted_keys = sorted(week_to_col.keys())

# Create ordered dict
sorted_week_to_col = OrderedDict((k, week_to_col[k]) for k in sorted_keys)

print(f"Sorted keys: {list(sorted_week_to_col.keys())}")


week_to_label = {}  

week_to_col['total'] = col + 2  # save the col to write assignee_total
coord = f"{get_column_letter(week_to_col['total'])}{row}"
entry = f"{coord} = Total Jira || " 
print (entry)
changes_list.append(entry)


# Now assign column numbers in order
for col_index, week_num in enumerate(sorted_week_to_col):
    print(f"week_to_col({week_num}) -> Column {col + 1 + col_index + 1}")
    week_to_label[week_num] = week_to_col[week_num]
    week_to_col[week_num] = col + 1 + col_index + 2     # + 2 because we added assignee_total col as well

# now write out the week headers
for i in sorted_week_to_col:
    label = week_to_label.get(i)
    coord = f"{get_column_letter(week_to_col.get(i))}{row}"
    #entry = f"{coord} = {runrate_params['mode']} {i} || "
    entry = f"{coord} = {label} || "
    print(entry)
    changes_list.append(entry)


# now loop through each assignee and write out their counts per week
for assignee in unique_assignee_list:
    weekly_buckets_resolved, week_info_resolved = bucketize_issues_by_interval(issues,"resolved",runrate_params.get("mode", "weeks"), assignee_filter=assignee)
    #weekly_buckets_closedrun, week_info_closed = bucketize_issues_by_weeks(issues,"closed")

    print(f"Here's the weekly breakdown for assignee={assignee}:")

    if weekly_buckets_resolved:
        # Print summary
        print_weekly_summary(weekly_buckets_resolved, week_info_resolved)

    assignee_total = 0
    print ("assignee_total reset to 0")
    # now loop through the weeks and write out the counts
    jql_ids = []  # to collect jira ids for this assignee across all weeks
    for i, (bucket, (year, week_num, start_date, end_date)) in enumerate(zip(weekly_buckets_resolved, week_info_resolved)):
        #week_str = f"Week {i+1} ({year}-W{week_num:02d})"
        week_str = f"{start_date.strftime('%Y-%m-%d')}"
        date_range = f"{start_date.strftime('%m/%d')} - {end_date.strftime('%m/%d/%Y')}"
        issue_count = len(bucket)

       
        #week_to_col[week_num] = col + 1 + i + 1
        #print(f"saving week_to_col('{week_num}' = {col + 1 + i + 1})")
        
        #coord = f"{get_column_letter(col + 1 + i + 1)}{row + 1}"
        #entry = f"{coord} = {len(bucket)} || "
        
        print(f"bucket len = {len(bucket)}")
        if (len(bucket) == 0):
            print (f"Skipping week {week_num} since no OPEN issues")
            continue

        # Populate week_to_col
        if interval == "years":
#            for year, dud, start_date, end_date in week_info_resolved:
            week_num = year

        elif interval == "months":
 #           for year, month_num, start_date, end_date in week_info_resolved:
                # ensure two-digit month formatting
                month_str = f"{week_num:02d}"
                week_num = f"{year}-{month_str}"

        elif interval == "weeks":
#            for year, week_num1, start_date, end_date in week_info_resolved:
                # BUG! BUG!  week_num needs to be crafted like "months" - it's multipart not single number
                week_num= f"{str(start_date.date())}"    # start date of the week

        else:
            print(f"Error: unrecognized period interval found in runrate_params mode={interval}")
            sys.exit(1)

        #print (f"Looking up week {week_num} in sorted_week_to_col, found column {week_to_col.get(week_num,'None')}")
        #coord = f"{get_column_letter(week_to_col.get(week_num))}{row + 1}"

        print(f"week_num={week_num!r} ({type(week_num)}), "
        f"keys={[ (k, type(k)) for k in week_to_col.keys() ]}")

        print(f"checking if week_num={week_num} exists week_to_col keys={week_to_col}")
        if week_num in week_to_col:
            print(f"Looking up week {week_num} in sorted_week_to_col, found column {week_to_col[week_num]}")
            coord = f"{get_column_letter(week_to_col[week_num])}{row + 1}"
        else:
            print(f"WARNING: Week {week_num} not found in sorted_week_to_col. Possibly because number of intervals exceeded hardcoded max")
            continue
            #coord = None  # or handle it however you want


        jql = "key in ("
        for issue in bucket:
            jql += issue.key + ","
            jql_ids.append(issue.key)   

        jql = jql.rstrip(",") + ") order by key asc"

        #entry = f"{coord} = {len(bucket)} || "
        hyperlink = _make_hyperlink_formula(f"{JIRA_URL}/issues/?jql={jql}", f"{len(bucket)}") + " || "
        entry = f"{coord} = {hyperlink} || " 
        print (entry)
        changes_list.append(entry)

        assignee_total += len(bucket)
        print(f"assignee_total updated to {assignee_total} for assignee={assignee}")

    # write out the  assignee.scope.yaml file for this assignee 
    write_execsummary_yaml(jql_ids,fileinfo, assignee.replace(' ','_'), row + 1, timestamp)

    # now write out the assignee_total
    coord = f"{get_column_letter(week_to_col['total'])}{row + 1}"
    entry = f"{coord} = {assignee_total} || " 
    print (entry)
    changes_list.append(entry)

    
    row = row + 1   # move to next row for next assignee    



changes_file = yaml_file.replace("scope.yaml","import.changes.txt")
print(f"Writing changes to {changes_file}")

if changes_list:
    with open(changes_file, "w") as f:
        #f.write("sheet = ", sheet)  #save sheet name for update_sharepoint downstream
        for entry in changes_list:
            if "||None" in entry:
                entry = entry.replace("||None", "||")
            f.write(entry + "\n")
            print(entry)
    print(f"Changes written to {changes_file} with ({len(changes_list)} entries).")
else:
    print(f"No changes to write. Not need to create {changes_file}")
